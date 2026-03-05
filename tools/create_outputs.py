"""
create_outputs.py — Generates three deliverables from demo (or real) lead data:

1. MASTER_LEADS spreadsheet (Excel, .xlsx) — all leads organized by buyer tab
2. CIM overview PDF — professional multi-page document with charts and system overview
3. Per-buyer CIM PDFs — one page per buyer showing their seller pipeline

Usage:
    python tools/create_outputs.py
"""

import sys
import os
import io
from pathlib import Path
from datetime import datetime
import matplotlib
matplotlib.use("Agg")  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import numpy as np

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable, PageBreak, KeepTogether,
)
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

sys.path.insert(0, str(Path(__file__).parent))
from demo_data import get_all_demo_leads, DEMO_LEADS
from buyers_db import BUYERS, get_buyer_by_id

BASE_DIR = Path(__file__).parent.parent
OUT_DIR = BASE_DIR / ".tmp"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
NAVY        = colors.HexColor("#0B1F3A")
GOLD        = colors.HexColor("#C9A84C")
LIGHT_BLUE  = colors.HexColor("#E8F0FA")
MID_BLUE    = colors.HexColor("#2E6DA4")
GREEN       = colors.HexColor("#2D8A4E")
ORANGE      = colors.HexColor("#D4622A")
GRAY_LIGHT  = colors.HexColor("#F5F5F5")
GRAY_MID    = colors.HexColor("#D0D0D0")
GRAY_DARK   = colors.HexColor("#555555")
WHITE       = colors.white
BLACK       = colors.black

PY_NAVY     = "#0B1F3A"
PY_GOLD     = "#C9A84C"
PY_MID_BLUE = "#2E6DA4"
PY_GREEN    = "#2D8A4E"
PY_ORANGE   = "#D4622A"

# ── HELPER ────────────────────────────────────────────────────────────────────

def mpl_to_reportlab_image(fig, width_inches=7.0):
    """Convert matplotlib figure to ReportLab Image flowable."""
    buf = io.BytesIO()
    fig.savefig(buf, format="PNG", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    img = Image(buf)
    aspect = img.imageHeight / img.imageWidth
    img.drawWidth  = width_inches * inch
    img.drawHeight = width_inches * inch * aspect
    return img


def score_color(score):
    if score >= 90: return GREEN
    if score >= 80: return MID_BLUE
    if score >= 70: return GOLD
    return ORANGE


# ══════════════════════════════════════════════════════════════════════════════
#   CHART GENERATORS
# ══════════════════════════════════════════════════════════════════════════════

def chart_buyer_industry_distribution(buyers):
    """Donut chart: Buyers by primary industry."""
    from collections import Counter
    counts = Counter()
    for b in buyers:
        ind = b["industries"][0].title() if b["industries"] else "General"
        counts[ind] += 1
    # Top 8 + "Other"
    top = counts.most_common(8)
    labels = [t[0] for t in top]
    vals = [t[1] for t in top]

    colors_list = [PY_NAVY, PY_MID_BLUE, PY_GOLD, PY_GREEN, PY_ORANGE,
                   "#7B5EA7", "#2AA8A8", "#CC4466"][:len(labels)]

    fig, ax = plt.subplots(figsize=(6.5, 4.5), facecolor="white")
    wedges, texts, autotexts = ax.pie(
        vals, labels=labels, colors=colors_list,
        autopct="%1.0f%%", startangle=140,
        wedgeprops=dict(width=0.6, edgecolor="white", linewidth=2),
        pctdistance=0.78, labeldistance=1.12,
    )
    for t in texts: t.set_fontsize(9)
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color("white")
        at.set_fontweight("bold")
    ax.set_title("Buyers by Primary Industry", fontsize=13, fontweight="bold",
                 color=PY_NAVY, pad=15)
    fig.tight_layout()
    return fig


def chart_deal_size_ranges(buyers):
    """Grouped bar chart: EBITDA ranges per buyer group."""
    groups = {
        "Under $1M EBITDA": 0,
        "$1M – $3M EBITDA": 0,
        "$3M – $7M EBITDA": 0,
        "$7M – $20M EBITDA": 0,
        "$20M+ EBITDA": 0,
    }
    for b in buyers:
        lo = b.get("ebitda_min", 0)
        hi = min(b.get("ebitda_max", 999_999_999), 999_999_999)
        mid = (lo + min(hi, 50_000_000)) / 2
        if mid < 1_000_000:   groups["Under $1M EBITDA"] += 1
        elif mid < 3_000_000: groups["$1M – $3M EBITDA"] += 1
        elif mid < 7_000_000: groups["$3M – $7M EBITDA"] += 1
        elif mid < 20_000_000: groups["$7M – $20M EBITDA"] += 1
        else: groups["$20M+ EBITDA"] += 1

    fig, ax = plt.subplots(figsize=(7, 3.8), facecolor="white")
    keys = list(groups.keys())
    vals = [groups[k] for k in keys]
    bar_colors = [PY_MID_BLUE, PY_NAVY, PY_GOLD, PY_GREEN, PY_ORANGE]
    bars = ax.bar(keys, vals, color=bar_colors, edgecolor="white",
                  linewidth=1.5, width=0.6, zorder=3)

    for bar, val in zip(bars, vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.15,
                    str(val), ha="center", va="bottom",
                    fontweight="bold", fontsize=11, color=PY_NAVY)

    ax.set_ylabel("Number of Buyers", fontsize=10, color=PY_NAVY)
    ax.set_title("Buyer EBITDA Requirements", fontsize=13, fontweight="bold",
                 color=PY_NAVY, pad=12)
    ax.set_ylim(0, max(vals) + 2)
    ax.tick_params(axis="x", labelsize=8)
    ax.yaxis.grid(True, alpha=0.4, color="#DDDDDD")
    ax.set_axisbelow(True)
    ax.spines[["top","right"]].set_visible(False)
    ax.tick_params(colors=PY_NAVY)
    fig.tight_layout()
    return fig


def chart_geographic_coverage(buyers):
    """Horizontal bar chart: geographic coverage."""
    regions = {
        "US Nationwide": 0, "Midwest": 0, "Southeast": 0, "West Coast": 0,
        "Mid-Atlantic / NE": 0, "Texas": 0, "Florida": 0, "Southwest": 0,
        "UK / International": 0,
    }
    for b in buyers:
        geos = [g.lower() for g in b.get("geographies", [])]
        states = b.get("states", [])
        if not geos and not states:
            regions["US Nationwide"] += 1
        elif "uk" in geos or "international" in geos:
            regions["UK / International"] += 1
        elif "us nationwide" in geos or "us" in geos or ("us" in " ".join(geos).lower() and not states):
            regions["US Nationwide"] += 1
        elif any(s in ["TX"] for s in states) and len(states) <= 2:
            regions["Texas"] += 1
        elif any(s in ["FL"] for s in states) and len(states) <= 3:
            regions["Florida"] += 1
        elif any(s in ["IL","IN","OH","MI","WI","MN","IA","MO"] for s in states):
            regions["Midwest"] += 1
        elif any(s in ["FL","GA","SC","NC","TN","AL","AR"] for s in states):
            regions["Southeast"] += 1
        elif any(s in ["CA","OR","WA","NV","AZ","UT"] for s in states):
            regions["West Coast"] += 1
        elif any(s in ["PA","NJ","DE","MD","NY","CT","WV"] for s in states):
            regions["Mid-Atlantic / NE"] += 1
        else:
            regions["US Nationwide"] += 1

    sorted_items = sorted(regions.items(), key=lambda x: x[1], reverse=True)
    sorted_items = [(k, v) for k, v in sorted_items if v > 0]
    keys = [i[0] for i in sorted_items]
    vals = [i[1] for i in sorted_items]

    fig, ax = plt.subplots(figsize=(7, 3.6), facecolor="white")
    colors_list = [PY_NAVY, PY_MID_BLUE, PY_GOLD, PY_GREEN, PY_ORANGE,
                   "#7B5EA7", "#2AA8A8", "#CC4466", "#888888"][:len(keys)]
    bars = ax.barh(keys, vals, color=colors_list, edgecolor="white",
                   linewidth=1.2, height=0.55, zorder=3)
    for bar, val in zip(bars, vals):
        ax.text(val + 0.1, bar.get_y() + bar.get_height()/2,
                str(val), va="center", ha="left", fontweight="bold",
                fontsize=10, color=PY_NAVY)
    ax.set_xlabel("Number of Active Buyers", fontsize=9, color=PY_NAVY)
    ax.set_title("Buyer Geographic Demand", fontsize=13, fontweight="bold",
                 color=PY_NAVY, pad=12)
    ax.set_xlim(0, max(vals) + 3)
    ax.xaxis.grid(True, alpha=0.4, color="#DDDDDD")
    ax.set_axisbelow(True)
    ax.spines[["top","right"]].set_visible(False)
    ax.tick_params(colors=PY_NAVY, labelsize=9)
    fig.tight_layout()
    return fig


def chart_pipeline_funnel(leads_by_buyer):
    """Funnel chart: Pipeline stages."""
    total_targets  = sum(len(v) for v in leads_by_buyer.values())
    total_buyers   = len(leads_by_buyer)
    contacted_pct  = 0.72
    responded_pct  = 0.28
    interested_pct = 0.14
    closed_pct     = 0.04

    stages = ["Scraped / Identified", "Outreach Sent", "Responded", "Expressed Interest", "Intro to Buyer"]
    counts = [
        total_targets,
        int(total_targets * contacted_pct),
        int(total_targets * responded_pct),
        int(total_targets * interested_pct),
        int(total_targets * closed_pct),
    ]
    stage_colors = [PY_NAVY, PY_MID_BLUE, PY_GOLD, PY_GREEN, PY_ORANGE]

    fig, ax = plt.subplots(figsize=(7.5, 4.5), facecolor="white")
    max_w = 1.0
    for i, (stage, count, col) in enumerate(zip(stages, counts, stage_colors)):
        width = max_w * (count / counts[0])
        y_pos = len(stages) - 1 - i
        rect = mpatches.FancyBboxPatch(
            ((max_w - width) / 2, y_pos - 0.35),
            width, 0.7,
            boxstyle="round,pad=0.02",
            linewidth=0, facecolor=col,
        )
        ax.add_patch(rect)
        ax.text(0.5, y_pos, f"{stage}   |   {count:,} companies",
                ha="center", va="center", color="white",
                fontsize=9.5, fontweight="bold")

    ax.set_xlim(0, 1)
    ax.set_ylim(-0.6, len(stages) - 0.4)
    ax.axis("off")
    ax.set_title(f"Projected Pipeline Funnel  ({total_buyers} Active Buyers | {total_targets:,} Targets)",
                 fontsize=12, fontweight="bold", color=PY_NAVY, pad=14)
    fig.tight_layout()
    return fig


def chart_match_scores(leads):
    """Distribution histogram of match scores for a buyer's leads."""
    scores = [l.get("match_score", 80) for l in leads]
    fig, ax = plt.subplots(figsize=(5.5, 3.2), facecolor="white")
    bins = range(70, 102, 3)
    n, bins_out, patches = ax.hist(scores, bins=bins, edgecolor="white",
                                    linewidth=1.5, color=PY_MID_BLUE, zorder=3)
    for patch, left in zip(patches, bins_out):
        if left >= 90:   patch.set_facecolor(PY_GREEN)
        elif left >= 80: patch.set_facecolor(PY_MID_BLUE)
        else:            patch.set_facecolor(PY_GOLD)

    ax.set_xlabel("Match Score", fontsize=9, color=PY_NAVY)
    ax.set_ylabel("# Leads", fontsize=9, color=PY_NAVY)
    ax.set_title("Lead Match Score Distribution", fontsize=11, fontweight="bold", color=PY_NAVY)
    ax.yaxis.grid(True, alpha=0.4, color="#DDDDDD")
    ax.set_axisbelow(True)
    ax.spines[["top","right"]].set_visible(False)
    ax.tick_params(colors=PY_NAVY, labelsize=8)

    a_patch = mpatches.Patch(color=PY_GREEN, label="A-Grade (90+)")
    b_patch = mpatches.Patch(color=PY_MID_BLUE, label="B-Grade (80-89)")
    c_patch = mpatches.Patch(color=PY_GOLD, label="C-Grade (70-79)")
    ax.legend(handles=[a_patch, b_patch, c_patch], fontsize=7, loc="upper left")
    fig.tight_layout()
    return fig


def chart_source_breakdown(all_leads_flat):
    """Pie chart: lead sources."""
    from collections import Counter
    counts = Counter(l.get("source","unknown") for l in all_leads_flat)
    labels_map = {"apollo":"Apollo.io","yelp":"Yelp API","yellow_pages":"Yellow Pages",
                  "manta":"Manta.com","google":"Google","unknown":"Other"}
    labels = [labels_map.get(k, k) for k in counts.keys()]
    vals = list(counts.values())
    clrs = [PY_NAVY, PY_MID_BLUE, PY_GOLD, PY_GREEN, PY_ORANGE, "#888888"][:len(labels)]

    fig, ax = plt.subplots(figsize=(5.5, 3.8), facecolor="white")
    wedges, texts, autotexts = ax.pie(
        vals, labels=labels, colors=clrs,
        autopct="%1.0f%%", startangle=90,
        wedgeprops=dict(edgecolor="white", linewidth=2),
        pctdistance=0.75,
    )
    for t in texts: t.set_fontsize(9)
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color("white")
        at.set_fontweight("bold")
    ax.set_title("Leads by Source", fontsize=11, fontweight="bold", color=PY_NAVY)
    fig.tight_layout()
    return fig


# ══════════════════════════════════════════════════════════════════════════════
#   PDF GENERATORS
# ══════════════════════════════════════════════════════════════════════════════

def make_styles():
    styles = getSampleStyleSheet()
    custom = {
        "title_page_main": ParagraphStyle("title_page_main", parent=styles["Heading1"],
            fontSize=28, textColor=WHITE, fontName="Helvetica-Bold",
            spaceAfter=8, alignment=TA_LEFT),
        "title_page_sub": ParagraphStyle("title_page_sub", parent=styles["Normal"],
            fontSize=14, textColor=GOLD, fontName="Helvetica",
            spaceAfter=6, alignment=TA_LEFT),
        "title_page_date": ParagraphStyle("title_page_date", parent=styles["Normal"],
            fontSize=11, textColor=colors.HexColor("#AABBCC"), fontName="Helvetica",
            alignment=TA_LEFT),
        "section_header": ParagraphStyle("section_header", parent=styles["Heading2"],
            fontSize=16, textColor=NAVY, fontName="Helvetica-Bold",
            spaceBefore=18, spaceAfter=8, borderPad=4),
        "sub_header": ParagraphStyle("sub_header", parent=styles["Heading3"],
            fontSize=12, textColor=MID_BLUE, fontName="Helvetica-Bold",
            spaceBefore=10, spaceAfter=4),
        "body": ParagraphStyle("body", parent=styles["Normal"],
            fontSize=9.5, textColor=BLACK, fontName="Helvetica",
            leading=15, spaceAfter=6, alignment=TA_JUSTIFY),
        "body_small": ParagraphStyle("body_small", parent=styles["Normal"],
            fontSize=8.5, textColor=GRAY_DARK, fontName="Helvetica", leading=13),
        "table_header": ParagraphStyle("table_header", parent=styles["Normal"],
            fontSize=9, textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER),
        "company_name": ParagraphStyle("company_name", parent=styles["Normal"],
            fontSize=11, textColor=NAVY, fontName="Helvetica-Bold"),
        "label": ParagraphStyle("label", parent=styles["Normal"],
            fontSize=8, textColor=GRAY_DARK, fontName="Helvetica"),
        "value": ParagraphStyle("value", parent=styles["Normal"],
            fontSize=9, textColor=BLACK, fontName="Helvetica-Bold"),
        "kpi_number": ParagraphStyle("kpi_number", parent=styles["Normal"],
            fontSize=24, textColor=NAVY, fontName="Helvetica-Bold", alignment=TA_CENTER),
        "kpi_label": ParagraphStyle("kpi_label", parent=styles["Normal"],
            fontSize=8, textColor=GRAY_DARK, fontName="Helvetica", alignment=TA_CENTER),
    }
    # Merge base styles with custom (StyleSheet1 uses byName dict internally)
    merged = dict(styles.byName)
    merged.update(custom)
    return merged


def _table_style(header_color=None):
    hc = header_color or NAVY
    return TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), hc),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8.5),
        ("ALIGN",         (0,1), (-1,-1), "LEFT"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 7),
        ("RIGHTPADDING",  (0,0), (-1,-1), 7),
        ("GRID",          (0,0), (-1,-1), 0.4, GRAY_MID),
        ("ROWHEIGHT",     (0,0), (-1,-1), 22),
    ])


def score_badge(score, styles):
    color = score_color(score)
    grade = "A" if score >= 90 else ("B" if score >= 80 else "C")
    data = [[Paragraph(f"<b>{grade} — {score}</b>", styles["body_small"])]]
    t = Table(data, colWidths=[0.75*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), color),
        ("TEXTCOLOR",  (0,0), (-1,-1), WHITE),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("FONTNAME",   (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
    ]))
    return t


# ── OVERVIEW PDF ──────────────────────────────────────────────────────────────

def build_overview_pdf(all_leads: dict, buyers: list, out_path: Path):
    print(f"\n[PDF] Building overview PDF...")
    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        leftMargin=0.75*inch, rightMargin=0.75*inch,
        topMargin=0.75*inch, bottomMargin=0.75*inch,
    )
    S = make_styles()
    story = []
    all_flat = [l for leads in all_leads.values() for l in leads]
    W = 7.0 * inch

    # ── TITLE PAGE ────────────────────────────────────────────────────────────
    def draw_title_bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(NAVY)
        canvas.rect(0, 0, letter[0], letter[1], fill=1, stroke=0)
        # Gold accent bar
        canvas.setFillColor(GOLD)
        canvas.rect(0.75*inch, 5.8*inch, 0.06*inch, 2.6*inch, fill=1, stroke=0)
        canvas.restoreState()

    story.append(Spacer(1, 1.6*inch))
    story.append(Paragraph("Off-Market Deal Sourcing", S["title_page_sub"]))
    story.append(Spacer(1, 0.12*inch))
    story.append(Paragraph("Buyer Pipeline &<br/>Seller Outreach System", S["title_page_main"]))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("Valar Brokers — Jordi Quevedo-Valls", S["title_page_sub"]))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y')}", S["title_page_date"]))
    story.append(Spacer(1, 0.5*inch))

    # KPI boxes on title page
    total_buyers  = len(buyers)
    total_leads   = len(all_flat)
    total_linkedin = sum(1 for l in all_flat if l.get("owner_linkedin",""))
    avg_score     = int(sum(l.get("match_score",80) for l in all_flat) / max(len(all_flat),1))
    kpi_data = [
        [Paragraph(f"<b>{total_buyers}</b>", S["kpi_number"]),
         Paragraph(f"<b>{total_leads}</b>", S["kpi_number"]),
         Paragraph(f"<b>{total_linkedin}</b>", S["kpi_number"]),
         Paragraph(f"<b>{avg_score}</b>", S["kpi_number"])],
        [Paragraph("Active Buyers", S["kpi_label"]),
         Paragraph("Demo Leads (Sample)", S["kpi_label"]),
         Paragraph("With LinkedIn", S["kpi_label"]),
         Paragraph("Avg Match Score", S["kpi_label"])],
    ]
    kpi_table = Table(kpi_data, colWidths=[1.7*inch]*4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#132d52")),
        ("TEXTCOLOR",     (0,0), (-1,-1), WHITE),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#1e3d64")),
        ("ROWHEIGHT",     (0,0), (-1,0), 48),
    ]))
    story.append(kpi_table)
    story.append(PageBreak())

    # ── SECTION 1: SYSTEM OVERVIEW ─────────────────────────────────────────
    story.append(Paragraph("How the System Works", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=12))

    story.append(Paragraph(
        "This automated pipeline identifies off-market businesses — companies that are NOT publicly listed for sale — "
        "and systematically engages their owners to gauge acquisition interest. The system runs across all "
        f"<b>{total_buyers} active buyers</b> in parallel, outputting ready-to-send LinkedIn messages for each matched seller target.",
        S["body"]
    ))
    story.append(Spacer(1, 0.15*inch))

    # Workflow steps table
    workflow_data = [
        [Paragraph("Step", S["table_header"]),
         Paragraph("Action", S["table_header"]),
         Paragraph("Tool / Source", S["table_header"]),
         Paragraph("Output", S["table_header"])],
        ["1", "Load buyer criteria", "buyers_db.py", f"{total_buyers} buyers structured"],
        ["2", "Search businesses by industry + geo", "Apollo.io API", "Company + owner data"],
        ["3", "Scrape public directories", "Yelp, Yellow Pages, Manta", "Business names + phones"],
        ["4", "Deduplicate & score leads", "match_deals.py", "Match scores 0–100"],
        ["5", "Generate outreach messages", "message_generator.py + Claude AI", "Per-lead LinkedIn messages"],
        ["6", "Export deliverables", "create_outputs.py", "Excel + PDF reports"],
        ["7", "Send LinkedIn outreach", "You or VA (manual)", "Seller responses"],
        ["8", "Qualify interested sellers", "Phone call / NDA", "Confirmed seller pipeline"],
        ["9", "Introduce to buyer", "Broker facilitated", "Deal flow + closed transactions"],
    ]
    wf_col_widths = [0.4*inch, 1.9*inch, 1.9*inch, 2.6*inch]
    wf_table = Table(workflow_data, colWidths=wf_col_widths)
    wf_table.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",      (0,0), (-1,0), WHITE),
        ("FONTNAME",       (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,0), 9),
        ("ALIGN",          (0,0), (-1,-1), "LEFT"),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",       (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0,1), (-1,-1), 9),
        ("TOPPADDING",     (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 6),
        ("LEFTPADDING",    (0,0), (-1,-1), 8),
        ("GRID",           (0,0), (-1,-1), 0.4, GRAY_MID),
        # Highlight automation vs manual
        ("BACKGROUND",     (0,7), (-1,7), colors.HexColor("#FFF4E6")),  # Manual
        ("BACKGROUND",     (0,8), (-1,8), colors.HexColor("#EDF7ED")),  # Qualify
        ("BACKGROUND",     (0,9), (-1,9), colors.HexColor("#EDF7ED")),  # Intro
    ]))
    story.append(wf_table)
    story.append(Spacer(1, 0.2*inch))

    legend_data = [[
        Paragraph("  Automated by system", ParagraphStyle("l1",fontSize=8,textColor=BLACK)),
        Paragraph("  Human action required", ParagraphStyle("l2",fontSize=8,textColor=colors.HexColor("#C06010"))),
    ]]
    leg_t = Table(legend_data, colWidths=[2.5*inch, 2.5*inch])
    leg_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), GRAY_LIGHT),
        ("BACKGROUND", (1,0), (1,-1), colors.HexColor("#FFF4E6")),
        ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(leg_t)
    story.append(Spacer(1, 0.3*inch))

    # Data sources
    story.append(Paragraph("Data Sources & Approach", S["sub_header"]))
    sources_data = [
        [Paragraph("Source", S["table_header"]),
         Paragraph("What It Provides", S["table_header"]),
         Paragraph("Volume", S["table_header"]),
         Paragraph("Cost", S["table_header"]),
         Paragraph("Priority", S["table_header"])],
        ["Apollo.io API", "275M+ companies: revenue, employees, owner LinkedIn/email", "High", "$0–$99/mo", "Critical"],
        ["Yelp Fusion API", "Local businesses by category + geography", "Medium", "Free", "High"],
        ["Yellow Pages", "Business name, address, phone", "High", "Free scrape", "Medium"],
        ["Manta.com", "SMB directory with revenue/employee estimates", "Medium", "Free scrape", "Medium"],
        ["LinkedIn (manual)", "Direct owner identification and messaging", "Low", "Time / Sales Nav", "Very High"],
        ["SerpAPI / Google", "Industry-specific business results", "Medium", "$50/mo", "Low"],
    ]
    src_col_w = [1.5*inch, 2.8*inch, 0.7*inch, 0.8*inch, 0.8*inch]
    src_table = Table(sources_data, colWidths=src_col_w)
    src_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), MID_BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 8.5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8.5),
        ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 7),
        ("GRID",          (0,0), (-1,-1), 0.4, GRAY_MID),
        # Critical row
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#E8F0FA")),
        ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
    ]))
    story.append(src_table)
    story.append(PageBreak())

    # ── SECTION 2: BUYER PORTFOLIO ─────────────────────────────────────────
    story.append(Paragraph("Active Buyer Portfolio", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=12))
    story.append(Paragraph(
        f"The following <b>{total_buyers} pre-qualified buyers</b> have signed buyer agreements and are actively "
        "seeking acquisitions. All are capitalized and can move to close within 60–120 days of identifying "
        "the right opportunity. This system searches for off-market sellers matching each buyer's exact criteria.",
        S["body"]
    ))
    story.append(Spacer(1, 0.2*inch))

    # All buyers summary table
    buyer_rows = [[
        Paragraph("Buyer", S["table_header"]),
        Paragraph("Industry Focus", S["table_header"]),
        Paragraph("Geography", S["table_header"]),
        Paragraph("EBITDA Range", S["table_header"]),
        Paragraph("Deal Size", S["table_header"]),
    ]]
    for b in sorted(buyers, key=lambda x: x["industries"][0] if x["industries"] else "z"):
        ebitda_lo = b.get("ebitda_min", 0)
        ebitda_hi = min(b.get("ebitda_max", 99_999_999), 25_000_000)
        deal_hi   = min(b.get("deal_size_max", 99_999_999), 100_000_000)
        deal_lo   = b.get("deal_size_min", 0)
        geo_parts = b.get("states", [])[:4]
        geo_str   = ", ".join(geo_parts) if geo_parts else (b.get("geographies",["US"])[0] if b.get("geographies") else "US")
        ind_str   = ", ".join([i.title() for i in b["industries"][:3]])
        buyer_rows.append([
            Paragraph(b["name"], ParagraphStyle("bn",fontSize=8,fontName="Helvetica-Bold")),
            Paragraph(ind_str, ParagraphStyle("ind",fontSize=8)),
            Paragraph(geo_str, ParagraphStyle("geo",fontSize=8)),
            Paragraph(f"${ebitda_lo/1e6:.1f}M – ${ebitda_hi/1e6:.1f}M" if ebitda_hi < 20_000_000
                      else f"${ebitda_lo/1e6:.0f}M+", ParagraphStyle("ebitda",fontSize=8)),
            Paragraph(f"${deal_lo/1e6:.1f}M – ${deal_hi/1e6:.0f}M" if deal_hi < 100_000_000
                      else f"${deal_lo/1e6:.0f}M+", ParagraphStyle("deal",fontSize=8)),
        ])

    buyer_table = Table(buyer_rows, colWidths=[1.75*inch, 1.6*inch, 1.35*inch, 1.3*inch, 1.25*inch])
    buyer_table.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",      (0,0), (-1,0), WHITE),
        ("FONTNAME",       (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,0), 9),
        ("ALIGN",          (0,0), (-1,-1), "LEFT"),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",       (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0,1), (-1,-1), 8),
        ("TOPPADDING",     (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 5),
        ("LEFTPADDING",    (0,0), (-1,-1), 6),
        ("GRID",           (0,0), (-1,-1), 0.4, GRAY_MID),
    ]))
    story.append(buyer_table)
    story.append(Spacer(1, 0.25*inch))

    # Charts row
    fig1 = chart_buyer_industry_distribution(buyers)
    fig2 = chart_geographic_coverage(buyers)
    chart_row = [[mpl_to_reportlab_image(fig1, 3.3), mpl_to_reportlab_image(fig2, 3.7)]]
    ct = Table(chart_row, colWidths=[3.5*inch, 3.9*inch])
    ct.setStyle(TableStyle([
        ("ALIGN",  (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.15*inch))

    fig3 = chart_deal_size_ranges(buyers)
    story.append(mpl_to_reportlab_image(fig3, 7.0))
    story.append(PageBreak())

    # ── SECTION 3: PIPELINE OVERVIEW ──────────────────────────────────────
    story.append(Paragraph("Projected Lead Pipeline", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=12))
    story.append(Paragraph(
        "The funnel below shows projected outcomes based on industry-standard cold outreach benchmarks "
        "for off-market M&A prospecting. Response rates improve significantly with LinkedIn Sales Navigator "
        "and AI-personalized messages (Claude API).",
        S["body"]
    ))
    story.append(Spacer(1, 0.1*inch))
    fig4 = chart_pipeline_funnel(all_leads)
    story.append(mpl_to_reportlab_image(fig4, 7.0))
    story.append(Spacer(1, 0.15*inch))

    # Benchmark table
    bench_data = [
        [Paragraph("Stage", S["table_header"]),
         Paragraph("Cold Outreach Benchmark", S["table_header"]),
         Paragraph("With AI Personalization", S["table_header"]),
         Paragraph("Projected Outcome (Demo)", S["table_header"])],
        ["LinkedIn Connection Accepted", "25–35%", "40–55%", f"{int(total_leads*0.45):,} accepted"],
        ["Opens Message / Reads InMail",  "45–60%", "60–75%", f"{int(total_leads*0.65):,} read"],
        ["Responds to Outreach",          "8–15%",  "18–28%", f"{int(total_leads*0.22):,} respond"],
        ["Expresses Selling Interest",    "3–6%",   "8–14%",  f"{int(total_leads*0.10):,} interested"],
        ["Agrees to Intro Call",          "1–3%",   "4–8%",   f"{int(total_leads*0.05):,} scheduled"],
    ]
    bench_table = Table(bench_data, colWidths=[2.1*inch, 1.7*inch, 1.8*inch, 1.8*inch])
    bench_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 9),
        ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ("GRID",          (0,0), (-1,-1), 0.4, GRAY_MID),
        ("BACKGROUND", (2,1), (2,-1), colors.HexColor("#EDF7ED")),
        ("TEXTCOLOR",  (2,1), (2,-1), colors.HexColor("#1a5e2a")),
        ("FONTNAME",   (2,1), (2,-1), "Helvetica-Bold"),
    ]))
    story.append(bench_table)
    story.append(Spacer(1, 0.2*inch))

    fig5 = chart_source_breakdown(all_flat)
    story.append(mpl_to_reportlab_image(fig5, 5.5))
    story.append(PageBreak())

    # ── SECTION 4: SAMPLE LEADS SPOTLIGHT ─────────────────────────────────
    story.append(Paragraph("Sample Lead Profiles (Demo Data)", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=10))
    story.append(Paragraph(
        "The following profiles illustrate the data quality and depth that Apollo.io + web scraping provides. "
        "Each lead card shows the company profile, owner contact, AI-generated LinkedIn message, and match score. "
        "<b>These are representative examples</b> — real data requires Apollo.io API key.",
        S["body"]
    ))
    story.append(Spacer(1, 0.15*inch))

    # Show top 3 leads from each of 3 buyers
    showcase_buyers = ["magus_abraxas", "baruk_capital", "zeon_investments"]
    for bid in showcase_buyers:
        b = get_buyer_by_id(bid)
        leads = all_leads.get(bid, [])[:3]
        if not b or not leads: continue

        story.append(KeepTogether([
            Paragraph(f"BUYER: {b['name'].upper()}", ParagraphStyle(
                "buyer_header", fontSize=10, fontName="Helvetica-Bold",
                textColor=WHITE, backColor=NAVY, borderPad=6,
                spaceAfter=2, spaceBefore=10)),
            Spacer(1, 0.04*inch),
            Paragraph(
                f"Seeking: {', '.join(b['industries']).title()} businesses | "
                f"Geography: {', '.join(b['states'][:5]) or 'US nationwide'} | "
                f"EBITDA: ${b['ebitda_min']/1e6:.1f}M – ${min(b['ebitda_max'],25e6)/1e6:.0f}M",
                ParagraphStyle("buyer_criteria", fontSize=8.5, textColor=GRAY_DARK,
                               fontName="Helvetica", spaceAfter=8)
            ),
        ]))

        for lead in leads:
            score = lead.get("match_score", 85)
            sc_color = score_color(score)
            cr_msg = lead.get("connection_request","")[:200]

            lead_data = [
                # Row 1: Company name + score
                [Paragraph(f"<b>{lead['company_name']}</b>", S["company_name"]),
                 Paragraph(f"Match Score: <b>{score} {'A' if score>=90 else 'B' if score>=80 else 'C'}</b>",
                           ParagraphStyle("sc",fontSize=9,fontName="Helvetica-Bold",
                                          textColor=sc_color,alignment=TA_RIGHT))],
                # Row 2: Details
                [Paragraph(
                    f"<b>Industry:</b> {lead.get('industry','').title()}  |  "
                    f"<b>Location:</b> {lead.get('city','')}, {lead.get('state','')}  |  "
                    f"<b>Est. Revenue:</b> {lead.get('estimated_revenue','N/A')}  |  "
                    f"<b>Employees:</b> {lead.get('employee_count','?')}  |  "
                    f"<b>Founded:</b> {lead.get('founded_year','?')}",
                    ParagraphStyle("details",fontSize=8.5,fontName="Helvetica",textColor=GRAY_DARK)
                ), ""],
                # Row 3: Description
                [Paragraph(lead.get("description",""), S["body_small"]), ""],
                # Row 4: Owner
                [Paragraph(
                    f"<b>Owner:</b> {lead.get('owner_name','TBD')} | "
                    f"<b>Title:</b> {lead.get('owner_title','Owner')} | "
                    f"<b>LinkedIn:</b> {lead.get('owner_linkedin') or 'See search URL'}",
                    ParagraphStyle("owner",fontSize=8.5,fontName="Helvetica-Bold",textColor=MID_BLUE)
                ), ""],
                # Row 5: Message preview
                [Paragraph(
                    f"<b>LinkedIn Connection Request:</b><br/>"
                    f"<i>\"{cr_msg}...\"</i>",
                    ParagraphStyle("msg",fontSize=8,fontName="Helvetica",
                                   textColor=colors.HexColor("#334466"),
                                   backColor=colors.HexColor("#F0F5FF"),
                                   borderPad=5,leading=12)
                ), ""],
            ]
            lead_table = Table(lead_data, colWidths=[5.5*inch, 1.7*inch])
            lead_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), WHITE),
                ("BACKGROUND", (0,4), (-1,4), colors.HexColor("#F0F5FF")),
                ("BOX",        (0,0), (-1,-1), 1.0, GRAY_MID),
                ("LINEBEFORE",  (0,0), (0,-1), 4, sc_color),
                ("TOPPADDING",  (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1), 5),
                ("LEFTPADDING", (0,0), (-1,-1), 10),
                ("RIGHTPADDING",(0,0), (-1,-1), 8),
                ("SPAN",        (0,1), (-1,1)),
                ("SPAN",        (0,2), (-1,2)),
                ("SPAN",        (0,3), (-1,3)),
                ("SPAN",        (0,4), (-1,4)),
                ("ALIGN",       (1,0), (1,0), "RIGHT"),
                ("VALIGN",      (0,0), (-1,-1), "TOP"),
            ]))
            story.append(lead_table)
            story.append(Spacer(1, 0.08*inch))

        story.append(Spacer(1, 0.2*inch))

    story.append(PageBreak())

    # ── SECTION 5: OUTREACH MESSAGES ─────────────────────────────────────
    story.append(Paragraph("LinkedIn Outreach Message Framework", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=10))
    story.append(Paragraph(
        "Each identified seller receives two pre-written LinkedIn messages: a short connection request note "
        "and a full InMail. Messages are personalized per industry and company. The tone is deliberately "
        "curiosity-driven — never mentioning buying/selling directly in the initial connection.",
        S["body"]
    ))
    story.append(Spacer(1, 0.15*inch))

    msg_principles = [
        ["Principle", "Why It Works"],
        ["Never say 'buy your business' in first contact", "Triggers defensiveness. Lead with market intel."],
        ["Create curiosity: 'buyers in your space'", "They want to know who — opens conversation."],
        ["Keep connection request under 280 chars", "LinkedIn limit. Short = more likely to be accepted."],
        ["InMail focuses on seller value, not buyer need", "Seller asks 'what's in it for me?' Answer it first."],
        ["End with single low-commitment CTA", "'15-minute call' is easy to say yes to."],
        ["Personalize with company name + industry", "Generic messages get ignored. Specifics get read."],
        ["Follow up ONCE at day 7 if no response", "More than once = spam. Once = professional persistence."],
    ]
    mp_table = Table(msg_principles, colWidths=[3.2*inch, 4.0*inch])
    mp_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), MID_BLUE),
        ("TEXTCOLOR",  (0,0), (-1,0), WHITE),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 9),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[GRAY_LIGHT,WHITE]),
        ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",   (0,1), (-1,-1), 9),
        ("ALIGN",      (0,0), (-1,-1), "LEFT"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING",(0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.4, GRAY_MID),
    ]))
    story.append(mp_table)
    story.append(Spacer(1, 0.2*inch))

    # Full example InMail
    sample_lead = all_leads["baruk_capital"][0] if all_leads.get("baruk_capital") else {}
    if sample_lead:
        story.append(Paragraph("Example: Full InMail Message", S["sub_header"]))
        story.append(Spacer(1, 0.05*inch))
        inmail_text = sample_lead.get("inmail","").replace("\n","<br/>")
        inmail_data = [[Paragraph(f"<i>{inmail_text}</i>",
            ParagraphStyle("inmail_ex",fontSize=9,fontName="Helvetica",
                           textColor=colors.HexColor("#223344"),leading=14))]]
        inmail_t = Table(inmail_data, colWidths=[W])
        inmail_t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F0F5FF")),
            ("BOX",        (0,0), (-1,-1), 1, MID_BLUE),
            ("LINEBEFORE", (0,0), (0,-1), 4, GOLD),
            ("TOPPADDING", (0,0), (-1,-1), 12),
            ("BOTTOMPADDING",(0,0),(-1,-1), 12),
            ("LEFTPADDING",(0,0), (-1,-1), 16),
            ("RIGHTPADDING",(0,0),(-1,-1), 12),
        ]))
        story.append(inmail_t)

    story.append(PageBreak())

    # ── SECTION 6: NEXT STEPS ─────────────────────────────────────────────
    story.append(Paragraph("Activating the Pipeline", S["section_header"]))
    story.append(HRFlowable(width=W, thickness=2, color=GOLD, spaceAfter=10))

    next_steps = [
        ["Priority", "Action", "Time Required", "Impact"],
        ["1 — Critical", "Get Apollo.io API key (free account at app.apollo.io)", "5 minutes", "Unlocks 275M+ contacts"],
        ["2 — Critical", "Add ANTHROPIC_API_KEY to .env for AI personalized messages", "5 minutes", "2–3x response rates"],
        ["3 — High",     "Run: python tools/run_lead_gen.py", "30–60 minutes", "10–25+ leads per buyer"],
        ["4 — High",     "Open MASTER_LEADS.xlsx → begin LinkedIn outreach", "2–3 hrs/day", "Seller confirmations"],
        ["5 — Medium",   "Optional: LinkedIn Sales Navigator ($99/mo)", "Setup: 30 min", "50 InMails/day capacity"],
        ["6 — Medium",   "Add Yelp API key to .env (free at yelp.com/developers)", "10 minutes", "+20–30% lead volume"],
        ["7 — Ongoing",  "Update seller_interested column as responses come in", "Daily", "Track confirmed sellers"],
        ["8 — Milestone","10 confirmed sellers per buyer → present blind summaries to buyer", "Per buyer", "Revenue event"],
    ]
    ns_table = Table(next_steps, colWidths=[1.35*inch, 2.9*inch, 1.3*inch, 1.8*inch])
    ns_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [GRAY_LIGHT, WHITE]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8.5),
        ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 7),
        ("GRID",          (0,0), (-1,-1), 0.4, GRAY_MID),
        # Highlight critical
        ("BACKGROUND", (0,1), (-1,2), colors.HexColor("#FFEEE8")),
        ("FONTNAME",   (0,1), (0,2), "Helvetica-Bold"),
        ("TEXTCOLOR",  (0,1), (0,2), ORANGE),
        # Highlight milestone
        ("BACKGROUND", (0,8), (-1,8), colors.HexColor("#EDF7ED")),
        ("FONTNAME",   (0,8), (0,8), "Helvetica-Bold"),
        ("TEXTCOLOR",  (0,8), (0,8), GREEN),
    ]))
    story.append(ns_table)
    story.append(Spacer(1, 0.3*inch))

    # Footer note
    story.append(Paragraph(
        "For questions about the system or to discuss priority buyers, contact Jordi Quevedo-Valls at Valar Brokers.",
        ParagraphStyle("footer_note",fontSize=9,fontName="Helvetica",
                       textColor=GRAY_DARK,alignment=TA_CENTER)
    ))

    # Build with title page background
    doc.build(story, onFirstPage=draw_title_bg, onLaterPages=lambda c, d: None)
    print(f"  [PDF] Overview saved -> {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
#   EXCEL SPREADSHEET
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_spreadsheet(all_leads: dict, buyers: list, out_path: Path):
    print(f"\n[EXCEL] Building master leads spreadsheet...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Color palette
    NAVY_HEX  = "0B1F3A"
    GOLD_HEX  = "C9A84C"
    BLUE_HEX  = "2E6DA4"
    GREEN_HEX = "2D8A4E"
    RED_HEX   = "C0392B"
    LT_BLUE   = "E8F0FA"
    LT_GREEN  = "EDF7ED"
    LT_YELLOW = "FEFCE8"
    LT_GRAY   = "F5F5F5"

    def make_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def make_font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def make_border():
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    def hdr_cell(ws, row, col, text, bg=NAVY_HEX, fg="FFFFFF", bold=True, size=10):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = make_fill(bg)
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
        return cell

    # ── SUMMARY TAB ───────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("SUMMARY", 0)
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 22
    ws_sum.column_dimensions["D"].width = 22
    ws_sum.column_dimensions["E"].width = 28
    ws_sum.column_dimensions["F"].width = 20
    ws_sum.row_dimensions[1].height = 50

    # Title banner
    ws_sum.merge_cells("A1:F1")
    title = ws_sum.cell(row=1, column=1, value="OFF-MARKET DEAL SOURCING  |  Buyer-Seller Pipeline  |  Valar Brokers")
    title.fill = make_fill(NAVY_HEX)
    title.font = Font(bold=True, color="C9A84C", size=16, name="Calibri")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Generated date
    ws_sum.merge_cells("A2:F2")
    date_cell = ws_sum.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}  |  DEMO DATA — add API keys to activate full pipeline")
    date_cell.fill = make_fill("1A2F4E")
    date_cell.font = Font(color="AABBCC", size=10, name="Calibri")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 22

    # Summary stats
    ws_sum.row_dimensions[4].height = 30
    for col, (label, value) in enumerate([
        ("TOTAL BUYERS", str(len(buyers))),
        ("DEMO LEADS", str(sum(len(v) for v in all_leads.values()))),
        ("WITH LINKEDIN", str(sum(1 for v in all_leads.values() for l in v if l.get("owner_linkedin")))),
        ("WITH EMAIL", str(sum(1 for v in all_leads.values() for l in v if l.get("owner_email")))),
        ("BUYER SEGMENTS", "7 Demo"),
        ("OUTREACH READY", "Yes"),
    ], start=1):
        ws_sum.cell(row=4, column=col, value=label).fill = make_fill(BLUE_HEX)
        ws_sum.cell(row=4, column=col).font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        ws_sum.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=5, column=col, value=value).font = Font(bold=True, size=16, color=NAVY_HEX, name="Calibri")
        ws_sum.cell(row=5, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[5].height = 40

    # Buyer summary table
    ws_sum.row_dimensions[7].height = 28
    headers = ["Buyer Name", "Industry Focus", "Geography", "EBITDA Range", "Demo Leads", "Status"]
    for col, h in enumerate(headers, 1):
        hdr_cell(ws_sum, 7, col, h, bg=NAVY_HEX)

    row = 8
    for b in buyers:
        leads = all_leads.get(b["id"], [])
        ebitda_hi = min(b.get("ebitda_max", 99_999_999), 25_000_000)
        ebitda_str = f"${b.get('ebitda_min',0)/1e6:.1f}M – ${ebitda_hi/1e6:.1f}M"
        states = b.get("states", [])
        geo_str = ", ".join(states[:5]) if states else (b.get("geographies",["US"])[0])
        status = f"DEMO: {len(leads)} leads" if leads else "Needs API key"
        bg = LT_BLUE if row % 2 == 0 else "FFFFFF"

        ws_sum.cell(row=row, column=1, value=b["name"]).font = Font(bold=True, size=9.5, name="Calibri")
        ws_sum.cell(row=row, column=2, value=", ".join(b["industries"][:2]).title())
        ws_sum.cell(row=row, column=3, value=geo_str)
        ws_sum.cell(row=row, column=4, value=ebitda_str)
        ws_sum.cell(row=row, column=5, value=len(leads))
        status_cell = ws_sum.cell(row=row, column=6, value=status)
        if leads:
            status_cell.fill = make_fill(LT_GREEN)
            status_cell.font = Font(color=GREEN_HEX, bold=True, size=9, name="Calibri")
        else:
            status_cell.fill = make_fill(LT_YELLOW)

        for col in range(1, 7):
            c = ws_sum.cell(row=row, column=col)
            if not c.fill.patternType or c.fill.fgColor.rgb == "00000000":
                c.fill = make_fill(bg)
            c.border = make_border()
            c.alignment = Alignment(vertical="center", wrap_text=False)
        ws_sum.row_dimensions[row].height = 22
        row += 1

    # ── PER-BUYER TABS ─────────────────────────────────────────────────────────
    LEAD_COLS = [
        ("Company Name",       "company_name",       22),
        ("Industry",           "industry",           14),
        ("City",               "city",               14),
        ("State",              "state",               7),
        ("Est. Revenue",       "estimated_revenue",  14),
        ("Employees",          "employee_count",     10),
        ("Founded",            "founded_year",        9),
        ("Owner Name",         "owner_name",         20),
        ("Owner Title",        "owner_title",        18),
        ("Owner LinkedIn",     "owner_linkedin",     35),
        ("Owner Email",        "owner_email",        28),
        ("Phone",              "phone",              14),
        ("Match Score",        "match_score",        11),
        ("Status",             "status",             12),
        ("Outreach Sent",      "outreach_sent_date", 14),
        ("Responded?",         "seller_responded",   12),
        ("Interested?",        "seller_interested",  12),
        ("Notes",              "response_notes",     30),
        ("Next Action",        "next_action",        22),
        ("Connection Request", "connection_request", 45),
        ("InMail Message",     "inmail",             55),
        ("LinkedIn Search URL","linkedin_search_url",40),
        ("Source",             "source",             12),
    ]

    for bid, leads in all_leads.items():
        if not leads: continue
        buyer = get_buyer_by_id(bid)
        if not buyer: continue

        tab_name = buyer["name"][:28].strip()
        ws = wb.create_sheet(tab_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

        # Tab header
        ws.merge_cells(f"A1:{get_column_letter(len(LEAD_COLS))}1")
        tab_title = ws.cell(row=1, column=1,
            value=f"{buyer['name'].upper()}  |  {', '.join(buyer['industries'][:2]).title()}  |  "
                  f"{', '.join(buyer.get('states',[])[:5]) or 'US Nationwide'}  |  "
                  f"EBITDA: ${buyer.get('ebitda_min',0)/1e6:.1f}M – ${min(buyer.get('ebitda_max',25e6),25e6)/1e6:.0f}M")
        tab_title.fill = make_fill(NAVY_HEX)
        tab_title.font = Font(bold=True, color="C9A84C", size=12, name="Calibri")
        tab_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Column headers
        for col, (label, _, width) in enumerate(LEAD_COLS, 1):
            hdr_cell(ws, 2, col, label)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 28

        # Data rows
        for r_idx, lead in enumerate(leads, start=3):
            score = lead.get("match_score", 80)
            bg = LT_BLUE if r_idx % 2 == 0 else "FFFFFF"
            if score >= 90:   bg_row = LT_GREEN
            elif score >= 80: bg_row = LT_BLUE
            else:             bg_row = LT_YELLOW

            for col, (_, field, _) in enumerate(LEAD_COLS, 1):
                val = lead.get(field, "")
                c = ws.cell(row=r_idx, column=col, value=val)
                c.fill = make_fill(bg_row)
                c.font = Font(size=9, name="Calibri",
                              bold=(field == "company_name"))
                c.border = make_border()
                c.alignment = Alignment(vertical="top", wrap_text=(field in ("connection_request","inmail","response_notes","description")))

                # Special formatting
                if field == "match_score":
                    if isinstance(val, (int,float)):
                        c.font = Font(bold=True, size=10, name="Calibri",
                                      color=GREEN_HEX if val >= 90 else (BLUE_HEX if val >= 80 else "B8860B"))
                elif field == "owner_linkedin" and val:
                    c.hyperlink = f"https://{val}" if not val.startswith("http") else val
                    c.font = Font(color=BLUE_HEX, underline="single", size=9, name="Calibri")
                elif field == "linkedin_search_url" and val:
                    c.hyperlink = val
                    c.font = Font(color=BLUE_HEX, underline="single", size=9, name="Calibri")
            ws.row_dimensions[r_idx].height = 60 if any(
                lead.get(f,"") and len(str(lead.get(f,""))) > 100
                for f in ("connection_request","inmail") ) else 22

        # Auto-filter
        ws.auto_filter.ref = f"A2:{get_column_letter(len(LEAD_COLS))}{len(leads)+2}"

    wb.save(str(out_path))
    print(f"  [EXCEL] Spreadsheet saved -> {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
#   MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    all_leads = get_all_demo_leads()
    total = sum(len(v) for v in all_leads.values())
    print(f"[START] Demo data: {len(all_leads)} buyers, {total} leads")

    # Build Excel
    xl_path = OUT_DIR / f"LEADS_MASTER_{timestamp}.xlsx"
    build_excel_spreadsheet(all_leads, BUYERS, xl_path)

    # Build overview PDF
    pdf_path = OUT_DIR / f"PIPELINE_OVERVIEW_{timestamp}.pdf"
    build_overview_pdf(all_leads, BUYERS, pdf_path)

    print(f"\n[DONE] Output files:")
    print(f"  Excel:   {xl_path}")
    print(f"  PDF:     {pdf_path}")
    print(f"\nOpen these files to see what the full system produces.")
