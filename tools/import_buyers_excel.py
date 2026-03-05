"""
import_buyers_excel.py
----------------------
Reads BUYERS_TEMPLATE.xlsx and converts it to .tmp/buyers.json.
Merges with any existing buyers already in buyers.json (won't duplicate by buyer_id).

Usage:
    python tools/import_buyers_excel.py
    python tools/import_buyers_excel.py --excel data/BUYERS_TEMPLATE.xlsx --out .tmp/buyers.json
    python tools/import_buyers_excel.py --replace   # overwrite buyers.json completely
"""

import argparse
import json
import os
import re
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()

TEMPLATE_PATH = Path("data/BUYERS_TEMPLATE.xlsx")
OUT_PATH = Path(os.getenv("BUYERS_JSON_PATH", ".tmp/buyers.json"))

# Row numbers in the template (1-indexed)
HEADER_ROW = 3   # Column headers
DATA_START  = 6  # First real data row (row 5 is the example row — skip it)

# Column index → field mapping (1-indexed, matches create_templates.py order)
COL = {
    "buyer_id":               1,
    "buyer_name":             2,
    "company_name":           3,
    "contact_email":          4,
    "contact_phone":          5,
    "agreement_signed_date":  6,
    "industry_preferences":   7,
    "industry_exclusions":    8,
    "geography_states":       9,
    "asking_price_min":       10,
    "asking_price_max":       11,
    "revenue_min":            12,
    "revenue_max":            13,
    "cash_flow_min":          14,
    "cash_flow_multiple_max": 15,
    "sba_loan_preferred":     16,
    "seller_financing_ok":    17,
    "all_cash_ok":            18,
    "employees_min":          19,
    "employees_max":          20,
    "years_in_business_min":  21,
    "recurring_revenue_preferred": 22,
    "weight_industry":        23,
    "weight_geography":       24,
    "weight_financials":      25,
    "weight_cf_multiple":     26,
    "weight_years":           27,
    "weight_structure":       28,
    "buyer_profile_summary":  29,
    "notes":                  30,
}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def cell_int(ws, row: int, col: int) -> int | None:
    v = ws.cell(row=row, column=col).value
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).replace(",", "").replace("$", "").strip()))
    except ValueError:
        return None


def cell_float(ws, row: int, col: int) -> float | None:
    v = ws.cell(row=row, column=col).value
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def cell_bool(ws, row: int, col: int) -> bool:
    v = str(ws.cell(row=row, column=col).value or "").strip().lower()
    return v in ("yes", "true", "1", "y")


def parse_list(raw: str) -> list[str]:
    """Split 'HVAC; Plumbing; Home Services' into ['HVAC', 'Plumbing', 'Home Services']."""
    if not raw:
        return []
    parts = re.split(r'[;,\n]', raw)
    return [p.strip() for p in parts if p.strip()]


def parse_states(raw: str) -> list[str]:
    """Extract valid 2-letter state codes from a semicolon/comma-separated string."""
    ALL_STATES = {
        "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
        "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
        "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
        "VA","WA","WV","WI","WY",
    }
    if not raw:
        return []
    parts = re.split(r'[;,\s]+', raw.upper())
    return [p.strip() for p in parts if p.strip() in ALL_STATES]


def validate_weights(weights: dict, buyer_name: str) -> dict:
    total = sum(weights.values())
    if total == 0:
        # All blank — use defaults
        return {"industry_match": 30, "geography_match": 20, "financials_in_range": 25,
                "cash_flow_multiple": 10, "years_established": 8, "deal_structure": 7}
    if total != 100:
        print(f"  [WARN] {buyer_name}: weights sum to {total}, not 100. Normalizing.")
        factor = 100 / total
        return {k: round(v * factor, 1) for k, v in weights.items()}
    return weights


# ---------------------------------------------------------------------------
# Row parser
# ---------------------------------------------------------------------------

def parse_row(ws, row: int, idx: int) -> dict | None:
    """Parse one data row from the worksheet. Returns None if row is empty."""
    buyer_name = cell_str(ws, row, COL["buyer_name"])
    email      = cell_str(ws, row, COL["contact_email"])

    if not buyer_name and not email:
        return None  # Empty row — skip

    # Buyer ID
    buyer_id = cell_str(ws, row, COL["buyer_id"])
    if not buyer_id:
        buyer_id = f"buyer_{idx:03d}"

    # Scoring weights
    raw_weights = {
        "industry_match":      cell_float(ws, row, COL["weight_industry"])  or 30,
        "geography_match":     cell_float(ws, row, COL["weight_geography"]) or 20,
        "financials_in_range": cell_float(ws, row, COL["weight_financials"]) or 25,
        "cash_flow_multiple":  cell_float(ws, row, COL["weight_cf_multiple"]) or 10,
        "years_established":   cell_float(ws, row, COL["weight_years"])     or 8,
        "deal_structure":      cell_float(ws, row, COL["weight_structure"]) or 7,
    }
    weights = validate_weights(raw_weights, buyer_name)

    return {
        "buyer_id":             buyer_id,
        "buyer_name":           buyer_name,
        "company_name":         cell_str(ws, row, COL["company_name"]),
        "contact_email":        email,
        "contact_phone":        cell_str(ws, row, COL["contact_phone"]),
        "agreement_type":       "buyer_broker_agreement",
        "agreement_signed_date": cell_str(ws, row, COL["agreement_signed_date"]),

        "criteria": {
            "industry_preferences": parse_list(cell_str(ws, row, COL["industry_preferences"])),
            "industry_exclusions":  parse_list(cell_str(ws, row, COL["industry_exclusions"])),
            "geography_states":     parse_states(cell_str(ws, row, COL["geography_states"])),

            "financials": {
                "asking_price_min":       cell_int(ws, row, COL["asking_price_min"]),
                "asking_price_max":       cell_int(ws, row, COL["asking_price_max"]),
                "revenue_min":            cell_int(ws, row, COL["revenue_min"]),
                "revenue_max":            cell_int(ws, row, COL["revenue_max"]),
                "cash_flow_min":          cell_int(ws, row, COL["cash_flow_min"]),
                "cash_flow_max":          None,
                "cash_flow_multiple_max": cell_float(ws, row, COL["cash_flow_multiple_max"]),
                "revenue_multiple_max":   None,
            },

            "deal_structure": {
                "seller_financing_ok":  cell_bool(ws, row, COL["seller_financing_ok"]),
                "sba_loan_preferred":   cell_bool(ws, row, COL["sba_loan_preferred"]),
                "all_cash_ok":          cell_bool(ws, row, COL["all_cash_ok"]),
                "real_estate_preferred": False,
            },

            "business_attributes": {
                "employees_min":             cell_int(ws, row, COL["employees_min"]),
                "employees_max":             cell_int(ws, row, COL["employees_max"]),
                "years_in_business_min":     cell_int(ws, row, COL["years_in_business_min"]),
                "absentee_owner_ok":         False,
                "recurring_revenue_preferred": cell_bool(ws, row, COL["recurring_revenue_preferred"]),
            },

            "scoring_weights": weights,
        },

        "buyer_profile_summary": cell_str(ws, row, COL["buyer_profile_summary"]),
        "notes":                 cell_str(ws, row, COL["notes"]),
        "_import_source":        "excel",
        "_imported_at":          str(date.today()),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import buyers from Excel template")
    parser.add_argument("--excel", default=str(TEMPLATE_PATH))
    parser.add_argument("--out",   default=str(OUT_PATH))
    parser.add_argument("--replace", action="store_true",
                        help="Replace buyers.json entirely (default: merge/update)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    out_path   = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}")
        print(f"[INFO]  Create it first: python tools/create_templates.py")
        return

    print(f"[INFO] Reading: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # Find the Buyers sheet
    sheet_name = "Buyers" if "Buyers" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"[INFO] Sheet: '{sheet_name}' — {ws.max_row} rows")

    # Parse all data rows
    new_buyers = []
    idx = 1
    for row in range(DATA_START, ws.max_row + 1):
        buyer = parse_row(ws, row, idx)
        if buyer:
            new_buyers.append(buyer)
            print(f"  [{buyer['buyer_id']}] {buyer['buyer_name'] or '(no name)'} | "
                  f"{buyer['criteria']['geography_states']} | "
                  f"{buyer['criteria']['industry_preferences'][:3]}")
            idx += 1

    if not new_buyers:
        print("[WARN] No buyer rows found in the Excel file.")
        print(f"[INFO] Fill in rows starting at row {DATA_START} and re-run.")
        return

    # Merge with existing buyers (unless --replace)
    existing = []
    if not args.replace and out_path.exists():
        try:
            existing = json.loads(out_path.read_text())
            print(f"[INFO] Existing buyers.json has {len(existing)} buyer(s)")
        except Exception:
            existing = []

    # Merge: update existing entries by buyer_id, append new ones
    existing_map = {b["buyer_id"]: b for b in existing}
    for buyer in new_buyers:
        existing_map[buyer["buyer_id"]] = buyer

    merged = list(existing_map.values())
    out_path.write_text(json.dumps(merged, indent=2))

    action = "Replaced" if args.replace else "Merged"
    print(f"\n[DONE] {action} buyers.json — {len(merged)} total buyer(s)")
    print(f"[INFO] Output: {out_path}")
    print(f"[INFO] Review the file, then run: python tools/run_pipeline.py")


if __name__ == "__main__":
    main()
