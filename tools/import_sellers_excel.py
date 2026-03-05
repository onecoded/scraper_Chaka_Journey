"""
import_sellers_excel.py
-----------------------
Reads SELLERS_TEMPLATE.xlsx and appends/merges deals into .tmp/all_listings.json.
Manually entered deals are treated exactly like scraped deals by the matching engine.

Usage:
    python tools/import_sellers_excel.py
    python tools/import_sellers_excel.py --excel data/SELLERS_TEMPLATE.xlsx --out .tmp/all_listings.json
    python tools/import_sellers_excel.py --replace   # overwrite all_listings.json
"""

import argparse
import json
import os
import re
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()

TEMPLATE_PATH = Path("data/SELLERS_TEMPLATE.xlsx")
OUT_PATH = Path(".tmp/all_listings.json")

HEADER_ROW = 3
DATA_START  = 6  # Skip example row (row 5)

COL = {
    "deal_id":            1,
    "title":              2,
    "source":             3,
    "url":                4,
    "industry":           5,
    "location_city":      6,
    "location_state":     7,
    "asking_price":       8,
    "annual_revenue":     9,
    "cash_flow":          10,
    "ebitda":             11,
    "employees":          12,
    "years_established":  13,
    "reason_for_selling": 14,
    "sba_eligible":       15,
    "financing_available": 16,
    "real_estate_included": 17,
    "inventory_value":    18,
    "ffe_value":          19,
    "broker_name":        20,
    "broker_email":       21,
    "broker_phone":       22,
    "date_listed":        23,
    "description":        24,
    "notes":              25,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def cell_str(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def cell_int(ws, row, col) -> int | None:
    v = ws.cell(row=row, column=col).value
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).replace(",", "").replace("$", "").strip()))
    except ValueError:
        return None


def cell_bool_tri(ws, row, col) -> bool | None:
    """Returns True/False/None for Yes/No/Unknown."""
    v = str(ws.cell(row=row, column=col).value or "").strip().lower()
    if v in ("yes", "y", "true", "1"):
        return True
    if v in ("no", "n", "false", "0"):
        return False
    return None  # Unknown


def fmt_price(val: int | None) -> str:
    if val is None:
        return ""
    if val >= 1_000_000:
        return f"${val/1_000_000:.1f}M"
    if val >= 1_000:
        return f"${val//1_000:,}K"
    return f"${val:,}"


def parse_row(ws, row: int, idx: int) -> dict | None:
    title = cell_str(ws, row, COL["title"])
    state = cell_str(ws, row, COL["location_state"]).upper()

    if not title and not state:
        return None

    # Generate deal ID
    deal_id = cell_str(ws, row, COL["deal_id"])
    if not deal_id:
        deal_id = f"manual_{idx:03d}"

    asking_price   = cell_int(ws, row, COL["asking_price"])
    annual_revenue = cell_int(ws, row, COL["annual_revenue"])
    cash_flow      = cell_int(ws, row, COL["cash_flow"])
    ebitda         = cell_int(ws, row, COL["ebitda"])
    sde            = cash_flow  # Treat CF as SDE for matching

    return {
        "deal_id":          deal_id,
        "source":           cell_str(ws, row, COL["source"]) or "manual",
        "url":              cell_str(ws, row, COL["url"]),
        "title":            title,
        "description":      cell_str(ws, row, COL["description"]),
        "industry":         cell_str(ws, row, COL["industry"]),
        "industry_category": cell_str(ws, row, COL["industry"]),
        "location_city":    cell_str(ws, row, COL["location_city"]),
        "location_state":   state,
        "asking_price":     asking_price,
        "asking_price_raw": fmt_price(asking_price),
        "annual_revenue":   annual_revenue,
        "annual_revenue_raw": fmt_price(annual_revenue),
        "cash_flow":        cash_flow,
        "cash_flow_raw":    fmt_price(cash_flow),
        "ebitda":           ebitda,
        "sde":              sde,
        "employees":        cell_int(ws, row, COL["employees"]),
        "years_established": cell_int(ws, row, COL["years_established"]),
        "reason_for_selling": cell_str(ws, row, COL["reason_for_selling"]),
        "real_estate_included": cell_bool_tri(ws, row, COL["real_estate_included"]),
        "inventory_included":   cell_bool_tri(ws, row, COL["real_estate_included"]),
        "inventory_value":  cell_int(ws, row, COL["inventory_value"]),
        "ff_e_value":       cell_int(ws, row, COL["ffe_value"]),
        "sba_eligible":     cell_bool_tri(ws, row, COL["sba_eligible"]),
        "financing_available": cell_bool_tri(ws, row, COL["financing_available"]),
        "broker_email":     cell_str(ws, row, COL["broker_email"]),
        "broker_phone":     cell_str(ws, row, COL["broker_phone"]),
        "broker_name":      cell_str(ws, row, COL["broker_name"]),
        "date_listed":      cell_str(ws, row, COL["date_listed"]),
        "date_scraped":     str(date.today()),
        "_notes":           cell_str(ws, row, COL["notes"]),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import seller deals from Excel template")
    parser.add_argument("--excel",   default=str(TEMPLATE_PATH))
    parser.add_argument("--out",     default=str(OUT_PATH))
    parser.add_argument("--replace", action="store_true",
                        help="Replace all_listings.json entirely (default: merge)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    out_path   = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}")
        print(f"[INFO]  Create it first: python tools/create_templates.py")
        return

    print(f"[INFO] Reading: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    sheet_name = "Seller Deals" if "Seller Deals" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"[INFO] Sheet: '{sheet_name}' — {ws.max_row} rows")

    new_listings = []
    idx = 1
    for row in range(DATA_START, ws.max_row + 1):
        listing = parse_row(ws, row, idx)
        if listing:
            new_listings.append(listing)
            print(f"  [{listing['deal_id']}] {listing['title'][:50]} | "
                  f"{listing['location_state']} | "
                  f"Ask: {listing['asking_price_raw']} | CF: {listing['cash_flow_raw']}")
            idx += 1

    if not new_listings:
        print("[WARN] No deals found in the Excel file.")
        print(f"[INFO] Fill in rows starting at row {DATA_START} and re-run.")
        return

    # Merge with existing listings
    existing = []
    if not args.replace and out_path.exists():
        try:
            existing = json.loads(out_path.read_text())
            print(f"[INFO] Existing listings: {len(existing)}")
        except Exception:
            existing = []

    existing_map = {l["deal_id"]: l for l in existing}
    for listing in new_listings:
        existing_map[listing["deal_id"]] = listing  # Upsert

    merged = list(existing_map.values())
    out_path.write_text(json.dumps(merged, indent=2))

    action = "Replaced" if args.replace else "Merged"
    print(f"\n[DONE] {action} listings — {len(merged)} total deals in {out_path}")
    print(f"[INFO] Run matching: python tools/match_deals.py")
    print(f"[INFO] Or full pipeline: python tools/run_pipeline.py --skip-scrape")


if __name__ == "__main__":
    main()
