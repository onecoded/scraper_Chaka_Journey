"""
dealflow_app.py — Valar Advisory Deal Flow Platform
5-Tab Streamlit app:
  Tab 1: Lead Stream    — CRM-style cards, auto-scrape, interest tracking
  Tab 2: Deal Pipeline  — Full broker pipeline, 1-click NDA/CIM, LOI to close
  Tab 3: Outreach       — Cross-post deals to buyers, LinkedIn, tracking
  Tab 4: Automation     — Auto-draft emails, task queue
  Tab 5: Buyers & Settings — Buyer mandate management
"""

import os
import sys
import json
import time
import imaplib
import smtplib
import email as email_lib
from datetime import datetime, timedelta
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.parse import quote as _urlquote

import streamlit as st
import pandas as pd
import plotly.graph_objects as go

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "tools"))

from dealflow_db import (
    DEAL_STAGES, STAGE_LABELS, STAGE_COLORS, INTEREST_COLORS, INTEREST_EMOJI,
    get_all_buyers, get_buyer, upsert_buyer, delete_buyer,
    get_all_deals, get_deal, add_deal, update_deal, delete_deal,
    deal_exists, get_pipeline_stats, mark_draft_created,
    log_outreach, get_outreach_log, mark_cross_posted,
    add_seller_intake, get_seller_intakes, mark_intake_reviewed,
    get_auto_draft_candidates,
)

# ── lazy imports ──────────────────────────────────────────────────────────────

def _buyers_db():
    try:
        import buyers_db
        return buyers_db.BUYERS
    except Exception:
        return []

def _apollo_search(buyer, max_results=10):
    try:
        import apollo_search
        return apollo_search.search_leads_for_buyer(buyer, max_results)
    except Exception:
        return []

def _web_scraper(buyer, max_per_source=5):
    try:
        import web_scraper
        return web_scraper.scrape_all_sources(buyer, max_per_source)
    except Exception:
        return []

def _generate_email(lead: dict, buyer: dict) -> str:
    try:
        import message_generator
        msg = message_generator.generate_messages_with_ai(lead, buyer)
        return msg.get("inmail", msg.get("connection_request", ""))
    except Exception:
        return _default_outreach_email(lead, buyer)

# ── config ────────────────────────────────────────────────────────────────────

from dotenv import load_dotenv
# Try project .env first, then parent
_ENV_FILE = ROOT / ".env"
if not _ENV_FILE.exists():
    _ENV_FILE = ROOT.parent / ".env"
load_dotenv(_ENV_FILE)

BROKER_NAME    = os.getenv("BROKER_NAME",    "Joseph Schneekloth")
BROKER_TITLE   = os.getenv("BROKER_TITLE",   "Chief Growth Officer")
BROKER_COMPANY = os.getenv("BROKER_COMPANY", "Valar Advisory")
BROKER_PHONE   = os.getenv("BROKER_PHONE",   "641-451-7288")
BROKER_EMAIL   = os.getenv("BROKER_EMAIL",   os.getenv("SMTP_USER", "Joseph.Schneek@gmail.com"))
SMTP_HOST      = os.getenv("SMTP_HOST",      "smtp.gmail.com")
SMTP_PORT      = int(os.getenv("SMTP_PORT",  587))
SMTP_USER      = os.getenv("SMTP_USER",      "")
SMTP_PASS      = os.getenv("SMTP_PASSWORD",  "")

LEADS_DIR = ROOT / ".tmp" / "leads"
TMP_DIR   = ROOT / ".tmp"

# Legacy status list kept for backward compat
STATUSES = ["identified", "emailed", "responded", "qualified", "dead"]

# ── helpers ───────────────────────────────────────────────────────────────────

# Industry synonym clusters — any two industries sharing a cluster are considered equivalent
INDUSTRY_CLUSTERS = {
    "home_services":   ["hvac","home services","plumbing","electrical","landscaping",
                        "roofing","pest control","cleaning","lawn care","home improvement"],
    "manufacturing":   ["manufacturing","fabrication","industrial","production","cnc",
                        "machining","assembly","metal","plastics","aerospace","defense"],
    "transportation":  ["transportation","trucking","logistics","freight","delivery",
                        "courier","shipping","warehousing","distribution","supply chain",
                        "last mile"],
    "technology":      ["technology","software","it","saas","digital","tech","cyber",
                        "cloud","app","platform","managed services","msp","data"],
    "healthcare":      ["healthcare","medical","dental","therapy","health","clinic",
                        "pharma","biotech","home health","behavioral health","veterinary"],
    "construction":    ["construction","contractor","general contractor","excavation",
                        "civil","concrete","roofing","framing","drywall","flooring"],
    "food_bev":        ["food","restaurant","beverage","hospitality","catering","bakery",
                        "grocery","food manufacturing","food distribution","brewery"],
    "financial":       ["financial services","finance","accounting","insurance",
                        "wealth management","tax","bookkeeping","banking","fintech"],
    "business_svcs":   ["business services","staffing","hr","marketing","consulting",
                        "advertising","printing","security","janitorial","facilities"],
    "automotive":      ["automotive","auto","car","truck","fleet","repair","mechanic",
                        "collision","towing","tire","parts"],
    "education":       ["education","training","tutoring","school","learning","childcare",
                        "daycare","workforce training"],
    "retail":          ["retail","store","shop","ecommerce","e-commerce","wholesale"],
    "real_estate":     ["real estate","property management","realty","rental","commercial"],
    "energy":          ["energy","oil","gas","solar","utilities","environmental"],
    "chemicals":       ["chemicals","chemical","industrial chemicals","specialty chemicals"],
}

def _get_industry_cluster(ind: str) -> str | None:
    il = ind.lower().strip()
    for cluster, terms in INDUSTRY_CLUSTERS.items():
        for t in terms:
            if t in il or il in t:
                return cluster
    return None

def _industry_score(lead_industry: str, buyer_industries: list) -> tuple:
    """Returns (hard_fail, pts) — pts up to 35."""
    if not buyer_industries:
        return False, 15
    ll = lead_industry.lower().strip()
    lead_cluster = _get_industry_cluster(ll)
    for b_ind in buyer_industries:
        bl = b_ind.lower().strip()
        if bl == ll:
            return False, 35          # exact
        if bl in ll or ll in bl:
            return False, 30          # substring
        if lead_cluster and lead_cluster == _get_industry_cluster(bl):
            return False, 25          # same cluster
    return True, 0                    # hard fail

# Geographic region expansions
_GEO_EXPAND = {
    "southeast":  ["FL","GA","AL","SC","NC","TN","MS","AR","LA"],
    "southwest":  ["TX","AZ","NM","NV","CO","UT","OK"],
    "midwest":    ["OH","IL","IN","MI","WI","MN","IA","MO","KS","NE","SD","ND"],
    "northeast":  ["NY","NJ","PA","CT","MA","RI","VT","NH","ME","MD","DE","DC"],
    "west":       ["CA","OR","WA","ID","MT","WY","AK","HI"],
    "south":      ["TX","FL","GA","AL","SC","NC","TN","MS","AR","LA","OK"],
    "mid-atlantic":["NY","NJ","PA","MD","DE","DC","VA"],
    "new england":["MA","CT","RI","VT","NH","ME"],
    "mountain":   ["CO","UT","NV","ID","MT","WY","AZ","NM"],
    "plains":     ["KS","NE","SD","ND","MN","IA","MO"],
    "sunbelt":    ["FL","TX","AZ","GA","NC","SC","NV","TN","AL"],
}

def _expand_buyer_states(buyer: dict) -> set:
    states = {s.strip().upper()[:2] for s in (buyer.get("states") or []) if s}
    for geo in (buyer.get("geographies") or []):
        gl = geo.lower()
        if any(x in gl for x in ("national","nationwide","all","us")):
            return set()              # no restriction
        for region, st_list in _GEO_EXPAND.items():
            if region in gl:
                states.update(st_list)
    return states

def _parse_dollars(val) -> float:
    if not val:
        return 0.0
    s = str(val).replace(",","").replace("$","").replace(" ","").upper()
    mult = 1.0
    if s.endswith("M"):   mult = 1_000_000;   s = s[:-1]
    elif s.endswith("K"): mult = 1_000;        s = s[:-1]
    elif s.endswith("B"): mult = 1_000_000_000; s = s[:-1]
    try:
        return float(s) * mult
    except Exception:
        return 0.0


def score_lead(lead: dict, buyer: dict) -> int:
    """
    Strict match scoring with hard gates.
    Returns 0 if any hard gate fails (wrong industry, wrong state, revenue way out of range).
    Max score: 100 pts.
    """
    score = 0

    # ── HARD GATE 1: Industry ─────────────────────────────────────────────────
    lead_industry = (lead.get("industry") or "").strip()
    buyer_industries = [i for i in (buyer.get("industries") or []) if i]
    hard_fail, ind_pts = _industry_score(lead_industry, buyer_industries)
    if hard_fail:
        return 0
    score += ind_pts  # 0–35 pts

    # ── HARD GATE 2: Location ─────────────────────────────────────────────────
    lead_state = (lead.get("state") or "").strip().upper()[:2]
    expanded = _expand_buyer_states(buyer)
    if expanded:                          # buyer has location preference
        if lead_state and lead_state not in expanded:
            return 0                      # wrong state — hard fail
        score += 25 if lead_state else 0  # known + matching state
    else:
        score += 15                       # national buyer — partial credit

    # ── HARD GATE 3: Revenue range ────────────────────────────────────────────
    lead_rev = _parse_dollars(
        lead.get("estimated_revenue") or lead.get("revenue_estimate") or
        lead.get("asking_price") or ""
    )
    b_min = float(buyer.get("deal_size_min") or 0)
    b_max = float(buyer.get("deal_size_max") or 0)

    if (b_min or b_max) and lead_rev > 0:
        if b_max and lead_rev > b_max * 2.5:
            return 0   # way too large
        if b_min and lead_rev < b_min * 0.25:
            return 0   # way too small
        if (not b_min or lead_rev >= b_min) and (not b_max or lead_rev <= b_max):
            score += 20  # perfect range fit
        else:
            score += 8   # close but not exact
    elif not b_min and not b_max:
        score += 10  # no revenue filter → partial credit

    # ── SOFT: EBITDA / cash flow ──────────────────────────────────────────────
    lead_ebitda = _parse_dollars(lead.get("ebitda_estimate") or lead.get("ebitda") or "")
    e_min = float(buyer.get("ebitda_min") or 0)
    e_max = float(buyer.get("ebitda_max") or 0)
    if lead_ebitda > 0 and (e_min or e_max):
        if (not e_min or lead_ebitda >= e_min) and (not e_max or lead_ebitda <= e_max):
            score += 10
        else:
            score += 3
    elif lead_ebitda > 0:
        score += 5

    # ── SOFT: Business age ────────────────────────────────────────────────────
    try:
        age = 2026 - int(str(lead.get("founded_year",""))[:4])
        if age >= 15:   score += 8
        elif age >= 10: score += 5
        elif age >= 5:  score += 2
    except Exception:
        pass

    # ── SOFT: Contact data richness ───────────────────────────────────────────
    if lead.get("owner_name"):   score += 4
    if lead.get("owner_email"):  score += 4
    if lead.get("owner_phone") or lead.get("phone"): score += 2
    if lead.get("company_domain") or lead.get("website"): score += 2

    return min(score, 100)


def days_since(dt_str: str) -> int:
    if not dt_str:
        return 9999
    try:
        dt = datetime.fromisoformat(dt_str[:19])
        return (datetime.now() - dt).days
    except Exception:
        return 9999


def load_cached_leads(buyer_id: str) -> list:
    path = LEADS_DIR / f"leads_{buyer_id}.json"
    if path.exists():
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return []


def rescore_all_deals() -> tuple:
    """Re-score every deal in the DB against its matched buyer. Returns (updated, zeroed)."""
    import sqlite3
    all_d = get_all_deals()
    all_b = {b["id"]: b for b in get_all_buyers()}
    updated = zeroed = 0
    for deal in all_d:
        buyer = all_b.get(deal.get("buyer_id"), {})
        new_score = score_lead(deal, buyer)
        new_interest = ("hot" if new_score >= 70 else "warm" if new_score >= 40 else "cold")
        update_deal(deal["id"], {"match_score": new_score, "interest_level": new_interest})
        if new_score == 0:
            zeroed += 1
        else:
            updated += 1
    return updated, zeroed


def purge_zero_score_deals() -> int:
    """Delete all deals with match_score = 0."""
    all_d = get_all_deals()
    count = 0
    for deal in all_d:
        if deal.get("match_score", 0) == 0:
            delete_deal(deal["id"])
            count += 1
    return count


def enrich_lead_serp(deal: dict) -> dict:
    """
    Use SerpAPI to fill in missing owner name, revenue, phone, domain for a lead.
    Returns dict of fields to update (may be empty).
    """
    serpapi_key = os.getenv("SERPAPI_KEY", "")
    if not serpapi_key or serpapi_key in ("", "REPLACE_ME", "your-serpapi-key"):
        return {}
    import re as _re
    import requests as _req
    company = deal.get("company_name", "")
    state   = deal.get("state", "")
    if not company:
        return {}
    query = f'"{company}" {state} owner CEO founder revenue employees'
    try:
        resp = _req.get("https://serpapi.com/search",
                        params={"api_key": serpapi_key, "q": query, "num": 5, "engine": "google"},
                        timeout=10)
        data = resp.json()
    except Exception:
        return {}
    enriched = {}
    kg = data.get("knowledge_graph", {})
    if kg.get("phone") and not deal.get("phone"):
        enriched["phone"] = kg["phone"]
    if kg.get("website") and not deal.get("company_domain"):
        enriched["company_domain"] = kg["website"]
    if kg.get("founded") and not deal.get("founded_year"):
        enriched["founded_year"] = str(kg["founded"])[:4]
    # Scan snippets for owner name and revenue
    owner_patterns = [r'(?:CEO|owner|founder|president)[,:\s]+([A-Z][a-z]+ [A-Z][a-z]+)']
    rev_patterns   = [r'\$([\d,.]+)\s*(?:million|M)\s*(?:in )?(?:revenue|sales|annual)',
                      r'(?:revenue|sales)[:\s]+\$([\d,.]+)\s*(?:million|M|K)?']
    for result in data.get("organic_results", [])[:4]:
        snippet = result.get("snippet") or result.get("title") or ""
        if not deal.get("owner_name"):
            for pat in owner_patterns:
                m = _re.search(pat, snippet, _re.IGNORECASE)
                if m:
                    enriched["owner_name"] = m.group(1)
                    break
        if not deal.get("revenue_estimate"):
            for pat in rev_patterns:
                m = _re.search(pat, snippet, _re.IGNORECASE)
                if m:
                    enriched["revenue_estimate"] = f"${m.group(1)}M"
                    break
    return enriched


def _clean_domain(raw: str) -> str:
    """Strip protocol + path from a URL to get just the domain."""
    import re as _re
    s = (raw or "").strip().lower()
    s = _re.sub(r"^https?://", "", s)
    s = s.split("/")[0].split("?")[0]
    s = s.replace("www.", "")
    return s

def enrich_emails_apollo(deal: dict) -> dict:
    """
    Use Apollo.io people search to find owner email + phone for a lead with a website.
    Apollo is much higher quality + cheaper than Hunter for this use case.
    Returns dict of fields to update.
    """
    apollo_key = os.getenv("APOLLO_API_KEY", "")
    if not apollo_key or apollo_key in ("", "REPLACE_ME", "your-apollo-api-key"):
        return {}
    raw_domain = deal.get("company_domain", "")
    if not raw_domain:
        return {}
    if any(d in raw_domain.lower() for d in ("yelp.com", "manta.com", "yellowpages.com", "bbb.org", "file://")):
        return {}  # Skip directory URLs and local files
    domain = _clean_domain(raw_domain)
    if not domain or "." not in domain:
        return {}
    import requests as _req
    # Search for senior people at the company domain
    try:
        resp = _req.post(
            "https://api.apollo.io/v1/mixed_people/search",
            headers={"X-Api-Key": apollo_key, "Content-Type": "application/json"},
            json={
                "q_organization_domains_list": [domain],
                "person_seniorities": ["c_suite", "owner", "founder", "vp"],
                "page": 1, "per_page": 5,
            },
            timeout=15,
        )
        data = resp.json()
    except Exception:
        return {}
    people = data.get("people") or []
    if not people:
        return {}
    # Prefer one with a verified email
    people.sort(key=lambda p: 0 if p.get("email_status") == "verified" else 1)
    pick = people[0]
    out = {}
    if pick.get("email"):
        out["owner_email"] = pick["email"]
    name = pick.get("name") or f"{pick.get('first_name','')} {pick.get('last_name','')}".strip()
    if name and not deal.get("owner_name"):
        out["owner_name"] = name
    if pick.get("phone_numbers") and not deal.get("owner_phone"):
        ph = pick["phone_numbers"][0].get("sanitized_number") or pick["phone_numbers"][0].get("raw_number")
        if ph:
            out["owner_phone"] = ph
    if pick.get("linkedin_url"):
        out["owner_linkedin"] = pick["linkedin_url"]
    return out


def enrich_via_clay_webhook(deal: dict) -> dict:
    """
    Push a lead to Clay.com via webhook. Clay handles enrichment async
    and writes back to our DB (requires Clay table + webhook return).
    For now this is a fire-and-forget push — Clay will write back later
    via a separate inbound webhook handler we'll add.
    """
    clay_webhook = os.getenv("CLAY_WEBHOOK_URL", "")
    if not clay_webhook:
        return {}
    import requests as _req
    try:
        _req.post(
            clay_webhook,
            json={
                "deal_id":      deal.get("id"),
                "company_name": deal.get("company_name"),
                "domain":       _clean_domain(deal.get("company_domain","")),
                "city":         deal.get("city",""),
                "state":        deal.get("state",""),
                "industry":     deal.get("industry",""),
            },
            timeout=8,
        )
        return {"clay_pushed": "yes"}
    except Exception:
        return {}


def top_n_leads(leads: list, buyer: dict, n: int = 3) -> list:
    for lead in leads:
        if not lead.get("match_score"):
            lead["match_score"] = score_lead(lead, buyer)
    return sorted(leads, key=lambda x: x.get("match_score", 0), reverse=True)[:n]


def create_gmail_draft(to: str, subject: str, body: str) -> tuple:
    if not SMTP_USER or not SMTP_PASS:
        return False, "SMTP credentials not set in .env (SMTP_USER / SMTP_PASSWORD)"
    try:
        msg = MIMEMultipart("alternative")
        msg["To"]      = to
        msg["From"]    = SMTP_USER
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(SMTP_USER, SMTP_PASS)
        imap.append("[Gmail]/Drafts", "", imaplib.Time2Internaldate(time.time()), msg.as_bytes())
        imap.logout()
        return True, "Draft created in Gmail."
    except Exception as e:
        return False, str(e)


def sync_hardcoded_buyers():
    existing = get_all_buyers()
    if existing:
        return len(existing)
    for b in _buyers_db():
        upsert_buyer(b)
    return len(_buyers_db())


def import_buyers_from_excel(uploaded_file) -> int:
    try:
        df = pd.read_excel(uploaded_file, sheet_name=0)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        count = 0
        for _, row in df.iterrows():
            name = str(row.get("name") or row.get("buyer_name") or row.get("company") or "").strip()
            if not name or name.lower() == "nan":
                continue
            def _list(val):
                if pd.isna(val) if hasattr(pd, 'isna') else (val != val):
                    return []
                return [x.strip() for x in str(val).split(",") if x.strip()]
            def _float(val, default=0.0):
                try:
                    return float(val) if not (val != val) else default
                except Exception:
                    return default
            upsert_buyer({
                "id":            name.lower().replace(" ", "_").replace("/", "_")[:40],
                "name":          name,
                "contact_name":  str(row.get("contact_name") or row.get("contact") or ""),
                "email":         str(row.get("email") or ""),
                "phone":         str(row.get("phone") or ""),
                "industries":    _list(row.get("industries") or row.get("industry")),
                "deal_size_min": _float(row.get("deal_size_min") or row.get("min_deal_size")),
                "deal_size_max": _float(row.get("deal_size_max") or row.get("max_deal_size")),
                "ebitda_min":    _float(row.get("ebitda_min") or row.get("min_ebitda")),
                "ebitda_max":    _float(row.get("ebitda_max") or row.get("max_ebitda")),
                "states":        _list(row.get("states") or row.get("state")),
                "geographies":   _list(row.get("geographies") or row.get("geography")),
                "notes":         str(row.get("notes") or ""),
                "broker":        str(row.get("broker") or ""),
            })
            count += 1
        return count
    except Exception as e:
        st.error(f"Excel import error: {e}")
        return 0


# ── email templates ───────────────────────────────────────────────────────────

def _default_outreach_email(lead: dict, buyer: dict) -> str:
    company  = lead.get("company_name", "your company")
    owner    = lead.get("owner_name", "")
    greeting = f"Dear {owner.split()[0]}," if owner else "Hello,"
    industries = ", ".join((buyer.get("industries") or [])[:2])
    geos = ", ".join((buyer.get("geographies") or buyer.get("states") or ["the US"])[:2])
    return f"""{greeting}

My name is {BROKER_NAME} — {BROKER_TITLE} at {BROKER_COMPANY}. I represent a \
well-capitalized buyer actively acquiring {industries} businesses in {geos}.

{company} came up in our search and fits their profile precisely. They offer \
clean, straightforward transactions with structures that honor what you've built \
— management continuity and cultural fit are priorities.

If you've considered an exit or partial liquidity event, even 2-3 years out, \
I'd welcome a brief 15-minute call. No obligation, strictly confidential.

{BROKER_NAME}
{BROKER_TITLE} | {BROKER_COMPANY}
{BROKER_PHONE} | {BROKER_EMAIL}"""


def _nda_email(deal: dict) -> tuple:
    owner   = deal.get("owner_name", "")
    company = deal.get("company_name", "the business")
    greeting = f"Dear {owner.split()[0]}," if owner else "Hello,"
    subject = f"Confidentiality Agreement — {company} Business Review"
    body = f"""{greeting}

Thank you for your openness to discussing a potential transaction. Before we proceed \
to share detailed buyer information, our qualified buyer requires a standard \
Non-Disclosure Agreement (NDA).

This is a mutual, standard NDA that protects both parties. It simply ensures all \
shared information remains strictly confidential during our evaluation process.

Please reply to this email confirming your willingness to execute the NDA, and I \
will send it via DocuSign immediately. The entire process takes less than 5 minutes.

Once signed, I'll send you the buyer's full profile and we can schedule an \
introductory call at your convenience.

Looking forward to moving this forward,

{BROKER_NAME}
{BROKER_TITLE} | {BROKER_COMPANY}
{BROKER_PHONE} | {BROKER_EMAIL}"""
    return subject, body


def _cim_email(deal: dict, buyer: dict) -> tuple:
    buyer_name    = buyer.get("name", "the buyer") if buyer else deal.get("buyer_name", "the buyer")
    contact_name  = buyer.get("contact_name", "") if buyer else ""
    greeting      = f"Dear {contact_name.split()[0]}," if contact_name else f"Dear {buyer_name},"
    company  = deal.get("company_name", "the business")
    industry = deal.get("industry", "")
    state    = deal.get("state", "")
    revenue  = deal.get("revenue_estimate", "")
    ebitda   = deal.get("ebitda_estimate", "")
    asking   = deal.get("asking_price", "")
    subject = f"Confidential Deal Summary — {industry} Business | {state}"
    body = f"""{greeting}

Following execution of the NDA, please find below the summary for a new acquisition \
opportunity that aligns with your mandate:

  Business:     {company}
  Industry:     {industry or '—'}
  Location:     {state or '—'}
  Revenue:      {revenue or '—'}
  EBITDA:       {ebitda or '—'}
  Asking Price: {asking or 'TBD'}

This opportunity is being presented on a confidential, off-market basis to a \
select group of qualified buyers. Seller is motivated and timeline is flexible.

Please review and let me know your level of interest. If you'd like to proceed, \
I can arrange an introductory call with the seller within 48 hours.

{BROKER_NAME}
{BROKER_TITLE} | {BROKER_COMPANY}
{BROKER_PHONE} | {BROKER_EMAIL}"""
    return subject, body


def _linkedin_message(deal: dict, buyer: dict) -> str:
    company  = deal.get("company_name", "a business")
    industry = deal.get("industry", "")
    state    = deal.get("state", "")
    buyer_name = buyer.get("contact_name") or buyer.get("name", "")
    first = buyer_name.split()[0] if buyer_name else "there"
    return f"""Hi {first},

I have a confidential off-market {industry} business in {state} that matches your acquisition criteria closely.

Revenue: {deal.get('revenue_estimate', 'TBD')} | EBITDA: {deal.get('ebitda_estimate', 'TBD')}

Seller is motivated. Happy to share more details under NDA. Interested in a quick call?

— {BROKER_NAME}, {BROKER_COMPANY}"""


def email_subject(lead: dict, buyer: dict) -> str:
    industry = (lead.get("industry") or "").title() or "Business"
    geo = (lead.get("state") or (buyer.get("states") or ["US"])[0])
    return f"Confidential — Qualified Buyer Interested in {geo} {industry} Companies"


def score_color(score: int) -> str:
    if score >= 70: return "#16A085"
    if score >= 40: return "#F39C12"
    return "#95A5A6"


# ── Phone area code → US state lookup ─────────────────────────────────────────
_AREA_STATE = {
    "205":"AL","251":"AL","256":"AL","334":"AL","659":"AL","938":"AL",
    "907":"AK","479":"AR","501":"AR","870":"AR",
    "480":"AZ","520":"AZ","602":"AZ","623":"AZ","928":"AZ",
    "209":"CA","213":"CA","279":"CA","310":"CA","323":"CA","341":"CA",
    "408":"CA","415":"CA","424":"CA","442":"CA","510":"CA","530":"CA",
    "559":"CA","562":"CA","619":"CA","626":"CA","628":"CA","650":"CA",
    "657":"CA","661":"CA","669":"CA","707":"CA","714":"CA","747":"CA",
    "760":"CA","764":"CA","805":"CA","818":"CA","820":"CA","831":"CA",
    "858":"CA","909":"CA","916":"CA","925":"CA","949":"CA","951":"CA",
    "303":"CO","719":"CO","720":"CO","970":"CO",
    "203":"CT","475":"CT","860":"CT","959":"CT",
    "302":"DE",
    "239":"FL","305":"FL","321":"FL","352":"FL","386":"FL","407":"FL",
    "448":"FL","561":"FL","567":"FL","689":"FL","727":"FL","754":"FL",
    "772":"FL","786":"FL","813":"FL","850":"FL","863":"FL","904":"FL",
    "941":"FL","954":"FL",
    "229":"GA","404":"GA","470":"GA","478":"GA","678":"GA","706":"GA",
    "762":"GA","770":"GA","912":"GA","943":"GA",
    "808":"HI","369":"HI",
    "208":"ID","986":"ID",
    "217":"IL","224":"IL","309":"IL","312":"IL","331":"IL","447":"IL",
    "464":"IL","618":"IL","630":"IL","708":"IL","730":"IL","773":"IL",
    "779":"IL","815":"IL","847":"IL","872":"IL",
    "219":"IN","260":"IN","317":"IN","463":"IN","574":"IN","765":"IN",
    "812":"IN","930":"IN",
    "319":"IA","515":"IA","563":"IA","641":"IA","712":"IA",
    "316":"KS","620":"KS","785":"KS","913":"KS",
    "270":"KY","364":"KY","502":"KY","606":"KY","859":"KY",
    "225":"LA","318":"LA","337":"LA","504":"LA","985":"LA",
    "207":"ME",
    "240":"MD","301":"MD","410":"MD","443":"MD","667":"MD",
    "339":"MA","351":"MA","413":"MA","508":"MA","617":"MA","774":"MA",
    "781":"MA","857":"MA","978":"MA",
    "231":"MI","248":"MI","269":"MI","313":"MI","517":"MI","586":"MI",
    "616":"MI","679":"MI","734":"MI","810":"MI","906":"MI","947":"MI",
    "989":"MI",
    "218":"MN","320":"MN","507":"MN","612":"MN","651":"MN","763":"MN",
    "952":"MN",
    "228":"MS","601":"MS","662":"MS","769":"MS",
    "314":"MO","417":"MO","573":"MO","636":"MO","660":"MO","816":"MO",
    "406":"MT",
    "308":"NE","402":"NE","531":"NE",
    "702":"NV","725":"NV","775":"NV",
    "603":"NH",
    "201":"NJ","551":"NJ","609":"NJ","640":"NJ","732":"NJ","848":"NJ",
    "856":"NJ","862":"NJ","908":"NJ","973":"NJ",
    "505":"NM","575":"NM",
    "212":"NY","315":"NY","332":"NY","347":"NY","516":"NY","518":"NY",
    "585":"NY","607":"NY","624":"NY","631":"NY","646":"NY","680":"NY",
    "716":"NY","718":"NY","725":"NV","838":"NY","845":"NY","914":"NY",
    "917":"NY","929":"NY","934":"NY",
    "252":"NC","336":"NC","472":"NC","704":"NC","743":"NC","828":"NC",
    "910":"NC","919":"NC","980":"NC","984":"NC",
    "701":"ND",
    "216":"OH","220":"OH","234":"OH","283":"OH","326":"OH","330":"OH",
    "380":"OH","419":"OH","436":"OH","440":"OH","513":"OH","567":"OH",
    "614":"OH","740":"OH","937":"OH",
    "405":"OK","539":"OK","572":"OK","580":"OK","918":"OK",
    "458":"OR","503":"OR","541":"OR","971":"OR",
    "215":"PA","223":"PA","267":"PA","272":"PA","412":"PA","445":"PA",
    "484":"PA","570":"PA","582":"PA","610":"PA","717":"PA","724":"PA",
    "814":"PA","835":"PA","878":"PA",
    "401":"RI",
    "803":"SC","839":"SC","843":"SC","854":"SC","864":"SC",
    "605":"SD",
    "423":"TN","615":"TN","629":"TN","731":"TN","865":"TN","901":"TN",
    "931":"TN",
    "210":"TX","214":"TX","254":"TX","281":"TX","325":"TX","346":"TX",
    "361":"TX","409":"TX","430":"TX","432":"TX","469":"TX","512":"TX",
    "682":"TX","713":"TX","726":"TX","737":"TX","806":"TX","817":"TX",
    "830":"TX","832":"TX","903":"TX","915":"TX","936":"TX","940":"TX",
    "945":"TX","956":"TX","972":"TX","979":"TX",
    "385":"UT","435":"UT","801":"UT",
    "802":"VT",
    "276":"VA","434":"VA","540":"VA","571":"VA","703":"VA","757":"VA",
    "804":"VA","826":"VA","948":"VA",
    "206":"WA","253":"WA","360":"WA","425":"WA","509":"WA","564":"WA",
    "304":"WV","681":"WV",
    "262":"WI","274":"WI","414":"WI","534":"WI","608":"WI","715":"WI",
    "920":"WI",
    "307":"WY",
}

def _infer_state_from_phone(phone: str) -> str:
    digits = "".join(c for c in str(phone or "") if c.isdigit())
    if len(digits) >= 10:
        area = digits[-10:-7]  # last 10 digits → first 3 = area code
        return _AREA_STATE.get(area, "")
    return ""


def enrich_and_rescore_all() -> tuple[int, int]:
    """Enrich missing state from phone, fill revenue from asking_price, rescore all.
    Returns (enriched_count, rescored_count)."""
    all_buyers = get_all_buyers()
    buyer_map  = {b["id"]: b for b in all_buyers}
    all_deals  = get_all_deals()
    enriched   = 0
    rescored   = 0

    for deal in all_deals:
        updates = {}
        did     = deal["id"]

        # Infer state from phone area code
        if not deal.get("state") and (deal.get("owner_phone") or deal.get("phone")):
            phone   = deal.get("owner_phone") or deal.get("phone", "")
            inferred = _infer_state_from_phone(phone)
            if inferred:
                updates["state"] = inferred
                enriched += 1

        # Estimate revenue from asking price (2× multiple is conservative floor)
        if not deal.get("revenue_estimate") and deal.get("asking_price"):
            ask = _parse_dollars(deal["asking_price"])
            if ask > 0:
                est = ask * 2.0
                updates["revenue_estimate"] = f"~${est:,.0f} (est.)"

        # Estimate EBITDA if revenue known but EBITDA missing (~15% margin)
        rev_val = _parse_dollars(
            updates.get("revenue_estimate") or deal.get("revenue_estimate", "")
        )
        if rev_val > 0 and not deal.get("ebitda_estimate"):
            updates["ebitda_estimate"] = f"~${rev_val * 0.15:,.0f} (est.)"

        if updates:
            update_deal(did, updates)
            deal.update(updates)

        # Rescore with enriched data
        buyer = buyer_map.get(deal.get("buyer_id", ""), {})
        new_score = score_lead(deal, buyer)
        if new_score != deal.get("match_score", -1):
            interest = "hot" if new_score >= 80 else "warm" if new_score >= 50 else "cold"
            update_deal(did, {"match_score": new_score, "interest_level": interest})
            rescored += 1

    return enriched, rescored


# ── page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Mithril — Deal Flow",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
  .block-container { padding-top: 1.2rem; }
  .stTabs [data-baseweb="tab-list"] { gap: 6px; }
  .stTabs [data-baseweb="tab"] {
      padding: 8px 20px; border-radius: 6px 6px 0 0;
      font-weight: 600; font-size: 0.93rem;
  }
  .lead-card {
      border-left: 5px solid #4A90D9;
      border-radius: 0 10px 10px 0;
      padding: 14px 18px;
      margin-bottom: 10px;
      background: #fafafa;
  }
  .lead-card-hot       { border-left-color: #E74C3C !important; background: #fff8f8 !important; }
  .lead-card-warm      { border-left-color: #F39C12 !important; background: #fffdf0 !important; }
  .lead-card-cold      { border-left-color: #3498DB !important; }
  /* Listing type */
  .card-onmarket       { border-left-color: #27AE60 !important; background: #f0fff4 !important; }
  .card-offmarket      { border-left-color: #8E44AD !important; background: #fdf4ff !important; }
  .card-perfect        { border-left-color: #F1C40F !important; background: #fffef0 !important;
                          box-shadow: 0 0 8px rgba(241,196,15,0.4); }
  .type-pill-on  { display:inline-block;padding:2px 10px;border-radius:12px;font-size:0.72rem;
                   font-weight:700;background:#27AE60;color:white;margin-right:4px; }
  .type-pill-off { display:inline-block;padding:2px 10px;border-radius:12px;font-size:0.72rem;
                   font-weight:700;background:#8E44AD;color:white;margin-right:4px; }
  .perfect-badge { display:inline-block;padding:2px 10px;border-radius:12px;font-size:0.72rem;
                   font-weight:700;background:#F1C40F;color:#333;margin-right:4px; }
  .badge {
      display: inline-block; padding: 2px 9px;
      border-radius: 12px; font-weight: 700;
      font-size: 0.78rem; color: white; margin-right: 4px;
  }
  .stage-step {
      display: inline-block; padding: 3px 10px;
      border-radius: 4px; font-size: 0.75rem;
      font-weight: 600; color: white; margin: 2px;
  }
  .intake-banner {
      background: #FEF3C7; border: 1px solid #F59E0B;
      border-radius: 8px; padding: 10px 16px; margin-bottom: 12px;
  }
</style>
""", unsafe_allow_html=True)

# ── startup: seed buyers from hardcoded list if DB is empty ───────────────────
sync_hardcoded_buyers()

# ── header ────────────────────────────────────────────────────────────────────

stats = get_pipeline_stats()

# Contact quality stat
_all_for_stats = get_all_deals()
_full_contact = sum(
    1 for d in _all_for_stats
    if (d.get("owner_email") or "").strip()
    and ((d.get("owner_phone") or d.get("phone") or "").strip())
    and (d.get("company_domain") or "").strip()
)

st.markdown(
    '<div style="display:flex;align-items:center;gap:14px;margin-bottom:2px">'
    '<span style="font-size:2.2rem">💎</span>'
    '<div>'
    '<div style="font-size:1.75rem;font-weight:800;letter-spacing:-0.5px;'
    'background:linear-gradient(90deg,#c9d6e8 0%,#7fa8d0 50%,#4a90d9 100%);'
    '-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text">'
    'Mithril <span style="color:#888;font-weight:400">— Deal Flow Forge</span>'
    '</div>'
    f'<div style="color:#666;font-size:0.78rem;letter-spacing:0.5px">'
    f'⛏ Off-market deal sourcing for {BROKER_COMPANY} · {BROKER_NAME} · {BROKER_PHONE}'
    '</div>'
    '</div></div>',
    unsafe_allow_html=True
)
st.markdown(
    '<div style="color:#4a5060;font-size:0.72rem;font-style:italic;margin:6px 0 12px 0;letter-spacing:0.3px">'
    '"All that is gold does not glitter, not all those who wander are lost — but every great deal leaves a trail." '
    '— Mithril, the rarest of finds'
    '</div>',
    unsafe_allow_html=True
)

s1, s2, s3, s4, s5, s6, s7 = st.columns(7)
s1.metric("📋 Leads",        stats["total"],              help="Total leads & matches in DB")
s2.metric("✅ Contactable",  _full_contact,               help="Leads with email + phone + website")
s3.metric("🔥 Hot",          stats["hot"],                help="Leads marked Hot")
s4.metric("📝 NDA Signed",   stats["nda_signed"],         help="Deals with NDA executed")
s5.metric("📄 Under LOI",    stats["under_contract"],     help="Under Contract / LOI")
s6.metric("✅ Closed",       stats["closed"],             help="Closed deals")
s7.metric("🆕 Intakes",      stats.get("new_intakes",0),  help="New seller intakes")

st.divider()

# ── tabs ──────────────────────────────────────────────────────────────────────

tab7, tab1, tab2, tab6, tab3, tab4, tab5 = st.tabs([
    "⚒  The Forge",
    "📋  The Mines",
    "🎯  The Pipeline",
    "🚀  Scout",
    "📜  Ravens",
    "🔮  Palantír",
    "👑  Fellowship",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DEAL FLOW STREAM
# ══════════════════════════════════════════════════════════════════════════════

# ── helpers for Excel imports ──────────────────────────────────────────────

def _import_buyers_merged(path) -> int:
    import openpyxl, io
    if hasattr(path, "read"):
        wb = openpyxl.load_workbook(io.BytesIO(path.read()))
    else:
        wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # Row 3 = headers, Row 4 = instructions, Row 5+ = data
    raw_headers = [str(c.value or "").strip() for c in ws[3]]
    count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue
        vals = [str(v or "").strip() for v in row]
        name = vals[1] if len(vals) > 1 else ""
        if not name or name.startswith("First") or name.startswith("Auto"):
            continue
        def _sp(idx):
            return [x.strip() for x in vals[idx].replace(",",";").split(";") if x.strip()] if len(vals) > idx else []
        def _num(idx):
            try: return float(str(vals[idx]).replace(",","").replace("$","")) if len(vals) > idx and vals[idx] else 0.0
            except: return 0.0
        buyer_id = (vals[0] or name.lower().replace(" ","_"))[:40]
        upsert_buyer({
            "id": buyer_id, "name": name,
            "contact_name": name,
            "email":    vals[3] if len(vals) > 3 else "",
            "phone":    vals[4] if len(vals) > 4 else "",
            "industries": _sp(6),
            "states":     _sp(8),
            "deal_size_min": _num(9),
            "deal_size_max": _num(10),
            "ebitda_min":    _num(13),
            "notes":    vals[11] if len(vals) > 11 else "",
            "broker":   "Jordi Quevedo-Valls",
        })
        count += 1
    return count


def _import_seller_offmarket(path) -> int:
    import openpyxl, io
    if hasattr(path, "read"):
        wb = openpyxl.load_workbook(io.BytesIO(path.read()))
    else:
        wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # Row 3 = headers, Row 4+ = data
    headers = [str(c.value or "").strip().lower() for c in ws[3]]
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = [str(v or "").strip() for v in row]
        if not any(vals) or vals[0].startswith("SEL_") is False:
            try:
                if not vals[1]: continue
            except: continue
        title = vals[1] if len(vals) > 1 else ""
        if not title or title.startswith("Business Title"):
            continue
        def _num(idx):
            try: return float(str(vals[idx]).replace(",","").replace("$","")) if len(vals) > idx and vals[idx] else 0
            except: return 0
        industry = vals[4] if len(vals) > 4 else ""
        state    = vals[6] if len(vals) > 6 else ""
        asking   = _num(7)
        revenue  = _num(8)
        cf       = _num(9)

        if deal_exists("__offmarket__", title):
            continue

        # Score against all buyers and find best
        all_byr = get_all_buyers()
        best_buyer = {}
        best_sc = 0
        for b in all_byr:
            sc = score_lead({"industry": industry, "state": state, "revenue_estimate": str(revenue)}, b)
            if sc > best_sc:
                best_sc, best_buyer = sc, b

        add_deal({
            "buyer_id":       best_buyer.get("id","__offmarket__"),
            "buyer_name":     best_buyer.get("name","Unmatched"),
            "company_name":   title,
            "industry":       industry,
            "city":           vals[5] if len(vals) > 5 else "",
            "state":          state,
            "asking_price":   f"${asking:,.0f}" if asking else "",
            "revenue_estimate": f"${revenue:,.0f}" if revenue else "",
            "ebitda_estimate":  f"${cf:,.0f}" if cf else "",
            "source":         vals[2] if len(vals) > 2 else "pdf_extract",
            "company_domain": vals[3] if len(vals) > 3 else "",
            "match_score":    best_sc,
            "interest_level": "hot" if best_sc >= 80 else "warm" if best_sc >= 50 else "cold",
            "listing_type":   "off-market",
            "deal_stage":     "identified",
            "status":         "identified",
        })
        count += 1
    return count


with tab1:

    st.markdown(
        '<div style="background:#0d1421;border-left:3px solid #4A90D9;padding:8px 14px;'
        'border-radius:0 6px 6px 0;margin-bottom:6px;font-size:0.85rem;color:#aab">'
        '<span style="color:#F39C12;font-weight:700">📦 Sellers</span> are off-market businesses scraped or imported. '
        '<span style="color:#9B59B6;font-weight:700">🤝 Matches</span> are seller×buyer pairs ready to work. '
        '<span style="color:#27AE60;font-weight:700">✉ Already Reached Out</span> are matches we have outreach drafted/sent on. '
        'Manage <strong style="color:#0077b5">👥 Buyers</strong> in Settings tab.'
        '</div>',
        unsafe_allow_html=True
    )

    all_buyers = get_all_buyers()
    _bnames1   = ["All Buyers"] + [b["name"] for b in all_buyers]

    # ── seller / match / all toggle ──────────────────────────────────────
    # Stats for each bucket
    _all_for_view = get_all_deals()
    _seller_count = sum(1 for d in _all_for_view if not (d.get("buyer_id") or "").strip())
    _match_count  = sum(1 for d in _all_for_view if (d.get("buyer_id") or "").strip())
    _contacted_ct = sum(1 for d in _all_for_view if (d.get("last_contacted") or "").strip() and (d.get("buyer_id") or "").strip())

    _vtoggle = st.radio(
        "View",
        [
            f"🤝 Matches ({_match_count})",
            f"✉ Reached Out ({_contacted_ct})",
            f"📦 Sellers Only ({_seller_count})",
            f"📊 All ({len(_all_for_view)})",
        ],
        horizontal=True,
        key="ls_view_toggle",
        label_visibility="collapsed",
        index=0,
    )
    if _vtoggle.startswith("📦"):
        f_view_mode = "sellers"
    elif _vtoggle.startswith("🤝"):
        f_view_mode = "matches"
    elif _vtoggle.startswith("✉"):
        f_view_mode = "reached_out"
    else:
        f_view_mode = "all"

    # ── search box ────────────────────────────────────────────────────────
    f_search = st.text_input(
        "🔎 Search",
        key="ls_search",
        placeholder="Search company, owner, industry, city, state…",
        label_visibility="collapsed"
    )

    CLUSTERS = {
        "All Industries": None,
        "🔧 HVAC / Home Services": ["hvac","plumbing","home services","electrical","landscaping"],
        "🏭 Manufacturing":        ["manufacturing","fabrication","industrial","aerospace","cnc"],
        "🚚 Transport / Logistics":["transport","logistics","freight","trucking","distribution"],
        "💻 Tech / Software":      ["software","technology","it","saas","digital"],
        "🏥 Healthcare":           ["healthcare","medical","dental","therapy","health"],
        "🍽️ Food / Hospitality":   ["food","restaurant","hospitality","catering"],
        "🏗️ Construction":         ["construction","contractor","excavation","civil"],
        "💼 Services":             ["services","consulting","staffing","accounting","finance"],
    }

    # ── single-row filter bar — fluid, no sliders ──────────────────────────
    fc1, fc2, fc3, fc4, fc5, fc6 = st.columns([3, 2, 2, 2, 2, 2])
    f_buyer    = fc1.selectbox("Buyer",    _bnames1, key="ls_buyer")
    f_interest = fc2.selectbox("Interest", ["All","Hot","Warm","Cold"], key="ls_interest")
    f_type     = fc3.selectbox("Type",     ["All","On-Market","Off-Market"], key="ls_type")
    f_stage    = fc4.selectbox("Stage",    ["All"] + [STAGE_LABELS[s] for s in DEAL_STAGES], key="ls_stage")
    f_cluster  = fc5.selectbox("Industry", list(CLUSTERS.keys()), key="ls_cluster")
    f_sort     = fc6.selectbox("Sort",     ["Interest + Newest","Highest Score","Newest Listed","Overdue Contact"], key="ls_sort")

    # Hidden filter values (no longer in UI but logic still uses them)
    f_min_score = 0
    f_contact   = "Any"
    f_perfect   = False

    # ── tools expander ────────────────────────────────────────────────────
    with st.expander("⚙️  Tools — Import · Scrape · Add Lead · Maintenance", expanded=False):
        _tool_tabs = st.tabs(["📥 Import", "🔄 Scrape", "➕ Add Lead", "🧹 Maintenance"])

        with _tool_tabs[0]:  # Import
            ti1, ti2, ti3 = st.columns(3)
            if ti1.button("Import BUYERS_MERGED.xlsx", key="imp_buyers_merged", use_container_width=True):
                p = ROOT / "data" / "BUYERS_MERGED.xlsx"
                if p.exists():
                    st.success(f"Imported {_import_buyers_merged(p)} buyers!")
                    st.rerun()
                else:
                    st.error("File not found: data/BUYERS_MERGED.xlsx")
            if ti2.button("Import Seller_OffMarket_Listings.xlsx", key="imp_sellers_om", use_container_width=True):
                p = ROOT / "Seller_OffMarket_Listings.xlsx"
                if p.exists():
                    st.success(f"Imported {_import_seller_offmarket(p)} deals!")
                    st.rerun()
                else:
                    st.error("File not found: Seller_OffMarket_Listings.xlsx")
            if ti3.button("🚀 Scrape All Buyers Now", key="imp_fresh_all", type="primary", use_container_width=True):
                targets = get_all_buyers()
                if not targets:
                    st.warning("No buyers configured.")
                else:
                    _prog = st.progress(0); _stat = st.empty(); _add = 0
                    for _i, _b in enumerate(targets):
                        _prog.progress(_i/len(targets), text=f"{_b['name']}…")
                        _ls = _apollo_search(_b, 10)
                        if len(_ls) < 2: _ls += _web_scraper(_b, 8)
                        for _l in _ls:
                            _l["match_score"] = score_lead(_l, _b)
                            if _l["match_score"] >= 40 and not deal_exists(_b["id"], _l.get("company_name","")):
                                add_deal({**_l,"buyer_id":_b["id"],"buyer_name":_b["name"],
                                          "status":"identified","deal_stage":"identified",
                                          "email_subject":email_subject(_l,_b),
                                          "email_body":_default_outreach_email(_l,_b)})
                                _add += 1
                    _prog.progress(1.0); st.success(f"Added {_add} new leads."); st.rerun()
            st.markdown("**Upload files**")
            _uc1, _uc2 = st.columns(2)
            _upl_s = _uc1.file_uploader("Seller listings (xlsx/csv)", type=["xlsx","csv"], key="upl_sellers")
            if _upl_s:
                st.success(f"Imported {_import_seller_offmarket(_upl_s)} listings!"); st.rerun()
            _upl_b = _uc2.file_uploader("Buyer mandates (xlsx/csv)", type=["xlsx","csv"], key="upl_buyers_tab1")
            if _upl_b:
                st.success(f"Imported {import_buyers_from_excel(_upl_b)} buyers!"); st.rerun()

        with _tool_tabs[1]:  # Scrape
            _ts1, _ts2, _ts3 = st.columns([3, 2, 2])
            _scrape_buyer = _ts1.selectbox("Buyer", _bnames1, key="ls_scrape_buyer")
            _min_s = _ts2.slider("Min score", 0, 100, 60, 5, key="ls_min_score")
            _max_r = _ts3.number_input("Max per source", 5, 50, 15, key="ls_max_res")
            if st.button("▶ Run Scrape", type="primary", key="ls_run_scrape", use_container_width=True):
                _targets = all_buyers if _scrape_buyer == "All Buyers" else [b for b in all_buyers if b["name"] == _scrape_buyer]
                _tadd = 0; _prog2 = st.progress(0)
                for _idx, _byr in enumerate(_targets):
                    _ls2 = _apollo_search(_byr, int(_max_r))
                    if len(_ls2) < 3:
                        _ex = {l["company_name"] for l in _ls2}
                        _ls2 += [l for l in _web_scraper(_byr, 5) if l.get("company_name") not in _ex]
                    for _l2 in _ls2:
                        _l2["match_score"] = score_lead(_l2, _byr)
                        if _l2["match_score"] >= _min_s and not deal_exists(_byr["id"], _l2.get("company_name","")):
                            add_deal({**_l2,"buyer_id":_byr["id"],"buyer_name":_byr["name"],
                                      "status":"identified","deal_stage":"identified",
                                      "email_subject":email_subject(_l2,_byr),
                                      "email_body":_default_outreach_email(_l2,_byr)})
                            _tadd += 1
                    _prog2.progress((_idx+1)/len(_targets))
                st.success(f"Added {_tadd} new leads."); st.rerun()

        with _tool_tabs[2]:  # Add Lead
            with st.form("manual_lead_form"):
                ml1, ml2, ml3 = st.columns(3)
                ml_company = ml1.text_input("Company Name *")
                ml_owner   = ml1.text_input("Owner Name")
                ml_industry= ml1.text_input("Industry")
                ml_buyer   = ml1.selectbox("Match to Buyer", [b["name"] for b in all_buyers] if all_buyers else ["(no buyers)"])
                ml_email   = ml2.text_input("Owner Email")
                ml_phone   = ml2.text_input("Owner Phone")
                ml_state   = ml2.text_input("State (2-letter)")
                ml_source  = ml2.text_input("Source", value="manual")
                ml_revenue = ml3.text_input("Revenue Estimate")
                ml_ebitda  = ml3.text_input("EBITDA Estimate")
                ml_asking  = ml3.text_input("Asking Price")
                ml_type    = ml3.selectbox("Listing Type", ["off-market","on-market"])
                ml_notes   = st.text_area("Notes", height=50)
                if st.form_submit_button("Add Lead", type="primary"):
                    if ml_company.strip():
                        _bo = next((b for b in all_buyers if b["name"] == ml_buyer), {})
                        _sc = score_lead({"industry":ml_industry,"state":ml_state,"owner_name":ml_owner,"owner_email":ml_email,"revenue_estimate":ml_revenue}, _bo)
                        add_deal({"company_name":ml_company,"owner_name":ml_owner,"owner_email":ml_email,"owner_phone":ml_phone,"industry":ml_industry,"state":ml_state,"revenue_estimate":ml_revenue,"ebitda_estimate":ml_ebitda,"asking_price":ml_asking,"listing_type":ml_type,"source":ml_source,"response_notes":ml_notes,"match_score":_sc,"buyer_id":_bo.get("id",""),"buyer_name":_bo.get("name",ml_buyer),"deal_stage":"identified","status":"identified","email_subject":email_subject({"industry":ml_industry,"state":ml_state},_bo),"email_body":_default_outreach_email({"company_name":ml_company,"industry":ml_industry,"state":ml_state,"owner_name":ml_owner},_bo)})
                        st.success(f"Added: {ml_company}"); st.rerun()
                    else:
                        st.error("Company name required.")

        with _tool_tabs[3]:  # Maintenance
            st.caption("⚡ Recommended workflow: 1) Find Emails (Hunter) →  2) Purge No-Contact Leads →  3) Enrich & Rescore")
            _m1, _m2, _m3, _m4, _m5 = st.columns(5)
            if _m1.button("🔄 Enrich & Rescore", type="primary", key="db_rescore", use_container_width=True,
                          help="Infer state from phone, estimate revenue from asking, rescore everything"):
                with st.spinner("Working…"):
                    _en, _re = enrich_and_rescore_all()
                st.success(f"Enriched {_en} · Rescored {_re}"); st.rerun()
            if _m2.button("🗑️ Purge Junk", type="primary", key="db_purge", use_container_width=True,
                          help="Delete leads with no name or 0 score"):
                with st.spinner("Purging…"):
                    _rm = 0
                    for _d in get_all_deals():
                        _nm = (_d.get("company_name") or "").strip()
                        if not _nm or _nm in ("—","-","Unknown","None") or _d.get("match_score",0)==0:
                            delete_deal(_d["id"]); _rm += 1
                st.success(f"Removed {_rm} junk leads."); st.rerun()
            if _m3.button("✉ Find Emails (Apollo)", key="db_apollo", type="primary", use_container_width=True,
                          help="Use Apollo.io people search to find verified owner email for every lead with a real website"):
                _ak = os.getenv("APOLLO_API_KEY","")
                if not _ak or _ak in ("REPLACE_ME","your-apollo-api-key"):
                    st.error("Add APOLLO_API_KEY in Settings.")
                else:
                    _need = [d for d in get_all_deals()
                             if (d.get("company_domain") or "").strip()
                             and not (d.get("owner_email") or "").strip()
                             and not any(skip in (d.get("company_domain") or "").lower()
                                         for skip in ("yelp.com","manta.com","yellowpages.com","bbb.org","file://"))][:300]
                    if not _need:
                        st.warning("No leads need Apollo enrichment.")
                    else:
                        _ph = st.progress(0); _ec = 0
                        for _i, _dl in enumerate(_need):
                            _upd = enrich_emails_apollo(_dl)
                            if _upd:
                                update_deal(_dl["id"], _upd); _ec += 1
                            _ph.progress((_i+1)/len(_need))
                        st.success(f"Apollo enriched {_ec}/{len(_need)} leads with verified contacts."); st.rerun()
            if _m4.button("✨ Enrich (SerpAPI)", key="db_enrich", use_container_width=True,
                          help="Fill in missing owner names + revenue via Google Search"):
                _sk = os.getenv("SERPAPI_KEY","")
                if not _sk or _sk == "REPLACE_ME":
                    st.error("Add SERPAPI_KEY in Settings.")
                else:
                    _thin = [d for d in get_all_deals() if not d.get("owner_name") or not d.get("revenue_estimate")][:50]
                    _p3 = st.progress(0); _ec = 0
                    for _i3, _dl in enumerate(_thin):
                        _upd = enrich_lead_serp(_dl)
                        if _upd: update_deal(_dl["id"], _upd); _ec += 1
                        _p3.progress((_i3+1)/len(_thin))
                    st.success(f"Enriched {_ec}/{len(_thin)} leads."); st.rerun()

            if _m5.button("☠ Delete No-Contact", key="db_purge_nocontact", type="primary", use_container_width=True,
                          help="Permanently delete every lead with no email AND no phone AND no website"):
                with st.spinner("Casting unworthy leads into Mount Doom…"):
                    _rm = 0
                    for _d in get_all_deals():
                        _has_em = bool((_d.get("owner_email") or "").strip())
                        _has_ph = bool((_d.get("owner_phone") or _d.get("phone") or "").strip())
                        _has_wb = bool((_d.get("company_domain") or "").strip())
                        if not (_has_em or _has_ph or _has_wb):
                            delete_deal(_d["id"]); _rm += 1
                st.success(f"☠ Cast {_rm} leads into the fire. Mithril remains."); st.rerun()

            # Second row of maintenance buttons
            st.markdown("")
            _mb1, _mb2 = st.columns(2)
            if _mb1.button("⚒ Push to Clay (Enrichment Waterfall)", key="db_clay_push", use_container_width=True,
                           help="Send leads with a website to Clay.com for multi-source enrichment"):
                _ck = os.getenv("CLAY_WEBHOOK_URL","")
                if not _ck or "your-webhook-id" in _ck:
                    st.error("Add CLAY_WEBHOOK_URL in Settings — sign up at clay.com first.")
                else:
                    _need = [d for d in get_all_deals()
                             if (d.get("company_domain") or "").strip()
                             and not any(skip in (d.get("company_domain") or "").lower()
                                         for skip in ("yelp.com","manta.com","yellowpages.com","file://"))
                             and not (d.get("owner_email") or "").strip()][:200]
                    if not _need:
                        st.warning("No leads need Clay enrichment.")
                    else:
                        _cp = st.progress(0); _csent = 0
                        for _i, _dl in enumerate(_need):
                            _r = enrich_via_clay_webhook(_dl)
                            if _r: _csent += 1
                            _cp.progress((_i+1)/len(_need))
                        st.success(f"⚒ Sent {_csent}/{len(_need)} leads to Clay. Enriched data will write back via webhook.")

    # ── build deal list ──
    buyer_id_filter = None
    if f_buyer != "All Buyers":
        bmap = {b["name"]: b["id"] for b in all_buyers}
        buyer_id_filter = bmap.get(f_buyer)

    interest_filter = None
    if f_interest == "Hot":   interest_filter = "hot"
    elif f_interest == "Warm": interest_filter = "warm"
    elif f_interest == "Cold": interest_filter = "cold"

    stage_filter = None
    if f_stage != "All":
        stage_filter = next((k for k, v in STAGE_LABELS.items() if v == f_stage), None)

    deals = get_all_deals(buyer_id=buyer_id_filter, interest_level=interest_filter, deal_stage=stage_filter)

    if f_type == "On-Market":
        deals = [d for d in deals if d.get("listing_type") == "on-market"]
    elif f_type == "Off-Market":
        deals = [d for d in deals if d.get("listing_type", "off-market") == "off-market"]

    if f_perfect:
        deals = [d for d in deals if d.get("match_score", 0) >= 80]

    # Apply minimum score filter
    deals = [d for d in deals if d.get("match_score", 0) >= f_min_score]

    # Seller / Match / Reached-Out view mode
    if f_view_mode == "sellers":
        deals = [d for d in deals if not (d.get("buyer_id") or "").strip()]
    elif f_view_mode == "matches":
        deals = [d for d in deals if (d.get("buyer_id") or "").strip()]
    elif f_view_mode == "reached_out":
        deals = [d for d in deals
                 if (d.get("buyer_id") or "").strip()
                 and (d.get("last_contacted") or "").strip()]

    # Free-text search filter
    if f_search and f_search.strip():
        _q = f_search.strip().lower()
        deals = [
            d for d in deals
            if _q in (d.get("company_name") or "").lower()
            or _q in (d.get("owner_name") or "").lower()
            or _q in (d.get("owner_email") or "").lower()
            or _q in (d.get("industry") or "").lower()
            or _q in (d.get("city") or "").lower()
            or _q in (d.get("state") or "").lower()
            or _q in (d.get("buyer_name") or "").lower()
        ]

    # Contact quality filter
    def _has_email(d):  return bool((d.get("owner_email") or "").strip())
    def _has_phone(d):  return bool((d.get("owner_phone") or d.get("phone") or "").strip())
    def _has_web(d):    return bool((d.get("company_domain") or "").strip())

    if f_contact == "Email + Phone + Web":
        deals = [d for d in deals if _has_email(d) and _has_phone(d) and _has_web(d)]
    elif f_contact == "Phone + Web":
        deals = [d for d in deals if _has_phone(d) and _has_web(d)]
    elif f_contact == "Phone or Web":
        deals = [d for d in deals if _has_phone(d) or _has_web(d)]
    # "Any" — no filter

    # Industry cluster filter
    if f_cluster and f_cluster != "All Industries":
        cluster_terms = CLUSTERS.get(f_cluster, [])
        if cluster_terms:
            deals = [d for d in deals if any(t in (d.get("industry","") or "").lower() for t in cluster_terms)]

    # Sort
    if f_sort == "Highest Score":
        deals = sorted(deals, key=lambda x: x.get("match_score", 0), reverse=True)
    elif f_sort == "Newest Listed":
        deals = sorted(deals, key=lambda x: x.get("date_added",""), reverse=True)
    elif f_sort == "Overdue Contact":
        deals = sorted(deals, key=lambda x: days_since(x.get("last_contacted","")), reverse=True)
    # default = interest + newest

    # New seller intakes alert
    new_intakes = get_seller_intakes(reviewed=False)
    if new_intakes:
        st.markdown(
            f'<div class="intake-banner">📬 <strong>{len(new_intakes)} new seller intake submissions</strong> — '
            f'review them in the <strong>Automation</strong> tab.</div>',
            unsafe_allow_html=True
        )

    if not deals:
        _total_in_db = len(get_all_deals())
        if _total_in_db > 0:
            _ne1, _ne2 = st.columns([3, 1])
            with _ne1:
                st.warning(
                    f"**0 of {_total_in_db} leads match current filters.** "
                    f"Your DB has phone on 47% and website on 27% of leads. "
                    f"Email is 0% until you enrich with Hunter.io."
                )
            with _ne2:
                if st.button("👀 Show All Leads", key="ls_show_all", type="primary", use_container_width=True):
                    st.session_state["ls_contact"]  = "Any"
                    st.session_state["ls_min_score_filter"] = 0
                    st.session_state["ls_buyer"]    = "All Buyers"
                    st.session_state["ls_interest"] = "All"
                    st.session_state["ls_type"]     = "All"
                    st.session_state["ls_stage"]    = "All"
                    st.session_state["ls_cluster"]  = "All Industries"
                    st.session_state["ls_perfect"]  = False
                    st.session_state["ls_search"]   = ""
                    st.rerun()
            st.caption(
                "→ Or open **⚙️ Tools → 🧹 Maintenance** and click **✉ Find Emails (Hunter)** to enrich emails. "
                "→ Or change **Contact Quality** above to relax the filter."
            )
        else:
            st.info("No leads yet. Use **⚙️ Tools → 🔄 Scrape** or **➕ Add Lead** to get started.")
    else:
        PAGE_SIZE = 25
        total_pages = max(1, (len(deals) + PAGE_SIZE - 1) // PAGE_SIZE)

        # Page navigation
        pg_col1, pg_col2, pg_col3 = st.columns([2, 4, 2])
        if "ls_page" not in st.session_state:
            st.session_state.ls_page = 0
        # Reset page when filters change
        filter_key = f"{f_buyer}|{f_interest}|{f_type}|{f_stage}"
        if st.session_state.get("_ls_filter_key") != filter_key:
            st.session_state.ls_page = 0
            st.session_state["_ls_filter_key"] = filter_key

        with pg_col1:
            if st.button("◀ Prev", disabled=st.session_state.ls_page == 0, key="ls_prev"):
                st.session_state.ls_page -= 1
                st.rerun()
        with pg_col2:
            start = st.session_state.ls_page * PAGE_SIZE
            end   = min(start + PAGE_SIZE, len(deals))
            st.markdown(
                f"<div style='text-align:center;padding:6px'>"
                f"<strong>{len(deals)} leads</strong> — Page {st.session_state.ls_page+1}/{total_pages} "
                f"(showing {start+1}–{end})</div>",
                unsafe_allow_html=True
            )
        with pg_col3:
            if st.button("Next ▶", disabled=st.session_state.ls_page >= total_pages - 1, key="ls_next"):
                st.session_state.ls_page += 1
                st.rerun()

        page_deals = deals[start:end]

        import html as _html

        def _e(v):
            """HTML-escape a raw value so it never breaks the template."""
            return _html.escape(str(v or ""))

        def _money_cell(v):
            s = str(v or "").strip()
            return f'<strong style="color:#e8eaf0">{_e(s)}</strong>' if s and s not in ("—","-","None","0") else '<span style="color:#3a3a3a">—</span>'

        def _dim(text):
            return f'<span style="color:#555;font-size:0.78rem">{text}</span>'

        for deal in page_deals:
            did     = deal["id"]
            company = (deal.get("company_name") or "").strip()
            if not company or company in ("—", "-", "Unknown", "None"):
                continue

            ilevel     = deal.get("interest_level", "cold")
            stage      = deal.get("deal_stage", "identified")
            score      = deal.get("match_score", 0)
            owner      = (deal.get("owner_name") or "").strip()
            o_email    = (deal.get("owner_email") or "").strip()
            phone      = (deal.get("owner_phone") or deal.get("phone") or "").strip()
            industry   = (deal.get("industry") or "").strip()
            state      = (deal.get("state") or "").strip()
            revenue    = (deal.get("revenue_estimate") or "").strip()
            ebitda     = (deal.get("ebitda_estimate") or "").strip()
            asking     = (deal.get("asking_price") or "").strip()
            ltype      = deal.get("listing_type", "off-market")
            buyer_name = (deal.get("buyer_name") or "").strip()
            last_c     = deal.get("last_contacted", "")
            days_c     = days_since(last_c)
            added      = (deal.get("date_added") or "")[:10]
            city_val   = (deal.get("city") or "").strip()
            domain_val = (deal.get("company_domain") or "").strip()
            notes_val  = (deal.get("response_notes") or "")[:120].strip()

            # Data quality 0-5
            dq = sum([bool(revenue), bool(ebitda or asking), bool(owner), bool(o_email), bool(state)])
            dq_stars = "★" * dq + "☆" * (5 - dq)
            dq_color = "#27AE60" if dq >= 4 else "#F39C12" if dq >= 2 else "#555"

            is_perfect = score >= 80
            if is_perfect:           border_c, bg_c = "#F1C40F", "#18160a"
            elif ltype == "on-market": border_c, bg_c = "#27AE60", "#0a140d"
            elif ilevel == "hot":    border_c, bg_c = "#E74C3C", "#180a0a"
            elif ilevel == "warm":   border_c, bg_c = "#F39C12", "#18130a"
            else:                    border_c, bg_c = "#8E44AD", "#110a18"

            scolor      = STAGE_COLORS.get(stage, "#4A90D9")
            sc_         = score_color(score)
            days_label  = f"{days_c}d ago" if days_c < 9999 else "Never"
            overdue     = days_c > 14
            location_str = ", ".join(filter(None, [city_val, state])) or "—"

            # Safe escaped values
            co_e   = _e(company)
            ind_e  = _e(industry) if industry else '<span style="color:#3a3a3a">—</span>'
            loc_e  = _e(location_str)
            byr_e  = _e(buyer_name)
            add_e  = _e(added)
            day_e  = _e(days_label)
            own_e  = f'<strong style="color:#e8eaf0">{_e(owner)}</strong>' if owner else _dim("Owner unknown")
            eml_e  = (f'<a href="mailto:{_e(o_email)}" style="color:#4A90D9;font-size:0.8rem">{_e(o_email)}</a>'
                      if o_email else _dim("no email"))
            phn_e  = f'<span style="color:#aaa;font-size:0.8rem">{_e(phone)}</span>' if phone else _dim("no phone")
            # Detect local file paths (Windows C:\ or file:/// or POSIX /home/)
            _is_local_path = (
                domain_val.startswith("file:///")
                or (len(domain_val) >= 3 and domain_val[1:3] == ":\\")
                or domain_val.startswith("/") and "://" not in domain_val
            )
            if domain_val and _is_local_path:
                # Show as copyable path text — browsers block file:// from web pages
                _path_only = domain_val.replace("file:///","").replace("/","\\") if domain_val.startswith("file:///") else domain_val
                _short = _path_only.split("\\")[-1] if "\\" in _path_only else _path_only.split("/")[-1]
                dom_e = (f'<span title="{_e(_path_only)}" '
                         f'style="color:#F39C12;font-size:0.75rem">📄 {_e(_short[:50])}</span>'
                         f' <span style="color:#444;font-size:0.7rem">(local file — copy path from tooltip)</span>')
            elif domain_val:
                _dom_href = domain_val if domain_val.startswith("http") else "https://" + domain_val
                _dom_label = domain_val if len(domain_val) <= 40 else domain_val[:37] + "…"
                dom_e = (f'<a href="{_e(_dom_href)}" target="_blank" '
                         f'style="color:#4A90D9;font-size:0.78rem">{_e(_dom_label)}</a>')
            else:
                _gq = _urlquote(f'"{company}" {city_val} {state}')
                _lq = _urlquote(f'owner {company}')
                dom_e = (f'<a href="https://www.google.com/search?q={_gq}" target="_blank" '
                         f'style="color:#666;font-size:0.75rem">🔍 Google</a>'
                         f'&nbsp;&nbsp;<a href="https://www.linkedin.com/search/results/people/?keywords={_lq}" '
                         f'target="_blank" style="color:#0077b5;font-size:0.75rem">💼 LinkedIn</a>')
            nts_e  = _e(notes_val)

            type_pill    = '<span class="type-pill-on">🟢 Listed</span>' if ltype == "on-market" else '<span class="type-pill-off">🟣 Off-Market</span>'
            perfect_html = '<span class="perfect-badge">🎯 Perfect</span>' if is_perfect else ""
            stage_lbl    = _e(STAGE_LABELS.get(stage, stage))
            overdue_col  = "#E74C3C" if overdue else "#555"

            # Build buyer-row block as a separate string
            if buyer_name and buyer_name not in ("", "—", "Unknown"):
                _buyer_block = (
                    f'<div style="background:#1a0d24;border-left:2px solid #9B59B6;padding:4px 10px;'
                    f'margin-bottom:8px;border-radius:0 4px 4px 0">'
                    f'<span style="font-size:0.62rem;color:#9B59B6;letter-spacing:1.2px;font-weight:700">🤝 MATCHED BUYER</span> &nbsp; '
                    f'<strong style="color:#d4b3e8;font-size:0.88rem">{byr_e}</strong>'
                    f'<span style="float:right;color:{overdue_col};font-size:0.72rem">Last contact: {day_e}{"  ⚠" if overdue else ""}</span>'
                    f'</div>'
                )
            else:
                _buyer_block = (
                    f'<div style="background:#241a0d;border-left:2px solid #F39C12;padding:4px 10px;'
                    f'margin-bottom:8px;border-radius:0 4px 4px 0;font-size:0.72rem;color:#F39C12">'
                    f'⚠ No buyer matched yet · go to <strong>📁 Pre-Pipeline</strong> to match this seller</div>'
                )

            card_html = (
                f'<div style="border-left:4px solid {border_c};border-radius:0 8px 8px 0;'
                f'padding:10px 14px;margin-bottom:6px;background:{bg_c};border:1px solid {border_c}33">'

                # title row with explicit SELLER ↔ BUYER labels
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'flex-wrap:wrap;gap:6px;margin-bottom:4px">'
                f'<span style="display:flex;flex-direction:column;gap:1px">'
                f'<span style="font-size:0.62rem;color:#F39C12;letter-spacing:1.2px;font-weight:700">📦 SELLER (target)</span>'
                f'<span style="font-size:1.02rem;font-weight:800;color:#f0f0f0">{INTEREST_EMOJI.get(ilevel,"●")} {co_e}</span>'
                f'</span>'
                f'<span style="display:flex;gap:4px;align-items:center;flex-wrap:wrap">'
                f'{type_pill}{perfect_html}'
                f'<span class="badge" style="background:{sc_}">{score}/100</span>'
                f'<span class="badge" style="background:{scolor}">{stage_lbl}</span>'
                f'<span style="color:{dq_color};font-size:0.75rem" title="Data quality (5=best)">{dq_stars}</span>'
                f'</span></div>'

                # buyer match row (seller-only or matched, set above as _buyer_block)
                f'{_buyer_block}'

                # 3-column grid
                f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;font-size:0.83rem">'

                # Business col
                f'<div style="background:#0a0c10;border:1px solid #1a1d24;border-radius:6px;padding:8px 12px">'
                f'<div style="color:#444;font-size:0.65rem;text-transform:uppercase;font-weight:700;margin-bottom:4px">Business</div>'
                f'<div style="color:#ccc">{ind_e}</div>'
                f'<div style="color:#777;font-size:0.8rem">{loc_e}</div>'
                f'<div style="margin-top:3px">{dom_e}</div>'
                f'</div>'

                # Financials col
                f'<div style="background:#0a0c10;border:1px solid #1a1d24;border-radius:6px;padding:8px 12px">'
                f'<div style="color:#444;font-size:0.65rem;text-transform:uppercase;font-weight:700;margin-bottom:4px">Financials</div>'
                f'<div style="margin-bottom:2px">Rev: {_money_cell(revenue)}</div>'
                f'<div style="margin-bottom:2px">EBITDA: {_money_cell(ebitda)}</div>'
                f'<div style="color:#666;font-size:0.8rem">Ask: {_money_cell(asking)}</div>'
                f'</div>'

                # Owner col
                f'<div style="background:#0a0c10;border:1px solid #1a1d24;border-radius:6px;padding:8px 12px">'
                f'<div style="color:#444;font-size:0.65rem;text-transform:uppercase;font-weight:700;margin-bottom:4px">Owner</div>'
                f'<div style="margin-bottom:2px">{own_e}</div>'
                f'<div style="margin-bottom:2px">{eml_e}</div>'
                f'<div>{phn_e}</div>'
                f'</div>'

                f'</div>'  # end grid

                + (f'<div style="margin-top:8px;font-size:0.76rem;color:#555;font-style:italic;'
                   f'border-top:1px solid #1a1d24;padding-top:6px">{nts_e}</div>' if notes_val else "")
                + '</div>'
            )

            st.markdown(card_html, unsafe_allow_html=True)

            _ba1, _ba2, _ba3, _ba4, _ba5, _ba6 = st.columns([2, 2, 2, 2, 2, 1])

            new_interest = _ba1.selectbox(
                "Interest", ["cold","warm","hot"],
                index=["cold","warm","hot"].index(ilevel),
                key=f"il_{did}", label_visibility="collapsed"
            )
            if new_interest != ilevel:
                update_deal(did, {"interest_level": new_interest,
                                  "seller_interested": 1 if new_interest == "hot" else 0})
                st.rerun()

            if _ba2.button("▶ Pipeline", key=f"pipe_{did}", use_container_width=True):
                _sl = [s for s in DEAL_STAGES if s != "dead"]
                _ci = _sl.index(stage) if stage in _sl else 0
                update_deal(did, {"deal_stage": _sl[min(_ci+1,len(_sl)-1)],
                                  "status": _sl[min(_ci+1,len(_sl)-1)]})
                st.rerun()

            if _ba3.button("📧 Draft", key=f"dft_{did}", use_container_width=True):
                st.session_state[f"expand_email_{did}"] = True

            if _ba4.button("✓ Contacted", key=f"cont_{did}", use_container_width=True):
                update_deal(did, {"last_contacted": datetime.now().isoformat(),
                                  "status": "emailed", "deal_stage": "contacted"})
                st.rerun()

            if _ba5.button("📝 NDA", key=f"nda_{did}", use_container_width=True):
                subj, body = _nda_email(deal)
                if o_email:
                    ok, msg = create_gmail_draft(o_email, subj, body)
                    if ok:
                        update_deal(did, {"nda_sent": 1, "nda_sent_at": datetime.now().isoformat(),
                                          "deal_stage": "nda_sent", "status": "emailed",
                                          "last_contacted": datetime.now().isoformat()})
                        log_outreach(did, deal.get("buyer_id",""), deal.get("buyer_name",""), "nda_email")
                        st.success("NDA draft created!"); st.rerun()
                    else:
                        st.error(f"Gmail: {msg}")
                else:
                    st.warning("No email on file.")

            if _ba6.button("🗑", key=f"deldeal_{did}", use_container_width=True, help="Delete"):
                delete_deal(did); st.rerun()

            # Email expander
            if st.session_state.get(f"expand_email_{did}"):
                with st.expander("Draft Email", expanded=True):
                    buyer_obj_for_email = next((b for b in all_buyers if b["id"] == deal.get("buyer_id")), {})
                    subj_key = f"esubj_{did}"
                    body_key = f"ebody_{did}"
                    to_key   = f"eto_{did}"
                    if subj_key not in st.session_state:
                        st.session_state[subj_key] = deal.get("email_subject") or email_subject(deal, buyer_obj_for_email)
                    if body_key not in st.session_state:
                        st.session_state[body_key] = deal.get("email_body") or _default_outreach_email(deal, buyer_obj_for_email)
                    if to_key not in st.session_state:
                        st.session_state[to_key] = o_email or ""
                    to_v   = st.text_input("TO:", value=st.session_state[to_key],   key=f"eti_{did}")
                    subj_v = st.text_input("SUBJECT:", value=st.session_state[subj_key], key=f"esi_{did}")
                    body_v = st.text_area("BODY:", value=st.session_state[body_key], key=f"ebi_{did}", height=200)
                    eb1, eb2, eb3 = st.columns(3)
                    if eb1.button("Create Gmail Draft", key=f"cgd_{did}", type="primary", use_container_width=True):
                        if to_v:
                            ok, msg = create_gmail_draft(to_v, subj_v, body_v)
                            if ok:
                                mark_draft_created(did)
                                st.success("Draft created!")
                            else:
                                st.error(f"Gmail: {msg}")
                        else:
                            st.error("Enter a TO address.")
                    if eb2.button("Regenerate (AI)", key=f"regen_{did}", use_container_width=True):
                        with st.spinner("Generating..."):
                            new_b = _generate_email(deal, buyer_obj_for_email)
                        st.session_state[body_key] = new_b
                        st.rerun()
                    if eb3.button("Close", key=f"cls_{did}", use_container_width=True):
                        st.session_state[f"expand_email_{did}"] = False
                        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — DEAL PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

with tab2:
    st.subheader("📋 Deal Pipeline")
    st.caption("Deals land here when you click **Pipeline +** in Lead Stream. Drag across columns as the deal progresses.")

    PIPELINE_STAGES = ["contacted", "nda_sent", "nda_signed", "cim_sent", "loi", "due_diligence", "under_contract", "closed"]

    dp_f1, dp_f2 = st.columns([4, 2])
    all_buyers = get_all_buyers()
    dp_buyer  = dp_f1.selectbox("Filter by Buyer", ["All Buyers"] + [b["name"] for b in all_buyers], key="dp_buyer")
    show_dead = dp_f2.checkbox("Show Dead Deals", key="dp_show_dead")

    dp_buyer_id = None
    if dp_buyer != "All Buyers":
        dp_buyer_id = next((b["id"] for b in all_buyers if b["name"] == dp_buyer), None)

    # Only deals that have been promoted from Lead Stream (stage != identified)
    all_pipeline = get_all_deals(buyer_id=dp_buyer_id)
    pipeline_deals = [d for d in all_pipeline if d.get("deal_stage", "identified") != "identified"]
    if not show_dead:
        pipeline_deals = [d for d in pipeline_deals if d.get("deal_stage") != "dead"]

    if not pipeline_deals:
        st.markdown("""
        <div style='text-align:center;padding:60px 20px;color:#888'>
            <div style='font-size:3rem'>📭</div>
            <div style='font-size:1.2rem;font-weight:600;margin-top:12px'>Pipeline is empty</div>
            <div style='margin-top:8px'>Go to <strong>Lead Stream</strong> → click <strong>Pipeline +</strong> on any lead to start tracking a deal here.</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        # Stage count summary bar
        stage_counts = {s: sum(1 for d in pipeline_deals if d.get("deal_stage") == s) for s in PIPELINE_STAGES}
        bar = st.columns(len(PIPELINE_STAGES))
        for stg, col in zip(PIPELINE_STAGES, bar):
            cnt = stage_counts[stg]
            color = STAGE_COLORS[stg] if cnt > 0 else "#333"
            col.markdown(
                f"<div style='background:{color};color:white;border-radius:6px;padding:4px 4px;"
                f"text-align:center;font-size:0.65rem;font-weight:700'>"
                f"{STAGE_LABELS[stg]}<br><span style='font-size:1rem'>{cnt}</span></div>",
                unsafe_allow_html=True
            )

        st.markdown("---")

        # ── Kanban board ── 4 grouped columns ──
        KB_GROUPS = [
            ("📞 Contacted",     ["contacted"],                          "#4A90D9"),
            ("📝 NDA Phase",     ["nda_sent", "nda_signed"],             "#9B59B6"),
            ("📊 CIM / LOI",     ["cim_sent", "loi"],                    "#E67E22"),
            ("🏁 Closing",       ["due_diligence", "under_contract", "closed"], "#16A085"),
        ]

        kb_cols = st.columns(4)
        for (gname, stage_keys, gcolor), col in zip(KB_GROUPS, kb_cols):
            group_deals = [d for d in pipeline_deals if d.get("deal_stage") in stage_keys]
            with col:
                st.markdown(
                    f"<div style='background:{gcolor};color:white;padding:7px 12px;"
                    f"border-radius:6px;font-weight:700;margin-bottom:10px'>"
                    f"{gname} <span style='opacity:0.8'>({len(group_deals)})</span></div>",
                    unsafe_allow_html=True
                )

                if not group_deals:
                    st.markdown(
                        "<div style='border:2px dashed #333;border-radius:6px;"
                        "padding:24px 10px;text-align:center;color:#555;font-size:0.8rem'>"
                        "Empty</div>",
                        unsafe_allow_html=True
                    )

                for d in group_deals:
                    did       = d["id"]
                    company   = d.get("company_name", "Unknown")
                    stage     = d.get("deal_stage", "contacted")
                    score     = d.get("match_score", 0)
                    ilevel    = d.get("interest_level", "cold")
                    o_email   = d.get("owner_email", "")
                    buyer_obj = next((b for b in all_buyers if b["id"] == d.get("buyer_id")), {})
                    buyer_email = buyer_obj.get("email", "")
                    scolor    = STAGE_COLORS.get(stage, "#999")
                    ie        = INTEREST_EMOJI.get(ilevel, "")

                    short_name = company[:30] + ("…" if len(company) > 30 else "")

                    with st.expander(f"{ie} {short_name}", expanded=False):
                        st.markdown(
                            f"<div style='font-size:0.78rem;color:#aaa;margin-bottom:6px'>"
                            f"Buyer: {d.get('buyer_name','—')}<br>"
                            f"Ask: {d.get('asking_price','—')} | Rev: {d.get('revenue_estimate','—')}<br>"
                            f"Owner: {d.get('owner_name','—')} | {o_email or 'no email'}"
                            f"</div>",
                            unsafe_allow_html=True
                        )

                        # Stage stepper (visual only)
                        step_html = ""
                        for s in PIPELINE_STAGES:
                            if s == stage:
                                step_html += f"<span style='background:{STAGE_COLORS[s]};color:white;padding:2px 6px;border-radius:4px;font-size:0.65rem;font-weight:700;margin:1px'>{STAGE_LABELS[s]}</span>"
                            elif PIPELINE_STAGES.index(s) < PIPELINE_STAGES.index(stage):
                                step_html += f"<span style='background:#555;color:#bbb;padding:2px 6px;border-radius:4px;font-size:0.65rem;margin:1px'>{STAGE_LABELS[s]}</span>"
                            else:
                                step_html += f"<span style='background:#2a2a2a;color:#555;padding:2px 6px;border-radius:4px;font-size:0.65rem;margin:1px'>{STAGE_LABELS[s]}</span>"
                        st.markdown(step_html, unsafe_allow_html=True)

                        # Move stage
                        new_stage = st.selectbox(
                            "Move to Stage",
                            PIPELINE_STAGES,
                            index=PIPELINE_STAGES.index(stage) if stage in PIPELINE_STAGES else 0,
                            format_func=lambda x: STAGE_LABELS[x],
                            key=f"kb_stage_{did}"
                        )
                        if new_stage != stage:
                            update_deal(did, {"deal_stage": new_stage, "status": new_stage})
                            st.rerun()

                        # Context-aware action buttons
                        if stage in ("contacted", "nda_sent"):
                            if st.button("1-Click NDA →Seller", key=f"kb_nda_{did}", type="primary", use_container_width=True):
                                subj, body = _nda_email(d)
                                if o_email:
                                    ok, msg = create_gmail_draft(o_email, subj, body)
                                    if ok:
                                        update_deal(did, {
                                            "nda_sent": 1, "nda_sent_at": datetime.now().isoformat(),
                                            "deal_stage": "nda_sent", "last_contacted": datetime.now().isoformat()
                                        })
                                        log_outreach(did, d.get("buyer_id",""), buyer_obj.get("name",""), "nda")
                                        st.success("NDA drafted in Gmail!")
                                        st.rerun()
                                    else:
                                        st.error(msg)
                                else:
                                    st.warning("No seller email.")
                                    st.text_area("NDA Body (copy manually)", body, height=100, key=f"kb_nda_txt_{did}")

                        if stage == "nda_sent":
                            if st.button("✅ NDA Signed", key=f"kb_ndasign_{did}", use_container_width=True):
                                update_deal(did, {
                                    "nda_signed": 1, "nda_signed_at": datetime.now().isoformat(),
                                    "deal_stage": "nda_signed"
                                })
                                st.success("NDA signed!")
                                st.rerun()

                        if stage in ("nda_signed", "cim_sent"):
                            if st.button("1-Click CIM →Buyer", key=f"kb_cim_{did}", type="primary", use_container_width=True):
                                subj, body = _cim_email(d, buyer_obj)
                                if buyer_email:
                                    ok, msg = create_gmail_draft(buyer_email, subj, body)
                                    if ok:
                                        update_deal(did, {
                                            "cim_sent": 1, "cim_sent_at": datetime.now().isoformat(),
                                            "deal_stage": "cim_sent"
                                        })
                                        log_outreach(did, buyer_obj.get("id",""), buyer_obj.get("name",""), "cim")
                                        st.success("CIM drafted to buyer!")
                                        st.rerun()
                                    else:
                                        st.error(msg)
                                else:
                                    st.warning(f"No email for buyer — add in Buyers & Settings.")

                        if stage in ("cim_sent", "loi"):
                            if st.button("📄 LOI Received", key=f"kb_loi_{did}", use_container_width=True):
                                update_deal(did, {
                                    "loi_received": 1, "loi_received_at": datetime.now().isoformat(),
                                    "deal_stage": "loi"
                                })
                                st.rerun()

                        if stage in ("loi", "due_diligence", "under_contract"):
                            if st.button("🎉 Closed!", key=f"kb_close_{did}", type="primary", use_container_width=True):
                                update_deal(did, {"deal_stage": "closed", "status": "qualified"})
                                st.balloons()
                                st.rerun()

                        notes_v = st.text_area("Notes", value=d.get("response_notes",""), key=f"kb_notes_{did}", height=50)
                        nc1, nc2 = st.columns(2)
                        if nc1.button("Save", key=f"kb_save_{did}", use_container_width=True):
                            update_deal(did, {"response_notes": notes_v})
                            st.success("Saved.")
                        if nc2.button("Dead ❌", key=f"kb_dead_{did}", use_container_width=True):
                            update_deal(did, {"deal_stage": "dead"})
                            st.rerun()

                        # Outreach log
                        log = get_outreach_log(deal_id=did)
                        if log:
                            with st.expander(f"History ({len(log)})", expanded=False):
                                for entry in log:
                                    st.caption(f"{entry.get('sent_at','')[:16]} | {entry.get('outreach_type','')} | {entry.get('notes','')}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — OUTREACH
# ══════════════════════════════════════════════════════════════════════════════

with tab3:
    st.subheader("Outreach")
    st.caption("Cross-post deals to buyers, generate LinkedIn messages, and track all outreach activity.")

    ot1, ot2 = st.tabs(["Cross-Post Deals", "Outreach History"])

    with ot1:
        st.markdown("#### Select a Deal to Cross-Post to Buyers")
        all_deals_for_outreach = get_all_deals()
        if not all_deals_for_outreach:
            st.info("No deals yet. Add some leads in the Lead Stream tab first.")
        else:
            deal_opts = {f"{d['company_name']} ({d.get('state','')}) — Score: {d.get('match_score',0)}": d["id"]
                         for d in all_deals_for_outreach}
            selected_deal_label = st.selectbox("Deal to cross-post:", list(deal_opts.keys()), key="ot_deal_sel")
            selected_deal_id    = deal_opts[selected_deal_label]
            selected_deal       = next((d for d in all_deals_for_outreach if d["id"] == selected_deal_id), {})

            # Show deal summary
            oc1, oc2 = st.columns(2)
            oc1.markdown(f"**Company:** {selected_deal.get('company_name','')}")
            oc1.markdown(f"**Industry:** {selected_deal.get('industry','—')}")
            oc1.markdown(f"**State:** {selected_deal.get('state','—')}")
            oc2.markdown(f"**Revenue:** {selected_deal.get('revenue_estimate','—')}")
            oc2.markdown(f"**EBITDA:** {selected_deal.get('ebitda_estimate','—')}")
            oc2.markdown(f"**Asking:** {selected_deal.get('asking_price','—')}")

            already_posted = json.loads(selected_deal.get("cross_posted_buyers") or "[]")

            st.markdown("#### Matching Buyers")
            all_buyers = get_all_buyers()
            buyer_scores = []
            for buyer in all_buyers:
                sc = score_lead(selected_deal, buyer)
                buyer_scores.append((buyer, sc))
            buyer_scores.sort(key=lambda x: x[1], reverse=True)

            for buyer, sc in buyer_scores[:15]:  # top 15 matches
                posted_badge = " ✅ Sent" if buyer["id"] in already_posted else ""
                bc1, bc2, bc3, bc4, bc5 = st.columns([3, 2, 2, 2, 2])
                bc1.markdown(f"**{buyer['name']}**{posted_badge}")
                bc2.markdown(f"Match: {sc}/100")
                bc3.markdown(f"{buyer.get('email','—')}")

                with bc4:
                    if st.button("Draft CIM Email", key=f"ot_cim_{selected_deal_id}_{buyer['id']}", use_container_width=True):
                        subj, body = _cim_email(selected_deal, buyer)
                        to_addr = buyer.get("email", "")
                        if to_addr:
                            ok, msg = create_gmail_draft(to_addr, subj, body)
                            if ok:
                                mark_cross_posted(selected_deal_id, buyer["id"])
                                log_outreach(selected_deal_id, buyer["id"], buyer["name"], "cim_cross_post")
                                st.success(f"Drafted to {buyer['name']}!")
                                st.rerun()
                            else:
                                st.error(f"Gmail: {msg}")
                        else:
                            st.warning(f"No email for {buyer['name']}")

                with bc5:
                    if st.button("LinkedIn Msg", key=f"ot_li_{selected_deal_id}_{buyer['id']}", use_container_width=True):
                        msg = _linkedin_message(selected_deal, buyer)
                        st.session_state[f"li_msg_{selected_deal_id}_{buyer['id']}"] = msg

                li_key = f"li_msg_{selected_deal_id}_{buyer['id']}"
                if li_key in st.session_state:
                    st.text_area(f"LinkedIn — {buyer['name']}", st.session_state[li_key],
                                 height=120, key=f"li_ta_{selected_deal_id}_{buyer['id']}")

            # Bulk send
            st.markdown("---")
            high_match_buyers = [(b, sc) for b, sc in buyer_scores if sc >= 60 and b["id"] not in already_posted and b.get("email")]
            if high_match_buyers:
                if st.button(f"Bulk Draft CIM to {len(high_match_buyers)} Buyers (Score >= 60)", type="primary", key="ot_bulk"):
                    count = 0
                    for buyer, sc in high_match_buyers:
                        subj, body = _cim_email(selected_deal, buyer)
                        ok, _ = create_gmail_draft(buyer["email"], subj, body)
                        if ok:
                            mark_cross_posted(selected_deal_id, buyer["id"])
                            log_outreach(selected_deal_id, buyer["id"], buyer["name"], "cim_bulk")
                            count += 1
                    st.success(f"Drafted CIM to {count} buyers!")
                    st.rerun()

    with ot2:
        st.markdown("#### Full Outreach History")
        all_log = get_outreach_log()
        if not all_log:
            st.info("No outreach logged yet.")
        else:
            df_log = pd.DataFrame(all_log)[["sent_at", "outreach_type", "buyer_name", "deal_id", "notes"]]
            df_log["sent_at"] = df_log["sent_at"].str[:16]
            st.dataframe(df_log, width="stretch", hide_index=True)
            st.download_button(
                "Export Outreach Log",
                data=df_log.to_csv(index=False),
                file_name=f"outreach_log_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — AUTOMATION
# ══════════════════════════════════════════════════════════════════════════════

with tab4:
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">'
        '<span style="font-size:1.4rem">🔮</span>'
        '<span style="font-size:1.3rem;font-weight:800;color:#e8eaf0">The Palantír — Automation</span>'
        '</div>'
        '<div style="color:#666;font-size:0.82rem;margin-bottom:12px">'
        'Auto-draft deal outreach. Recruit new buyers into the Fellowship. Review seller intakes.'
        '</div>',
        unsafe_allow_html=True
    )

    auto1, auto_rec, auto2, auto3 = st.tabs([
        "📧 Auto-Draft Deal Emails",
        "👑 Recruit New Buyers",
        "📥 Seller Intakes",
        "📋 Task Queue",
    ])

    with auto_rec:
        st.markdown("#### 👑 Recruit New Buyers into the Fellowship")
        st.caption(
            "Generate personalized teaser emails to PE firms, family offices, and individual buyers. "
            "Emails reference your active deals (anonymized) and drop the buyer agreement link. "
            "Replies → sign Buyer Agreement → portal access."
        )

        # Quick stats
        _rec_all_buyers = get_all_buyers()
        _rec_no_agreement = [b for b in _rec_all_buyers if not b.get("agreement_signed")]
        _ms1, _ms2 = st.columns(2)
        _ms1.metric("👑 Buyers in Fellowship", len(_rec_all_buyers))
        _ms2.metric("⏳ Without signed agreement", len(_rec_no_agreement))

        st.markdown("##### ✍ Compose Recruitment Email")
        _rc1, _rc2 = st.columns(2)
        with _rc1:
            _prospect_name = st.text_input("Prospect Name *", key="rec_name",
                                            placeholder="e.g. Sarah Chen")
            _prospect_firm = st.text_input("Firm / Company", key="rec_firm",
                                            placeholder="e.g. Audax Private Equity")
            _prospect_email = st.text_input("Email *", key="rec_email",
                                             placeholder="sarah@firm.com")
        with _rc2:
            _prospect_focus = st.text_input("Acquisition Focus", key="rec_focus",
                                             placeholder="e.g. industrial services, healthcare, $5M-$20M EBITDA")
            _prospect_geo = st.text_input("Geographic Focus", key="rec_geo",
                                           placeholder="e.g. Southeast, National")
            _custom_hook = st.text_area("Optional: personalization hook", key="rec_hook",
                                         placeholder="e.g. 'Just saw your Pipe View America exit announcement'",
                                         height=70)

        _gen_col1, _gen_col2 = st.columns([1, 3])
        if _gen_col1.button("✨ Generate Email", type="primary", key="rec_gen",
                             use_container_width=True, disabled=not _prospect_name):
            try:
                import buyer_outreach as _bo
                _prospect = {
                    "name":       _prospect_name,
                    "firm":       _prospect_firm,
                    "email":      _prospect_email,
                    "industries": [i.strip() for i in (_prospect_focus or "").split(",") if i.strip()],
                    "geography":  _prospect_geo,
                    "hook":       _custom_hook,
                }
                with st.spinner("Forging email…"):
                    _drafted = _bo.generate_buyer_recruitment_email(_prospect, get_all_deals())
                st.session_state["rec_drafted_subject"] = _drafted["subject"]
                st.session_state["rec_drafted_body"]    = _drafted["body"]
                st.success("✅ Email drafted below.")
            except Exception as _ex:
                st.error(f"Generation failed: {_ex}")

        if "rec_drafted_subject" in st.session_state:
            st.markdown("##### 📧 Drafted Email")
            _sub_v = st.text_input("Subject", value=st.session_state["rec_drafted_subject"], key="rec_sub_edit")
            _body_v = st.text_area("Body", value=st.session_state["rec_drafted_body"], height=300, key="rec_body_edit")

            _sc1, _sc2, _sc3 = st.columns(3)
            if _sc1.button("📨 Create Gmail Draft", type="primary", key="rec_draft",
                            use_container_width=True, disabled=not _prospect_email):
                ok, msg = create_gmail_draft(_prospect_email, _sub_v, _body_v)
                if ok:
                    st.success(f"Gmail draft created for {_prospect_email} — open Gmail to send.")
                else:
                    st.error(f"Gmail: {msg}")
            if _sc2.button("📋 Copy to Clipboard", key="rec_copy", use_container_width=True):
                _full = f"To: {_prospect_email}\nSubject: {_sub_v}\n\n{_body_v}"
                st.code(_full, language=None)
                st.caption("Select the text above and copy.")
            if _sc3.button("🔄 Regenerate", key="rec_regen", use_container_width=True):
                st.session_state.pop("rec_drafted_subject", None)
                st.session_state.pop("rec_drafted_body", None)
                st.rerun()

        st.markdown("---")
        st.markdown("##### 🔗 Direct Links")
        _l1, _l2 = st.columns(2)
        _l1.markdown(
            "**Buyer Form (Google):**  \n"
            "[forms.gle/Ka7VuFtryfkmvpp48](https://forms.gle/Ka7VuFtryfkmvpp48)"
        )
        _l2.markdown(
            "**Buyer Portal (live deals):**  \n"
            "`http://localhost:8601/Buyer_Portal?token=<buyer_id>`"
        )

    auto1, auto2, auto3 = auto1, auto2, auto3  # keep names

    with auto1:
        st.markdown("#### Auto-Draft Emails for Uncontacted Leads")
        st.info(
            "This scans all leads not contacted recently and creates Gmail draft emails. "
            "Only runs for leads with a valid seller email address."
        )
        all_buyers = get_all_buyers()

        ac1, ac2 = st.columns(2)
        draft_days = ac1.number_input("Days since last contact (threshold)", 1, 365, 30, key="ad_days")
        draft_min_score = ac2.slider("Min match score", 0, 100, 40, 5, key="ad_score")

        candidates = get_auto_draft_candidates(days_since_contact=int(draft_days))
        candidates = [d for d in candidates if d.get("match_score", 0) >= draft_min_score]
        candidates_with_email = [d for d in candidates if d.get("owner_email")]

        st.markdown(f"**{len(candidates)} leads** qualify for auto-draft ({len(candidates_with_email)} have email addresses)")

        if candidates_with_email:
            if st.button(f"Create {len(candidates_with_email)} Gmail Drafts", type="primary", key="ad_run"):
                prog = st.progress(0)
                success_count = 0
                for i, deal in enumerate(candidates_with_email):
                    buyer_obj = next((b for b in all_buyers if b["id"] == deal.get("buyer_id")), {})
                    subj = deal.get("email_subject") or email_subject(deal, buyer_obj)
                    body = deal.get("email_body") or _default_outreach_email(deal, buyer_obj)
                    ok, _ = create_gmail_draft(deal["owner_email"], subj, body)
                    if ok:
                        update_deal(deal["id"], {
                            "auto_drafted": 1,
                            "gmail_draft_created": 1,
                            "last_contacted": datetime.now().isoformat(),
                            "deal_stage": "contacted",
                        })
                        success_count += 1
                    prog.progress((i + 1) / len(candidates_with_email))
                st.success(f"Created {success_count} Gmail drafts. Open Gmail to review and send.")

            # Preview table
            df_cand = pd.DataFrame([{
                "Company":    d.get("company_name",""),
                "Owner":      d.get("owner_name",""),
                "Email":      d.get("owner_email",""),
                "Score":      d.get("match_score",0),
                "Last Contact": (d.get("last_contacted","Never") or "Never")[:10],
                "Buyer":      d.get("buyer_name",""),
            } for d in candidates_with_email])
            st.dataframe(df_cand, width="stretch", hide_index=True,
                         column_config={"Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d")})
        else:
            st.success("No uncontacted leads with email addresses at this threshold.")

    with auto2:
        st.markdown("#### Seller Intake Submissions")
        st.caption("Leads submitted via the Seller Intake Form (public URL).")

        show_reviewed = st.checkbox("Show reviewed submissions", value=False, key="si_show_reviewed")
        intakes = get_seller_intakes(reviewed=False if not show_reviewed else None)

        if not intakes:
            st.info(
                "No seller intake submissions yet.\n\n"
                "Share the intake form URL with potential sellers:\n"
                "- Run: `streamlit run tools/dealflow_app.py`\n"
                "- Navigate to: **Seller Intake** page in the sidebar\n"
                "- For a public URL: deploy to Streamlit Cloud (free)"
            )
        else:
            for intake in intakes:
                iid = intake["id"]
                is_reviewed = bool(intake.get("reviewed"))
                badge = "✅ Reviewed" if is_reviewed else "🆕 New"

                with st.expander(
                    f"{badge} — {intake.get('business_name','Unknown')} | {intake.get('industry','')} | {intake.get('state','')} | {(intake.get('submitted_at',''))[:10]}",
                    expanded=not is_reviewed
                ):
                    ic1, ic2, ic3 = st.columns(3)
                    ic1.markdown(f"**Business:** {intake.get('business_name','')}")
                    ic1.markdown(f"**Owner:** {intake.get('owner_name','')}")
                    ic1.markdown(f"**Email:** {intake.get('owner_email','')}")
                    ic1.markdown(f"**Phone:** {intake.get('owner_phone','')}")
                    ic2.markdown(f"**Industry:** {intake.get('industry','')}")
                    ic2.markdown(f"**State:** {intake.get('state','')}")
                    ic2.markdown(f"**Revenue:** {intake.get('annual_revenue','')}")
                    ic2.markdown(f"**EBITDA:** {intake.get('ebitda','')}")
                    ic3.markdown(f"**Asking Price:** {intake.get('asking_price','')}")
                    ic3.markdown(f"**Years Est.:** {intake.get('years_established','')}")
                    ic3.markdown(f"**Employees:** {intake.get('num_employees','')}")
                    ic3.markdown(f"**Reason Selling:** {intake.get('reason_selling','')}")
                    if intake.get("notes"):
                        st.markdown(f"**Notes:** {intake['notes']}")

                    ia1, ia2, ia3 = st.columns(3)
                    with ia1:
                        if st.button("Add to Lead Stream", key=f"si_add_{iid}", type="primary", use_container_width=True):
                            # Find best matching buyer
                            best_buyer = all_buyers[0] if all_buyers else {}
                            best_score = 0
                            for b in all_buyers:
                                sc = score_lead({"industry": intake.get("industry",""), "state": intake.get("state",""),
                                                 "owner_name": intake.get("owner_name",""), "owner_email": intake.get("owner_email",""),
                                                 "revenue_estimate": intake.get("annual_revenue","")}, b)
                                if sc > best_score:
                                    best_score, best_buyer = sc, b
                            lead_data = {
                                "company_name":   intake.get("business_name",""),
                                "owner_name":     intake.get("owner_name",""),
                                "owner_email":    intake.get("owner_email",""),
                                "owner_phone":    intake.get("owner_phone",""),
                                "industry":       intake.get("industry",""),
                                "state":          intake.get("state",""),
                                "revenue_estimate": intake.get("annual_revenue",""),
                                "ebitda_estimate":  intake.get("ebitda",""),
                                "asking_price":   intake.get("asking_price",""),
                                "source":         "seller_intake",
                                "match_score":    best_score,
                                "listing_type":   "off-market",
                                "buyer_id":       best_buyer.get("id",""),
                                "buyer_name":     best_buyer.get("name",""),
                                "deal_stage":     "identified",
                                "status":         "identified",
                            }
                            lead_data["email_subject"] = email_subject(lead_data, best_buyer)
                            lead_data["email_body"]    = _default_outreach_email(lead_data, best_buyer)
                            add_deal(lead_data)
                            mark_intake_reviewed(iid)
                            st.success(f"Added to Lead Stream! Best buyer match: {best_buyer.get('name','')}")
                            st.rerun()
                    with ia2:
                        if st.button("Mark Reviewed", key=f"si_rev_{iid}", use_container_width=True):
                            mark_intake_reviewed(iid)
                            st.rerun()

    with auto3:
        st.markdown("#### Broker Task Queue")
        st.caption("Manual task management for broker workflow.")

        if "tasks" not in st.session_state:
            st.session_state.tasks = []

        with st.form("add_task_form"):
            tc1, tc2, tc3 = st.columns([4, 2, 2])
            task_text     = tc1.text_input("Task description")
            task_priority = tc2.selectbox("Priority", ["High", "Medium", "Low"])
            task_due      = tc3.date_input("Due Date", value=datetime.now().date())
            if st.form_submit_button("Add Task", type="primary"):
                if task_text.strip():
                    st.session_state.tasks.append({
                        "text": task_text, "priority": task_priority,
                        "due": str(task_due), "done": False,
                        "added": datetime.now().strftime("%Y-%m-%d")
                    })

        # Task list
        priority_order = {"High": 0, "Medium": 1, "Low": 2}
        sorted_tasks = sorted(
            enumerate(st.session_state.tasks),
            key=lambda x: (x[1]["done"], priority_order.get(x[1]["priority"], 99))
        )
        priority_colors = {"High": "#E74C3C", "Medium": "#F39C12", "Low": "#3498DB"}
        for orig_idx, task in sorted_tasks:
            tc1, tc2, tc3, tc4 = st.columns([5, 2, 2, 1])
            style = "text-decoration:line-through;color:#999" if task["done"] else ""
            pcolor = priority_colors.get(task["priority"], "#999")
            pname  = task["priority"]
            ttext  = task["text"]
            tc1.markdown(
                f"<span style='{style}'>{ttext}</span> "
                f"<span class='badge' style='background:{pcolor}'>{pname}</span>",
                unsafe_allow_html=True
            )
            tc2.caption(f"Due: {task['due']}")
            done_val = tc3.checkbox("Done", value=task["done"], key=f"task_done_{orig_idx}")
            if done_val != task["done"]:
                st.session_state.tasks[orig_idx]["done"] = done_val
                st.rerun()
            if tc4.button("X", key=f"task_del_{orig_idx}"):
                st.session_state.tasks.pop(orig_idx)
                st.rerun()

        # Pipeline funnel chart
        st.markdown("---")
        st.markdown("#### Pipeline Funnel")
        funnel_stages = ["identified", "contacted", "nda_sent", "nda_signed", "cim_sent", "loi", "due_diligence", "under_contract", "closed"]
        funnel_vals   = [stats.get(f"stage_{s}", 0) for s in funnel_stages]
        funnel_labels = [STAGE_LABELS[s] for s in funnel_stages]
        funnel_colors = [STAGE_COLORS[s] for s in funnel_stages]

        fig = go.Figure(go.Funnel(
            y=funnel_labels, x=funnel_vals,
            textinfo="value+percent initial",
            marker=dict(color=funnel_colors),
        ))
        fig.update_layout(height=300, margin=dict(t=10, b=10, l=0, r=0),
                          paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, width="stretch")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — BUYERS & SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

with tab5:
    st.subheader("Buyer Mandates & Settings")

    bs1, bs2 = st.tabs(["Buyer Mandates", "App Settings"])

    with bs1:
        st.caption("Import, edit, and manage all buyer mandates. Changes sync to the pipeline database.")

        ib1, ib2, ib3 = st.columns([3, 3, 2])
        with ib1:
            uploaded = st.file_uploader(
                "Upload Buyer Excel / CSV", type=["xlsx", "xls", "csv"],
                help="Columns: name, industries, states, ebitda_min, ebitda_max, deal_size_min, deal_size_max, email, phone"
            )
            if uploaded:
                with st.spinner("Importing..."):
                    n = import_buyers_from_excel(uploaded)
                st.success(f"Imported {n} buyers.")

        with ib2:
            if st.button("Load Built-in 39 Buyers", use_container_width=True):
                n = sync_hardcoded_buyers()
                st.success(f"Loaded {n} buyers.")

        with ib3:
            all_buyers = get_all_buyers()
            st.download_button(
                "Export Buyers CSV",
                data=pd.DataFrame(all_buyers).to_csv(index=False),
                file_name="valar_buyers.csv", mime="text/csv",
                use_container_width=True
            )

        all_buyers = get_all_buyers()
        if not all_buyers:
            st.info("No buyers loaded yet.")
        else:
            st.markdown(f"**{len(all_buyers)} buyers**")
            df_buyers = pd.DataFrame([{
                "ID":         b["id"],
                "Name":       b["name"],
                "Contact":    b.get("contact_name") or "",
                "Email":      b.get("email") or "",
                "Phone":      b.get("phone") or "",
                "Industries": ", ".join(b.get("industries") or []),
                "States":     ", ".join(b.get("states") or []),
                "EBITDA Min": b.get("ebitda_min") or 0,
                "EBITDA Max": b.get("ebitda_max") or 0,
                "Deal Min":   b.get("deal_size_min") or 0,
                "Deal Max":   b.get("deal_size_max") or 0,
                "Broker":     b.get("broker") or "",
            } for b in all_buyers])

            edited = st.data_editor(
                df_buyers, width="stretch", num_rows="dynamic", hide_index=True,
                column_config={
                    "EBITDA Min": st.column_config.NumberColumn(format="$%.1fM"),
                    "EBITDA Max": st.column_config.NumberColumn(format="$%.1fM"),
                    "Deal Min":   st.column_config.NumberColumn(format="$%.1fM"),
                    "Deal Max":   st.column_config.NumberColumn(format="$%.1fM"),
                },
                key="buyer_editor"
            )

            if st.button("Save Changes", type="primary", key="save_buyers_btn"):
                for _, row in edited.iterrows():
                    if not str(row.get("Name","")).strip():
                        continue
                    upsert_buyer({
                        "id":            str(row["ID"]),
                        "name":          str(row["Name"]),
                        "contact_name":  str(row.get("Contact") or ""),
                        "email":         str(row.get("Email") or ""),
                        "phone":         str(row.get("Phone") or ""),
                        "industries":    [x.strip() for x in str(row.get("Industries","")).split(",") if x.strip()],
                        "states":        [x.strip() for x in str(row.get("States","")).split(",") if x.strip()],
                        "ebitda_min":    float(row.get("EBITDA Min") or 0),
                        "ebitda_max":    float(row.get("EBITDA Max") or 0),
                        "deal_size_min": float(row.get("Deal Min") or 0),
                        "deal_size_max": float(row.get("Deal Max") or 0),
                        "broker":        str(row.get("Broker") or ""),
                    })
                st.success("Buyers saved.")
                st.rerun()

            with st.expander("Add New Buyer"):
                with st.form("new_buyer_form"):
                    nb1, nb2, nb3 = st.columns(3)
                    nb_name  = nb1.text_input("Buyer Name *")
                    nb_cont  = nb1.text_input("Contact Name")
                    nb_email = nb2.text_input("Email")
                    nb_phone = nb2.text_input("Phone")
                    nb_ind   = nb1.text_input("Industries (comma-sep)", placeholder="manufacturing, aerospace")
                    nb_st    = nb2.text_input("States (comma-sep)", placeholder="TX, OH, IL")
                    nb_geo   = nb3.text_input("Geography notes")
                    nb_emin  = nb3.number_input("EBITDA Min ($M)", 0.0, value=0.0, step=0.5)
                    nb_emax  = nb3.number_input("EBITDA Max ($M)", 0.0, value=5.0, step=0.5)
                    nb_dmin  = nb3.number_input("Deal Min ($M)",   0.0, value=0.0, step=1.0)
                    nb_dmax  = nb3.number_input("Deal Max ($M)",   0.0, value=20.0, step=1.0)
                    nb_notes = st.text_area("Notes", height=60)
                    nb_broker= nb2.text_input("Broker")
                    if st.form_submit_button("Add Buyer", type="primary"):
                        if nb_name.strip():
                            upsert_buyer({
                                "id":            nb_name.lower().replace(" ","_")[:40],
                                "name":          nb_name, "contact_name": nb_cont,
                                "email":         nb_email, "phone": nb_phone,
                                "industries":    [x.strip() for x in nb_ind.split(",") if x.strip()],
                                "states":        [x.strip() for x in nb_st.split(",") if x.strip()],
                                "geographies":   [x.strip() for x in nb_geo.split(",") if x.strip()],
                                "ebitda_min":    nb_emin, "ebitda_max": nb_emax,
                                "deal_size_min": nb_dmin, "deal_size_max": nb_dmax,
                                "notes":         nb_notes, "broker": nb_broker,
                            })
                            st.success(f"Added: {nb_name}")
                            st.rerun()
                        else:
                            st.error("Buyer name is required.")

    with bs2:
        st.markdown("#### Broker Configuration")
        st.caption("Changes save to `.env` in your project folder and reload on next app restart.")

        _cfg_path = ROOT / ".env"

        def _read_env() -> dict:
            vals = {}
            if _cfg_path.exists():
                for line in _cfg_path.read_text(encoding="utf-8").splitlines():
                    if "=" in line and not line.strip().startswith("#"):
                        k, _, v = line.partition("=")
                        vals[k.strip()] = v.strip()
            return vals

        def _save_env_keys(updates: dict):
            lines = []
            if _cfg_path.exists():
                lines = _cfg_path.read_text(encoding="utf-8").splitlines()
            updated_keys = set()
            new_lines = []
            for line in lines:
                if "=" in line and not line.strip().startswith("#"):
                    k = line.partition("=")[0].strip()
                    if k in updates:
                        new_lines.append(f"{k}={updates[k]}")
                        updated_keys.add(k)
                        continue
                new_lines.append(line)
            for k, v in updates.items():
                if k not in updated_keys:
                    new_lines.append(f"{k}={v}")
            _cfg_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")

        cur = _read_env()

        with st.form("broker_settings_form"):
            st.markdown("**Your Info**")
            bf1, bf2 = st.columns(2)
            s_name    = bf1.text_input("Broker Name",    value=cur.get("BROKER_NAME",    BROKER_NAME))
            s_title   = bf1.text_input("Title",          value=cur.get("BROKER_TITLE",   BROKER_TITLE))
            s_company = bf1.text_input("Company",        value=cur.get("BROKER_COMPANY", BROKER_COMPANY))
            s_phone   = bf2.text_input("Phone",          value=cur.get("BROKER_PHONE",   BROKER_PHONE))
            s_email   = bf2.text_input("Email",          value=cur.get("BROKER_EMAIL",   BROKER_EMAIL))
            s_li      = bf2.text_input("LinkedIn URL",   value=cur.get("BROKER_LINKEDIN",""))
            st.markdown("**Gmail SMTP (for auto-drafts)**")
            gf1, gf2 = st.columns(2)
            s_smtp_user = gf1.text_input("Gmail Address",    value=cur.get("SMTP_USER",""))
            s_smtp_pass = gf2.text_input("App Password",     value=cur.get("SMTP_PASSWORD",""), type="password",
                                         help="Gmail App Password (16 chars) — not your real password")
            st.markdown("**API Keys** — the rings of power that bind the forge together")
            ak1, ak2 = st.columns(2)
            s_anthropic = ak1.text_input("🔮 Anthropic API Key", value=cur.get("ANTHROPIC_API_KEY",""), type="password",
                                          help="Claude API for AI email generation")
            s_apollo    = ak2.text_input("⚔ Apollo.io API Key",    value=cur.get("APOLLO_API_KEY",""), type="password",
                                          help="Apollo.io for contact + email enrichment (replaces Hunter)")
            s_serpapi   = ak1.text_input("🔍 SerpAPI Key",       value=cur.get("SERPAPI_KEY",""), type="password",
                                          help="Google Search enrichment for owner names + revenue")
            s_yelp      = ak2.text_input("📍 Yelp Fusion API",    value=cur.get("YELP_API_KEY",""), type="password",
                                          help="Yelp Fusion API for local business scraping")
            s_clay_key  = ak1.text_input("⚒ Clay API Key",        value=cur.get("CLAY_API_KEY",""), type="password",
                                          help="Clay.com API key — paste here after signup at clay.com")
            s_clay_url  = ak2.text_input("⚒ Clay Webhook URL",    value=cur.get("CLAY_WEBHOOK_URL",""),
                                          placeholder="https://api.clay.com/v1/connect/webhook/...",
                                          help="Clay table webhook URL — push scraped leads to Clay for enrichment")

            if st.form_submit_button("💾 Save Settings", type="primary"):
                _save_env_keys({
                    "BROKER_NAME":       s_name,
                    "BROKER_TITLE":      s_title,
                    "BROKER_COMPANY":    s_company,
                    "BROKER_PHONE":      s_phone,
                    "BROKER_EMAIL":      s_email,
                    "BROKER_LINKEDIN":   s_li,
                    "SMTP_USER":         s_smtp_user,
                    "SMTP_PASSWORD":     s_smtp_pass,
                    "ANTHROPIC_API_KEY": s_anthropic,
                    "APOLLO_API_KEY":    s_apollo,
                    "SERPAPI_KEY":       s_serpapi,
                    "YELP_API_KEY":      s_yelp,
                    "CLAY_API_KEY":      s_clay_key,
                    "CLAY_WEBHOOK_URL":  s_clay_url,
                })
                st.success("⚒ Saved to .env — restart the forge to apply API key changes.")

        bc2_status = st.columns(5)
        bc2_status[0].metric("📧 SMTP",      "✅" if SMTP_USER else "❌")
        bc2_status[1].metric("🔮 Anthropic", "✅" if os.getenv("ANTHROPIC_API_KEY","").startswith("sk-") else "❌")
        bc2_status[2].metric("⚔ Apollo",    "✅" if os.getenv("APOLLO_API_KEY","") not in ("","your-apollo-api-key") else "❌")
        bc2_status[3].metric("🔍 SerpAPI",   "✅" if os.getenv("SERPAPI_KEY","") not in ("","your-serpapi-key") else "❌")
        bc2_status[4].metric("⚒ Clay",      "✅" if os.getenv("CLAY_API_KEY","") and os.getenv("CLAY_WEBHOOK_URL","") else "❌")

        st.markdown("---")
        st.markdown("#### 📜 Portal Forms")
        st.caption("Share these public Google Forms to onboard new sellers and buyers — submissions feed into your pipeline.")
        _pf1, _pf2 = st.columns(2)
        with _pf1:
            st.markdown(
                '<div style="background:#241a0d;border-left:3px solid #F39C12;padding:10px 14px;border-radius:0 6px 6px 0">'
                '<div style="font-weight:700;color:#F39C12;font-size:0.9rem;margin-bottom:4px">📦 SELLER FORM</div>'
                '<div style="font-size:0.78rem;color:#aab;margin-bottom:8px">For business owners who want to sell. Captures financials, industry, ask price.</div>'
                '<a href="https://forms.gle/pVNQwEBsnWJQMymm7" target="_blank" '
                'style="display:inline-block;background:#F39C12;color:#0d0d0d;padding:6px 14px;'
                'border-radius:6px;text-decoration:none;font-weight:700;font-size:0.82rem">'
                '🔗 Open Seller Form</a>'
                '</div>',
                unsafe_allow_html=True
            )
            st.code("https://forms.gle/pVNQwEBsnWJQMymm7", language=None)
        with _pf2:
            st.markdown(
                '<div style="background:#1a0d24;border-left:3px solid #9B59B6;padding:10px 14px;border-radius:0 6px 6px 0">'
                '<div style="font-weight:700;color:#d4b3e8;font-size:0.9rem;margin-bottom:4px">👑 BUYER FORM</div>'
                '<div style="font-size:0.78rem;color:#aab;margin-bottom:8px">For PE / family office / individuals who want to buy. Captures mandate, EBITDA range, geo.</div>'
                '<a href="https://forms.gle/Ka7VuFtryfkmvpp48" target="_blank" '
                'style="display:inline-block;background:#9B59B6;color:#fff;padding:6px 14px;'
                'border-radius:6px;text-decoration:none;font-weight:700;font-size:0.82rem">'
                '🔗 Open Buyer Form</a>'
                '</div>',
                unsafe_allow_html=True
            )
            st.code("https://forms.gle/Ka7VuFtryfkmvpp48", language=None)

        st.caption(
            "💡 To wire form submissions directly into Mithril, set up a Google Sheets sync: "
            "1) Link both forms to one Google Sheet, 2) share the Sheet ID in the GOOGLE_SHEET_ID env var, "
            "3) Mithril will auto-poll for new entries."
        )

        st.markdown("---")
        st.markdown("#### Export Deal Flow")
        all_deals_export = get_all_deals()
        if all_deals_export:
            df_export = pd.DataFrame(all_deals_export)
            st.download_button(
                "Export All Deals to CSV",
                data=df_export.to_csv(index=False),
                file_name=f"valar_dealflow_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — LIVE SCRAPE  (results stream in as each source completes)
# ══════════════════════════════════════════════════════════════════════════════

with tab6:
    st.subheader("🚀 Live Lead Scrape")
    st.caption(
        "Pick a buyer · click Start · leads pop up as each source finishes (Yelp first — usually 2–3s). "
        "Each card shows website + phone + LinkedIn search + 1-click Add to Pipeline. "
        "Switch to the 📋 Deal Flow Stream tab anytime to browse your existing 2,000+ deals."
    )

    _all_buyers_s6 = get_all_buyers()
    if not _all_buyers_s6:
        st.warning("No buyers configured. Add buyers in the Buyers & Settings tab first.")
        st.stop()

    _buyer_labels_s6 = {
        f"{b['name']}  ·  {', '.join((b.get('industries') or [])[:3])}": b
        for b in _all_buyers_s6
    }
    _s6_c1, _s6_c2, _s6_c3 = st.columns([4, 2, 2])
    _sel_label   = _s6_c1.selectbox("Buyer to scrape for", list(_buyer_labels_s6.keys()), key="s6_buyer")
    _s6_buyer    = _buyer_labels_s6[_sel_label]
    _s6_max      = _s6_c2.slider("Max per source", 5, 40, 15, key="s6_max")
    _s6_minscore = _s6_c3.slider("Min score", 0, 70, 25, key="s6_minscore")

    _s6_start = st.button(f"🚀 Start Scrape for {_s6_buyer['name']}", type="primary", key="s6_start", use_container_width=True)

    # ── shared rich card renderer (used by live + persisted branches) ──
    def _s6_build_card(r):
        _sn    = _e(r.get("company_name",""))
        _sind  = _e(r.get("industry",""))
        _city  = r.get("city","")
        _state = r.get("state","")
        _sloc  = _e(", ".join(filter(None, [_city, _state])))
        _ssrc  = _e(r.get("source",""))
        _sdom  = (r.get("company_domain") or "").strip()
        _sphn  = (r.get("phone") or r.get("owner_phone") or "").strip()
        _semail= (r.get("owner_email") or "").strip()
        _owner = (r.get("owner_name") or "").strip()
        _ssc   = r.get("match_score", 0)
        _scc   = "#27AE60" if _ssc >= 70 else "#F39C12" if _ssc >= 45 else "#4A90D9"

        _href = _sdom if _sdom.startswith("http") else ("https://" + _sdom if _sdom else "")
        _is_dir = any(d in _sdom.lower() for d in ("yelp.com","manta.com","yellowpages.com","bbb.org"))
        if _href and not _is_dir:
            _web_h = (f'<a href="{_e(_href)}" target="_blank" '
                      f'style="color:#27AE60;font-size:0.78rem;font-weight:600">🌐 Website</a>')
        elif _href:
            _dn = _sdom.split(".")[0].title() if _sdom else "Listing"
            _web_h = (f'<a href="{_e(_href)}" target="_blank" '
                      f'style="color:#FF9F40;font-size:0.78rem">📋 {_e(_dn)} listing</a>')
        else:
            _web_h = '<span style="color:#3a3a3a;font-size:0.76rem">— no website —</span>'

        _phn_h = (f'<a href="tel:{_e(_sphn)}" style="color:#4A90D9;font-size:0.78rem">📞 {_e(_sphn)}</a>'
                  if _sphn else '<span style="color:#3a3a3a;font-size:0.76rem">no phone</span>')

        _eml_h = (f'<a href="mailto:{_e(_semail)}" style="color:#9B59B6;font-size:0.78rem">✉ {_e(_semail)}</a>'
                  if _semail else "")

        _li_co  = _urlquote(f'{r.get("company_name","")} {_state}')
        _li_h   = (f'<a href="https://www.linkedin.com/search/results/companies/?keywords={_li_co}" '
                   f'target="_blank" style="color:#0077B5;font-size:0.74rem">💼 Co</a>')
        if _owner:
            _li_p = _urlquote(f'{_owner} {r.get("company_name","")}')
            _li_h2 = (f' &nbsp; <a href="https://www.linkedin.com/search/results/people/?keywords={_li_p}" '
                      f'target="_blank" style="color:#0077B5;font-size:0.74rem">👤 {_e(_owner)}</a>')
        else:
            _li_p = _urlquote(f'owner CEO founder {r.get("company_name","")}')
            _li_h2 = (f' &nbsp; <a href="https://www.linkedin.com/search/results/people/?keywords={_li_p}" '
                      f'target="_blank" style="color:#0077B5;font-size:0.74rem">👤 Find Owner</a>')

        _gq = _urlquote(f'"{r.get("company_name","")}" {_city} {_state} owner CEO founder')
        _g_h = (f'<a href="https://www.google.com/search?q={_gq}" target="_blank" '
                f'style="color:#888;font-size:0.74rem">🔍 Google</a>')

        return (
            f'<div style="border-left:3px solid {_scc};padding:10px 14px;margin:4px 0;'
            f'background:#0a0c10;border-radius:0 8px 8px 0;border:1px solid {_scc}22">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">'
            f'<strong style="color:#f0f0f0;font-size:0.95rem">{_sn}</strong>'
            f'<span style="background:{_scc}22;color:{_scc};font-size:0.72rem;'
            f'padding:2px 9px;border-radius:10px;white-space:nowrap">{_ssc}/100</span>'
            f'</div>'
            f'<div style="color:#666;font-size:0.74rem;margin-bottom:6px">{_sind}  ·  {_sloc}  ·  via {_ssrc}</div>'
            f'<div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:3px">'
            f'{_web_h} &nbsp;·&nbsp; {_phn_h}' + (f' &nbsp;·&nbsp; {_eml_h}' if _eml_h else "") + f'</div>'
            f'<div style="display:flex;gap:8px;flex-wrap:wrap;font-size:0.74rem;color:#666">'
            f'{_li_h}{_li_h2} &nbsp;·&nbsp; {_g_h}'
            f'</div>'
            f'</div>'
        )

    if _s6_start:
        st.session_state["s6_results"] = []

    if _s6_start or st.session_state.get("_s6_running"):
        st.session_state["_s6_running"] = True

        try:
            import web_scraper as _ws
        except ImportError:
            st.error("web_scraper module not found in tools/")
            st.session_state.pop("_s6_running", None)
            st.stop()

        _s6_industries = (_s6_buyer.get("industries") or [])[:2]
        _s6_states_raw = list((_s6_buyer.get("states") or {}).keys()) if isinstance(_s6_buyer.get("states"), dict) else []
        _s6_states = _s6_states_raw[:4] or ["FL", "TX", "CA", "GA", "NC"]

        # Build per-source task list — FASTEST FIRST so user sees results immediately
        _s6_tasks = []
        for _ind in _s6_industries:
            _s6_tasks += [
                # Yelp API — usually 2-3s, returns 10-20 leads with phone + url
                (f"⚡ Yelp — {_ind} ({_s6_states[0]})",
                 lambda i=_ind, st0=_s6_states[0]: _ws.scrape_yelp(i, st0, _s6_max)),
                # YP — fairly quick HTML scrape
                (f"⚡ Yellow Pages — {_ind}",
                 lambda i=_ind, loc=_s6_states[0]: _ws.scrape_yellow_pages(f"{i} company", loc)),
                # Manta — directory listings
                (f"Manta — {_ind} ({_s6_states[0]})",
                 lambda i=_ind, st0=_s6_states[0]: _ws.scrape_manta(i, st0)),
                # BizBuySell — slower (Playwright), but quality listings
                (f"BizBuySell — {_ind}",
                 lambda i=_ind: _ws.scrape_bizbuysell(i, _s6_states, 0, 0, _s6_max)),
                # BizQuest — similar to BizBuySell
                (f"BizQuest — {_ind}",
                 lambda i=_ind: _ws.scrape_bizquest(i, _s6_states, _s6_max)),
                # Craigslist
                (f"Craigslist — {_ind}",
                 lambda i=_ind: _ws.scrape_craigslist(i, _s6_states, _s6_max)),
            ]

        _s6_progress  = st.progress(0.0)
        _s6_status    = st.empty()
        _s6_results   = st.session_state.get("s6_results", [])
        _s6_seen      = {r.get("company_name","").lower() for r in _s6_results}
        _s6_box       = st.empty()

        def _s6_render(results):
            filtered = [
                r for r in results
                if (r.get("company_domain") or r.get("phone"))
                   and r.get("match_score", 0) >= _s6_minscore
                   and (r.get("company_name") or "").strip()
            ]
            with _s6_box.container():
                if not filtered:
                    st.caption("⏳ No qualifying leads yet — scraping…")
                    return
                st.markdown(f"### 🎯 {len(filtered)} qualifying leads found")
                for r in filtered[:80]:
                    st.markdown(_s6_build_card(r), unsafe_allow_html=True)

        for _idx, (_tname, _tfn) in enumerate(_s6_tasks):
            _s6_status.markdown(f"⏳ **{_tname}**…")
            try:
                _new = _tfn()
                for _lead in _new:
                    _lkey = (_lead.get("company_name") or "").lower().strip()
                    if _lkey and _lkey not in _s6_seen:
                        _s6_seen.add(_lkey)
                        _lead["match_score"] = score_lead(_lead, _s6_buyer)
                        _lead["buyer_id"]    = _s6_buyer.get("id", "")
                        _lead["buyer_name"]  = _s6_buyer.get("name", "")
                        _s6_results.append(_lead)
            except Exception as _ex:
                st.toast(f"⚠ {_tname}: {_ex}", icon="⚠")

            _s6_progress.progress((_idx + 1) / len(_s6_tasks))
            _s6_render(_s6_results)

        st.session_state["s6_results"]  = _s6_results
        st.session_state["_s6_running"] = False
        _s6_status.markdown(f"✅ **Done!** {len(_s6_results)} total leads scraped.")
        _s6_progress.progress(1.0)

        # Save to pipeline button
        _qualifying = [
            r for r in _s6_results
            if (r.get("company_domain") or r.get("phone"))
               and r.get("match_score", 0) >= _s6_minscore
        ]
        if _qualifying:
            if st.button(f"💾 Save {len(_qualifying)} qualifying leads to Pipeline", key="s6_save"):
                _saved_s6 = 0
                for _ld in _qualifying:
                    if not deal_exists(_ld.get("company_name",""), _ld.get("buyer_id","")):
                        add_deal(_ld)
                        _saved_s6 += 1
                st.success(f"Saved {_saved_s6} new leads to Deal Flow Stream.")
                st.rerun()

    elif st.session_state.get("s6_results"):
        # Show persisted results from previous scrape
        _prev = st.session_state["s6_results"]
        _filt = [
            r for r in _prev
            if (r.get("company_domain") or r.get("phone"))
               and r.get("match_score", 0) >= _s6_minscore
               and (r.get("company_name") or "").strip()
        ]
        st.info(f"Showing **{len(_filt)} leads** from last scrape for **{_s6_buyer['name']}**. Adjust min score above or click Start Scrape to refresh.")

        # Bulk save button
        if _filt and st.button(f"💾 Save All {len(_filt)} to Pipeline", key="s6_save_persisted", type="primary"):
            _saved = 0
            for _ld in _filt:
                if not deal_exists(_ld.get("company_name",""), _ld.get("buyer_id","")):
                    add_deal(_ld); _saved += 1
            st.success(f"Saved {_saved} new leads.")
            st.rerun()

        for r in _filt[:80]:
            st.markdown(_s6_build_card(r), unsafe_allow_html=True)
    else:
        st.info("👆 Select a buyer and click **Start Scrape** to begin. Yelp results appear in 2-3 seconds.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — MANUAL TEASERS  (BIZBROKER/TEASERS matched to buyers)
# ══════════════════════════════════════════════════════════════════════════════

with tab7:
    st.markdown(
        '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">'
        '<span style="font-size:1.4rem;font-weight:800;color:#e8eaf0">⚒ The Forge</span>'
        '<span style="color:#666;font-size:0.82rem">Pick a seller on the <span style="color:#F39C12">left</span>, '
        'see best-matched buyers gravitate to the <span style="color:#9B59B6">right</span>. '
        'When you find a fit, forge the match into the Pipeline.</span>'
        '</div>',
        unsafe_allow_html=True
    )

    # Load all sellers (deals without buyer_id) and buyers
    _forge_all_deals  = get_all_deals()
    _forge_sellers    = [d for d in _forge_all_deals if (d.get("company_name") or "").strip()
                         and (d.get("company_name") or "").strip() not in ("—","-","Unknown","None")]
    _forge_buyers     = get_all_buyers()

    # Optional: also pull live teaser parses
    if st.button("🔄 Refresh from BIZBROKER/TEASERS", key="forge_refresh"):
        with st.spinner("Re-scanning TEASERS folder…"):
            try:
                import teaser_parser as _tp
                _t_deals = _tp.scan_all_teasers()
                _t_matches = _tp.match_teasers(_t_deals)
                st.session_state["forge_teaser_matches"] = _t_matches
                st.success(f"Loaded {len(_t_deals)} teaser deals · {len(_t_matches)} buyer matches")
            except Exception as _ex:
                st.error(f"Could not load: {_ex}")

    # Filter row
    _ff1, _ff2, _ff3 = st.columns([2, 2, 3])
    _seller_filter = _ff1.selectbox("Filter sellers by", ["All","Off-Market","On-Market","Teaser"], key="forge_seller_type")
    _min_match     = _ff2.selectbox("Min match %", ["Any","50%+","70%+","90%+"], key="forge_min_match", index=1)
    _search_forge  = _ff3.text_input("🔎 Search company / industry / state", key="forge_search")

    # Apply seller filter
    _forge_filtered_sellers = list(_forge_sellers)
    if _seller_filter == "Off-Market":
        _forge_filtered_sellers = [s for s in _forge_filtered_sellers if s.get("listing_type") != "on-market"]
    elif _seller_filter == "On-Market":
        _forge_filtered_sellers = [s for s in _forge_filtered_sellers if s.get("listing_type") == "on-market"]
    elif _seller_filter == "Teaser":
        _forge_filtered_sellers = [s for s in _forge_filtered_sellers if (s.get("source") or "").startswith("teaser:")]

    if _search_forge and _search_forge.strip():
        _q = _search_forge.lower().strip()
        _forge_filtered_sellers = [s for s in _forge_filtered_sellers
                                   if _q in (s.get("company_name") or "").lower()
                                   or _q in (s.get("industry") or "").lower()
                                   or _q in (s.get("state") or "").lower()
                                   or _q in (s.get("city") or "").lower()]

    _min_pct = {"Any":0, "50%+":50, "70%+":70, "90%+":90}[_min_match]

    # Sort sellers by score desc to show best at top
    _forge_filtered_sellers.sort(key=lambda x: x.get("match_score", 0), reverse=True)

    # Selected seller from session
    _sel_seller_id = st.session_state.get("forge_selected_seller_id")
    _sel_seller    = next((s for s in _forge_filtered_sellers if s.get("id") == _sel_seller_id), None)
    if not _sel_seller and _forge_filtered_sellers:
        _sel_seller = _forge_filtered_sellers[0]
        st.session_state["forge_selected_seller_id"] = _sel_seller.get("id")

    # ── two-column playground ──────────────────────────────────────────────
    _fc_left, _fc_right = st.columns(2)

    # LEFT: Sellers
    with _fc_left:
        st.markdown(
            f'<div style="background:#241a0d;border-left:3px solid #F39C12;padding:6px 12px;'
            f'border-radius:0 6px 6px 0;margin-bottom:6px">'
            f'<span style="color:#F39C12;font-weight:800;font-size:0.95rem">📦 SELLERS</span> '
            f'<span style="color:#666;font-size:0.78rem">· {len(_forge_filtered_sellers)} listings — click one to match</span>'
            f'</div>',
            unsafe_allow_html=True
        )
        # Render top 30 sellers as compact pickable cards
        for _s in _forge_filtered_sellers[:30]:
            _sname = (_s.get("company_name") or "").strip()
            if not _sname or _sname in ("—","Unknown"):
                continue
            _sid    = _s.get("id")
            _sind   = (_s.get("industry") or "").strip()[:50]
            _sloc   = ", ".join(filter(None, [_s.get("city",""), _s.get("state","")]))
            _srev   = (_s.get("revenue_estimate") or "").strip()
            _sebit  = (_s.get("ebitda_estimate") or "").strip()
            _ssc    = _s.get("match_score", 0)
            _is_selected = _sid == _sel_seller_id
            _border = "#F39C12" if _is_selected else "#1a1d24"
            _bg     = "#1a1308" if _is_selected else "#0a0c10"

            _scol1, _scol2 = st.columns([8, 2])
            with _scol1:
                st.markdown(
                    f'<div style="border-left:2px solid {_border};padding:6px 10px;margin:2px 0;'
                    f'background:{_bg};border-radius:0 4px 4px 0">'
                    f'<div style="font-weight:700;color:#f0f0f0;font-size:0.86rem">{_e(_sname)}</div>'
                    f'<div style="color:#666;font-size:0.72rem">{_e(_sind)} · {_e(_sloc)}</div>'
                    + (f'<div style="color:#888;font-size:0.7rem">Rev {_e(_srev)} · EBITDA {_e(_sebit)}</div>' if _srev or _sebit else "")
                    + '</div>',
                    unsafe_allow_html=True
                )
            with _scol2:
                if st.button("Pick", key=f"fpick_s_{_sid}", use_container_width=True,
                             type="primary" if _is_selected else "secondary"):
                    st.session_state["forge_selected_seller_id"] = _sid
                    st.rerun()

    # RIGHT: Buyers, ranked by combined Grade (match × capacity × recency)
    with _fc_right:
        if not _sel_seller:
            st.info("👈 Pick a seller on the left to see matching buyers.")
        else:
            # ── Buyer Grade = match (60) + capacity (30) + recency (10) ──
            _all_deals_for_capacity = get_all_deals()
            def _buyer_grade(buyer):
                """0-100. Match quality + capacity + recency."""
                fit = score_lead(_sel_seller, buyer)  # 0-100 strict
                match_pts = round(fit * 0.60)        # max 60

                # Capacity: active (not dead/closed/passed) deals owned by this buyer
                _bid = buyer.get("id","")
                _active = [d for d in _all_deals_for_capacity
                           if d.get("buyer_id") == _bid
                           and d.get("deal_stage") not in ("dead","closed","passed")]
                _n = len(_active)
                if   _n == 0: cap_pts = 30
                elif _n <= 2: cap_pts = 25
                elif _n <= 5: cap_pts = 18
                elif _n <= 10: cap_pts = 10
                else:         cap_pts = 0

                # Recency: when was buyer last contacted on any deal?
                from datetime import datetime as _dt
                _last_contacts = [d.get("last_contacted","") for d in _all_deals_for_capacity
                                  if d.get("buyer_id") == _bid and d.get("last_contacted")]
                _days = 9999
                if _last_contacts:
                    _last = max(_last_contacts)
                    try:
                        _dtv = _dt.fromisoformat(_last.split("+")[0].split("Z")[0])
                        _days = (_dt.now() - _dtv).days
                    except Exception:
                        _days = 9999
                if   _days >= 30: rec_pts = 10  # cool, capacity available
                elif _days >= 14: rec_pts = 6
                elif _days >= 7:  rec_pts = 3
                else:             rec_pts = 0   # contacted very recently

                grade = min(100, match_pts + cap_pts + rec_pts)
                return {
                    "grade":     grade,
                    "match":     fit,
                    "match_pts": match_pts,
                    "cap_pts":   cap_pts,
                    "rec_pts":   rec_pts,
                    "active":    _n,
                    "days":      _days,
                }

            _ranked_buyers = [(_buyer_grade(b), b) for b in _forge_buyers]
            _ranked_buyers.sort(key=lambda x: x[0]["grade"], reverse=True)
            _ranked_buyers = [(g,b) for g,b in _ranked_buyers if g["match"] >= _min_pct]

            # Show 100% matches count
            _perfect_n = sum(1 for g,_ in _ranked_buyers if g["grade"] >= 90)
            st.markdown(
                f'<div style="background:#1a0d24;border-left:3px solid #9B59B6;padding:6px 12px;'
                f'border-radius:0 6px 6px 0;margin-bottom:6px">'
                f'<span style="color:#d4b3e8;font-weight:800;font-size:0.95rem">👑 BUYERS</span> '
                f'<span style="color:#666;font-size:0.78rem">· '
                + (f'<strong style="color:#F1C40F">{_perfect_n} 100% match(es)</strong> · ' if _perfect_n else "")
                + f'ranked by Grade vs <strong style="color:#F39C12">{_e(_sel_seller.get("company_name",""))}</strong>'
                + '</span></div>'
                '<div style="color:#444;font-size:0.68rem;margin-bottom:6px">'
                'Grade = match (60) + capacity (30, fewer active deals = more) + recency (10, longer since contact = more)'
                '</div>',
                unsafe_allow_html=True
            )

            for _g, _b in _ranked_buyers[:30]:
                _bscore  = _g["grade"]
                _match   = _g["match"]
                _bname   = _b.get("name","") or _b.get("firm","")
                _binds   = ", ".join((_b.get("industries") or [])[:3])
                _bstates = ", ".join((list(_b.get("states") or {}) if isinstance(_b.get("states"), dict) else (_b.get("states") or []))[:3])
                _b_email = _b.get("email","")
                _b_phone = _b.get("phone","")
                _b_web   = _b.get("website","")

                # Color tiers based on Grade
                _bclr = "#F1C40F" if _bscore >= 90 else "#27AE60" if _bscore >= 75 else "#F39C12" if _bscore >= 55 else "#888"
                _is_perfect = _bscore >= 90

                # Capacity + recency badges
                _cap_label = f"📊 {_g['active']} active" if _g['active'] else "📊 0 active (ready)"
                if _g["days"] >= 9999:
                    _rec_label = "⏳ never contacted"
                elif _g["days"] >= 30:
                    _rec_label = f"⏳ {_g['days']}d cool"
                elif _g["days"] >= 14:
                    _rec_label = f"⏳ {_g['days']}d ago"
                else:
                    _rec_label = f"🔥 {_g['days']}d ago (recent)"

                _perfect_html = ' <span style="color:#F1C40F;font-size:0.7rem;font-weight:800;background:#3a2a08;padding:1px 6px;border-radius:4px">🎯 100%</span>' if _is_perfect else ""

                _cap_pts_v = _g["cap_pts"]
                _rec_pts_v = _g["rec_pts"]
                _grade_tooltip = f"match {_match} + cap {_cap_pts_v} + rec {_rec_pts_v}"

                _bcol1, _bcol2 = st.columns([8, 2])
                with _bcol1:
                    st.markdown(
                        f'<div style="border-left:3px solid {_bclr};padding:7px 11px;margin:2px 0;'
                        f'background:#0d0a14;border-radius:0 4px 4px 0">'
                        f'<div style="display:flex;justify-content:space-between;align-items:center">'
                        f'<span style="font-weight:700;color:#f0f0f0;font-size:0.86rem">{_e(_bname)}{_perfect_html}</span>'
                        f'<span style="background:{_bclr}22;color:{_bclr};font-size:0.7rem;'
                        f'padding:2px 8px;border-radius:8px;font-weight:800" title="{_grade_tooltip}">Grade {_bscore}</span>'
                        f'</div>'
                        f'<div style="color:#666;font-size:0.72rem">{_e(_binds)}</div>'
                        + (f'<div style="color:#888;font-size:0.7rem">{_e(_bstates)}</div>' if _bstates else "")
                        + f'<div style="display:flex;gap:8px;font-size:0.66rem;color:#666;margin-top:3px">'
                        + f'<span title="Active deals already on this buyer">{_cap_label}</span>'
                        + f'<span title="Last time we contacted them">{_rec_label}</span>'
                        + f'<span style="color:#444">· fit {_match}</span>'
                        + '</div>'
                        + '<div style="font-size:0.7rem;margin-top:3px">'
                        + (f'<a href="mailto:{_e(_b_email)}" style="color:#4A90D9">✉</a> ' if _b_email else "")
                        + (f'<a href="tel:{_e(_b_phone)}" style="color:#4A90D9">📞</a> ' if _b_phone else "")
                        + (f'<a href="{_e(_b_web)}" target="_blank" style="color:#27AE60">🌐</a>' if _b_web else "")
                        + '</div>'
                        + '</div>',
                        unsafe_allow_html=True
                    )
                with _bcol2:
                    _btn_label = "🎯 Forge" if _is_perfect else "➕ Add"
                    if st.button(_btn_label, key=f"fmatch_{_sel_seller.get('id')}_{_b.get('id','')}", use_container_width=True,
                                 type="primary" if _is_perfect else "secondary"):
                        # Add match to pipeline — refuse duplicates
                        _proj_name = _sel_seller.get("company_name","")
                        if deal_exists(_proj_name, _b.get("id","")):
                            st.info(f"⚒ Already forged: {_proj_name} × {_bname}")
                        else:
                            _new_match = {
                                "company_name":     _proj_name,
                                "industry":         _sel_seller.get("industry",""),
                                "state":            _sel_seller.get("state",""),
                                "city":             _sel_seller.get("city",""),
                                "revenue_estimate": _sel_seller.get("revenue_estimate",""),
                                "ebitda_estimate":  _sel_seller.get("ebitda_estimate",""),
                                "asking_price":     _sel_seller.get("asking_price",""),
                                "owner_email":      _sel_seller.get("owner_email",""),
                                "owner_name":       _sel_seller.get("owner_name",""),
                                "owner_phone":      _sel_seller.get("owner_phone",""),
                                "company_domain":   _sel_seller.get("company_domain",""),
                                "buyer_id":         _b.get("id",""),
                                "buyer_name":       _bname,
                                "match_score":      _bscore,
                                "interest_level":   "hot" if _bscore >= 90 else "warm" if _bscore >= 60 else "cold",
                                "listing_type":     _sel_seller.get("listing_type","off-market"),
                                "deal_stage":       "identified",
                                "status":           "identified",
                                "source":           f"forge:{_sel_seller.get('source','')}",
                            }
                            add_deal(_new_match)
                            st.success(f"⚒ Forged {_proj_name} × {_bname} into the Pipeline!")
                            st.rerun()
