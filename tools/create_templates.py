"""
create_templates.py
-------------------
Creates two Excel templates in the data/ directory:

  data/BUYERS_TEMPLATE.xlsx   — Fill in one row per buyer
  data/SELLERS_TEMPLATE.xlsx  — Fill in one row per seller deal (manual entry)

Run once to create them, or re-run to regenerate clean copies.

Usage:
    python tools/create_templates.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # Dark blue
SECTION_FILL  = PatternFill("solid", fgColor="2E75B6")   # Mid blue
REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")   # Soft orange — required cells
OPTIONAL_FILL = PatternFill("solid", fgColor="EBF3FB")   # Light blue — optional cells
EXAMPLE_FILL  = PatternFill("solid", fgColor="E2EFDA")   # Light green — example row
LOCKED_FILL   = PatternFill("solid", fgColor="F2F2F2")   # Gray — instructions

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT  = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT    = Font(bold=True, size=14, color="1F4E79")
NORMAL_FONT   = Font(size=10)
EXAMPLE_FONT  = Font(size=10, italic=True, color="375623")

THIN_BORDER   = Border(
    left=Side(style="thin"),   right=Side(style="thin"),
    top=Side(style="thin"),    bottom=Side(style="thin"),
)


def style_header(cell, text):
    cell.value = text
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_cell(cell, fill=None, bold=False, wrap=False):
    cell.fill = fill or OPTIONAL_FILL
    cell.font = Font(size=10, bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = THIN_BORDER


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# BUYERS TEMPLATE
# ---------------------------------------------------------------------------

BUYER_COLUMNS = [
    # (header, width, required, example, notes)
    ("Buyer ID",           14, False, "buyer_001",         "Auto-assigned if blank (buyer_001, buyer_002…)"),
    ("Full Name",          22, True,  "Michael Johnson",   "First and last name"),
    ("Company / Entity",   25, False, "MJ Capital LLC",    "LLC, Inc, or individual name"),
    ("Email",              28, True,  "mjohnson@email.com","Primary contact email"),
    ("Phone",              18, False, "305-555-0100",      "Cell or direct line"),
    ("Agreement Date",     18, False, "2025-11-01",        "Date buyer-broker agreement signed (YYYY-MM-DD)"),
    # Criteria
    ("Industry Preferences", 35, True,  "HVAC; Plumbing; Home Services",
     "Separate multiple with semicolons. E.g. HVAC; Landscaping; Pest Control"),
    ("Industry Exclusions",  30, False, "Restaurant; Retail; Franchise",
     "Separate with semicolons. These are HARD disqualifiers."),
    ("Target States",        20, True,  "FL; TX; GA; NC; SC",
     "2-letter codes, semicolon-separated. Leave blank = any state."),
    # Financials
    ("Min Asking Price ($)", 20, False, "500000",  "Integers only. E.g. 500000 = $500K"),
    ("Max Asking Price ($)", 20, True,  "2000000", "Integers only. E.g. 2000000 = $2M"),
    ("Min Revenue ($)",      18, False, "750000",  "Annual revenue minimum"),
    ("Max Revenue ($)",      18, False, "5000000", "Annual revenue maximum"),
    ("Min Cash Flow / SDE ($)", 22, False, "150000", "Min SDE or EBITDA"),
    ("Max CF Multiple",      18, False, "5.0",     "Max price/cash-flow multiple. E.g. 5.0 = 5x"),
    # Deal structure
    ("SBA Preferred?",       16, False, "Yes",     "Yes / No"),
    ("Seller Financing OK?", 20, False, "Yes",     "Yes / No"),
    ("All Cash?",            14, False, "No",      "Yes / No — buyer can do all-cash"),
    # Business attributes
    ("Min Employees",        16, False, "5",       "Minimum headcount"),
    ("Max Employees",        16, False, "50",      "Maximum headcount"),
    ("Min Years in Business", 20, False, "3",      "Minimum years established"),
    ("Recurring Revenue Preferred?", 24, False, "Yes", "Yes / No"),
    # Scoring weights (must sum to 100)
    ("Weight: Industry (default 30)",    24, False, "30", "Score weight 0-100. All 6 weights must sum to 100."),
    ("Weight: Geography (default 20)",   24, False, "20", ""),
    ("Weight: Financials (default 25)",  24, False, "25", ""),
    ("Weight: CF Multiple (default 10)", 25, False, "10", ""),
    ("Weight: Years Est. (default 8)",   24, False, "8",  ""),
    ("Weight: Deal Structure (default 7)", 26, False, "7", ""),
    # Notes
    ("Buyer Summary",  45, False,
     "Experienced operator, Southeast focus. SBA pre-approved to $2M.",
     "1-2 sentence profile. Used in email generation."),
    ("Notes / Financing Details", 45, False,
     "$400K liquid. Pre-qualified SBA. Close in 90 days.",
     "Internal notes. Also used to personalize outreach emails."),
]


def create_buyers_template(out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyers"

    # --- Title row ---
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:AD1")
    title_cell = ws["A1"]
    title_cell.value = "BUYER CRITERIA — Deal Flow & Matching App"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Instructions row ---
    ws.row_dimensions[2].height = 45
    ws.merge_cells("A2:AD2")
    inst = ws["A2"]
    inst.value = (
        "HOW TO USE: Fill in one row per buyer. Required fields are orange. "
        "Use semicolons to separate multiple values (e.g. 'FL; TX; GA'). "
        "Scoring weights must sum to exactly 100 per buyer. "
        "Save this file and run: python tools/import_buyers_excel.py"
    )
    inst.fill = LOCKED_FILL
    inst.font = Font(size=10, italic=True, color="595959")
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Column headers ---
    ws.row_dimensions[3].height = 50
    for col_idx, (header, width, required, _, _) in enumerate(BUYER_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        style_header(cell, header)
        if required:
            cell.fill = PatternFill("solid", fgColor="C00000")  # Red = required
        set_col_width(ws, get_column_letter(col_idx), width)

    # --- Notes row (row 4) ---
    ws.row_dimensions[4].height = 40
    for col_idx, (_, _, required, _, notes) in enumerate(BUYER_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = notes
        cell.fill = LOCKED_FILL
        cell.font = Font(size=8, italic=True, color="595959")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER

    # --- Example row (row 5) ---
    ws.row_dimensions[5].height = 22
    for col_idx, (_, _, required, example, _) in enumerate(BUYER_COLUMNS, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = example
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        cell.border = THIN_BORDER

    # --- Data rows (6 onward — pre-fill 20 empty rows) ---
    for row in range(6, 26):
        ws.row_dimensions[row].height = 20
        for col_idx, (_, _, required, _, _) in enumerate(BUYER_COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="top")
            cell.border = THIN_BORDER

    # --- Yes/No dropdowns for boolean columns ---
    yes_no_cols = [16, 17, 18, 22]  # SBA, Seller Fin, All Cash, Recurring
    for col_idx in yes_no_cols:
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}6:{col_letter}100")

    # --- Freeze panes: keep headers visible while scrolling ---
    ws.freeze_panes = "A6"

    # --- Legend tab ---
    legend = wb.create_sheet("Legend & Instructions")
    legend["A1"] = "COLOR LEGEND"
    legend["A1"].font = Font(bold=True, size=12)
    rows = [
        ("Red header", "C00000", "FFFFFF", "Required field — must be filled in"),
        ("Blue header", "1F4E79", "FFFFFF", "Optional field"),
        ("Orange cell", "FCE4D6", "000000", "Required data cell"),
        ("Light blue cell", "EBF3FB", "000000", "Optional data cell"),
        ("Green row", "E2EFDA", "375623", "Example data (row 5) — replace with your data"),
        ("Gray cell", "F2F2F2", "595959", "Instructions / notes — do not edit"),
    ]
    for i, (label, bg, fg, desc) in enumerate(rows, start=3):
        legend.cell(row=i, column=1).value = label
        legend.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
        legend.cell(row=i, column=1).font = Font(color=fg, bold=True)
        legend.cell(row=i, column=2).value = desc
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 55

    legend["A10"] = "SCORING WEIGHTS"
    legend["A10"].font = Font(bold=True, size=12)
    legend["A11"] = "The 6 weight columns (columns W-AB) must sum to exactly 100 per buyer."
    legend["A12"] = "Default weights: Industry=30, Geography=20, Financials=25, CF Multiple=10, Years=8, Structure=7"
    legend["A13"] = "Adjust to reflect each buyer's priorities. Example: a PE firm might weight CF Multiple higher."

    legend["A15"] = "RUNNING THE IMPORT"
    legend["A15"].font = Font(bold=True, size=12)
    legend["A16"] = "1. Fill in your buyer rows (start from row 6)"
    legend["A17"] = "2. Save this file"
    legend["A18"] = "3. Run: python tools/import_buyers_excel.py"
    legend["A19"] = "4. Check .tmp/buyers.json — verify the data looks correct"
    legend["A20"] = "5. Run the pipeline: python tools/run_pipeline.py"

    wb.save(str(out_path))
    print(f"[OK] Buyers template: {out_path}")


# ---------------------------------------------------------------------------
# SELLERS / LISTINGS TEMPLATE
# ---------------------------------------------------------------------------

SELLER_COLUMNS = [
    ("Deal ID",            14, False, "manual_001",          "Leave blank to auto-generate"),
    ("Business Title",     40, True,  "HVAC Company - Tampa FL", "Name or description from listing"),
    ("Source",             16, False, "manual",              "Where you found it: manual / bizbuysell / referral / etc."),
    ("Listing URL",        40, False, "https://...",         "Direct link to listing if available"),
    ("Industry",           25, True,  "HVAC Services",       "Be specific: HVAC / Plumbing / Landscaping / etc."),
    ("City",               20, False, "Tampa",               ""),
    ("State",              10, True,  "FL",                  "2-letter state code"),
    ("Asking Price ($)",   20, True,  "850000",              "Integer. E.g. 850000 = $850K"),
    ("Revenue ($)",        18, False, "1200000",             "Annual revenue"),
    ("Cash Flow / SDE ($)", 20, False, "220000",             "SDE or cash flow (annual)"),
    ("EBITDA ($)",         16, False, "",                    "If different from SDE"),
    ("Employees",          14, False, "12",                  "Headcount"),
    ("Years in Business",  20, False, "15",                  "Years established"),
    ("Reason for Selling", 30, False, "Retirement",          "Owner's stated reason"),
    ("SBA Eligible?",      16, False, "Yes",                 "Yes / No / Unknown"),
    ("Seller Financing?",  18, False, "Yes",                 "Yes / No / Unknown"),
    ("Real Estate Included?", 20, False, "No",              "Yes / No / Unknown"),
    ("Inventory Value ($)", 18, False, "45000",              "If inventory included"),
    ("FF&E Value ($)",     16, False, "120000",              "Furniture, fixtures, equipment"),
    ("Broker Name",        25, False, "John Smith",          "Listing broker or direct seller"),
    ("Broker Email",       30, False, "john@broker.com",     "Contact email for outreach"),
    ("Broker Phone",       18, False, "813-555-0100",        ""),
    ("Date Listed",        16, False, "2026-01-15",          "YYYY-MM-DD"),
    ("Description",        60, False,
     "Profitable HVAC company serving Tampa Bay area for 15 years. Strong recurring maintenance contracts.",
     "Copy the listing description or add your own notes"),
    ("Notes",              40, False, "Owner motivated, wants to close in 60 days",
     "Internal notes — not shared with anyone"),
]


def create_sellers_template(out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Seller Deals"

    # Title
    ws.row_dimensions[1].height = 28
    last_col = get_column_letter(len(SELLER_COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = "SELLER DEALS — Manual Entry for Deal Flow Matching"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Instructions
    ws.row_dimensions[2].height = 40
    ws.merge_cells(f"A2:{last_col}2")
    inst = ws["A2"]
    inst.value = (
        "HOW TO USE: Add one row per business listing. Required fields are orange. "
        "These deals will be matched against all buyer criteria and email drafts will be generated. "
        "Save this file and run: python tools/import_sellers_excel.py"
    )
    inst.fill = LOCKED_FILL
    inst.font = Font(size=10, italic=True, color="595959")
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Headers
    ws.row_dimensions[3].height = 50
    for col_idx, (header, width, required, _, _) in enumerate(SELLER_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        style_header(cell, header)
        if required:
            cell.fill = PatternFill("solid", fgColor="C00000")
        set_col_width(ws, get_column_letter(col_idx), width)

    # Notes row
    ws.row_dimensions[4].height = 38
    for col_idx, (_, _, _, _, notes) in enumerate(SELLER_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = notes
        cell.fill = LOCKED_FILL
        cell.font = Font(size=8, italic=True, color="595959")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER

    # Example row
    ws.row_dimensions[5].height = 22
    for col_idx, (_, _, _, example, _) in enumerate(SELLER_COLUMNS, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = example
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(vertical="top")
        cell.border = THIN_BORDER

    # Data rows
    for row in range(6, 106):  # 100 blank rows
        ws.row_dimensions[row].height = 20
        for col_idx, (_, _, required, _, _) in enumerate(SELLER_COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="top")
            cell.border = THIN_BORDER

    # Yes/No/Unknown dropdowns for boolean cols
    bool_cols = [15, 16, 17]  # SBA, Financing, Real Estate
    for col_idx in bool_cols:
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}6:{col_letter}200")

    ws.freeze_panes = "A6"
    wb.save(str(out_path))
    print(f"[OK] Sellers template: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    data_dir = Path(__file__).parent.parent / "data"
    data_dir.mkdir(exist_ok=True)

    buyers_path = data_dir / "BUYERS_TEMPLATE.xlsx"
    sellers_path = data_dir / "SELLERS_TEMPLATE.xlsx"

    create_buyers_template(buyers_path)
    create_sellers_template(sellers_path)

    print(f"\n[DONE] Templates created in data/")
    print(f"  {buyers_path}")
    print(f"  {sellers_path}")
    print(f"\nNext steps:")
    print(f"  1. Open data/BUYERS_TEMPLATE.xlsx and fill in your buyer rows")
    print(f"  2. Open data/SELLERS_TEMPLATE.xlsx and add any deals to test")
    print(f"  3. Run: python tools/import_buyers_excel.py")
    print(f"  4. Run: python tools/import_sellers_excel.py  (optional)")
    print(f"  5. Run: python tools/run_pipeline.py --skip-scrape  (to use your manual data)")


if __name__ == "__main__":
    main()
